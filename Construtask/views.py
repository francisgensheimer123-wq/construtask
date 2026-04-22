from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
import os
import struct
import unicodedata
import zlib

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Case, Count, DecimalField, ExpressionWrapper, F, Prefetch, Q, Sum, When
from django.db.models.functions import Coalesce
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from django.http import Http404, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView

from .forms import (
    AnexoOperacionalForm,
    AditivoContratoForm,
    AditivoContratoItemFormSet,
    CompromissoForm,
    CompromissoItemFormSet,
    MedicaoForm,
    MedicaoItemFormSet,
    NotaFiscalCentroCustoFormSet,
    NotaFiscalForm,
    ObraForm,
    PlanoContasForm,
    obter_centros_da_origem_nota,
    obter_centros_do_contrato,
)
from .models import AditivoContrato, AlertaOperacional, PlanoEmpresa, AnexoOperacional, Compromisso, CompromissoItem, FechamentoMensal, HistoricoOperacional, Medicao, MedicaoItem, NotaFiscal, NotaFiscalCentroCusto, Obra, OrcamentoBaseline, OrcamentoBaselineItem, PlanoContas
from .models_aquisicoes import Cotacao, OrdemCompra, SolicitacaoCompra
from .models_qualidade import NaoConformidade
from .models_risco import Risco
from .permissions import (
    filtrar_por_obra_contexto as _filtrar_por_obra_contexto,
    get_empresa_operacional as _get_empresa_operacional,
    get_obra_do_contexto as _obter_obra_contexto,
    get_obras_permitidas as _get_obras_permitidas,
)
from .services import importar_plano_contas_excel, obter_dados_contrato, obter_dados_medicao
from .services_aprovacao import (
    can_assume_alert,
    can_approve_aditivo,
    can_approve_value,
    can_close_alert,
    can_justify_alert,
    can_submit_for_approval,
    get_limite_aprovacao,
    get_papel_aprovacao,
)
from .services_eva import EVAService
from .services_indicadores import IndicadoresService
from .services_integracao import IntegracaoService
from .services_lgpd import registrar_acesso_dado_pessoal
from .services_alertas import (
    atualizar_status_alerta,
    listar_execucoes_regras_operacionais,
    listar_alertas_operacionais_ativos,
    listar_alertas_planejamento_suprimentos,
    obter_regra_operacional,
    resumo_executivo_alertas_operacionais,
    resumo_alertas_operacionais,
    sincronizar_alertas_operacionais_obra,
)
from .domain import arredondar_moeda
from .templatetags.formatters import money_br
from .audit import AuditService
from .application.alertas import (
    acoes_alerta_permitidas,
    obter_contexto_central_alertas,
    obter_dados_painel_executivo_alertas,
)
from .cache_utils import critical_cache_add, request_local_get_or_set, resilient_cache_get_or_set
from .application.financeiro import (
    dados_fechamento_mensal_request,
    dados_projecao_financeira_request,
    registrar_fechamento_mensal,
)
from .services_jobs import listar_jobs_recentes
from .text_normalization import corrigir_mojibake

from .services_tenant import TenantService, LimitePlanoExcedido #inserido por mim

_STATIC_APP_DIR = os.path.join(os.path.dirname(__file__), "static", "app")
_PDF_LOGO_PATH = os.path.join(_STATIC_APP_DIR, "logo-construtask.png")
_EXCEL_FILL_RED = PatternFill(fill_type="solid", fgColor="840B0B")
_EXCEL_FONT_WHITE = Font(color="FFFFFF", bold=True)
_EXCEL_FONT_BLACK_BOLD = Font(color="000000", bold=True)
_EXCEL_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def _calcular_percentual(valor, total):
    if not total:
        return 0
    return round((float(valor) / float(total)) * 100, 1)


def _home_cache_ttl():
    return max(30, int(getattr(settings, "CONSTRUTASK_HOME_CACHE_TTL", 120)))


def _cache_get_or_set_local(chave, builder, ttl=None, request=None):
    return request_local_get_or_set(
        request,
        chave,
        lambda: resilient_cache_get_or_set(chave, builder, timeout=ttl or _home_cache_ttl()),
    )


def _sincronizar_alertas_operacionais_rate_limited(obra):
    chave = f"alertas:sync:obra:{obra.pk}"
    if critical_cache_add(chave, True, max(30, int(getattr(settings, "CONSTRUTASK_ALERTAS_SYNC_TTL", 120)))):
        sincronizar_alertas_operacionais_obra(obra)


def _coletar_post_int(request, campo):
    valor = (request.POST.get(campo) or "").strip()
    if not valor:
        return None
    valor_normalizado = valor.replace(".", "").replace(",", "")
    if not valor_normalizado.isdigit():
        return None
    return int(valor_normalizado)


def _nivel_resumo_alerta(total):
    if total >= 8:
        return "critico"
    if total >= 4:
        return "alto"
    if total >= 1:
        return "medio"
    return "baixo"


def _grafico_score_operacional(score_operacional):
    componentes = list(score_operacional.get("componentes") or [])
    cores = ["#a61e1e", "#d4a017", "#4f9a2f", "#3f6fd1"]
    fatias = []
    offset = Decimal("0.00")
    for indice, componente in enumerate(componentes):
        maximo = Decimal(str(componente.get("maximo") or 0))
        pontuacao = Decimal(str(componente.get("pontuacao") or 0))
        percentual = Decimal("0.00")
        if maximo:
            percentual = (pontuacao / maximo * Decimal("25.00")).quantize(Decimal("0.01"))
        inicio = offset
        fim = min(Decimal("100.00"), offset + Decimal("25.00"))
        fatias.append({
            "cor": cores[indice % len(cores)],
            "inicio": inicio,
            "fim": fim,
            "percentual_componente": percentual,
            "indice": indice + 1,
            "nome": componente.get("nome"),
            "pontuacao": pontuacao,
            "nivel": componente.get("nivel"),
            "detalhe": componente.get("detalhe"),
        })
        offset = fim
    gradiente = ", ".join(
        f"{item['cor']} {item['inicio']}% {item['fim']}%" for item in fatias
    ) or "#d1d5db 0 100%"
    return {
        "gradiente": gradiente,
        "fatias": fatias,
    }


def _obter_grupos_navegacao():
    return {
        "planejamento": {
            "slug": "planejamento",
            "titulo": "Planejamento",
            "descricao": "Organize o orcamento, acompanhe o cronograma e monitore os riscos da obra em um unico fluxo.",
            "itens": [
                {
                    "titulo": "Plano de Contas",
                    "descricao": "Estruture a EAP, acompanhe o orcado e mantenha as baselines de referencia.",
                    "url_name": "plano_contas_list",
                },
                {
                    "titulo": "Cronograma",
                    "descricao": "Importe, revise e acompanhe as atividades planejadas e realizadas da obra.",
                    "url_name": "plano_fisico_list",
                },
                {
                    "titulo": "Riscos",
                    "descricao": "Registre, trate e acompanhe os riscos operacionais e gerenciais da obra.",
                    "url_name": "risco_list",
                },
                {
                    "titulo": "Alertas Operacionais",
                    "descricao": "Centralize desvios, justificativas e encerramentos dos alertas automaticos da obra.",
                    "url_name": "alerta_operacional_list",
                },
            ],
        },
        "qualidade": {
            "slug": "qualidade",
            "titulo": "Qualidade",
            "descricao": "Concentre o controle documental, as nao conformidades e as evidencias formais da obra.",
            "itens": [
                {
                    "titulo": "Documentos",
                    "descricao": "Controle revisoes, aprovacoes e rastreabilidade dos documentos da obra.",
                    "url_name": "documento_list",
                },
                {
                    "titulo": "Nao Conformidades",
                    "descricao": "Gerencie tratativas, evidencias e encerramentos das ocorrencias de qualidade.",
                    "url_name": "nao_conformidade_list",
                },
                {
                    "titulo": "Central de Evidencias",
                    "descricao": "Acesse rapidamente os comprovantes formais e registros probatorios da operacao.",
                    "url_name": "central_evidencias",
                },
            ],
        },
        "aquisicoes": {
            "slug": "aquisicoes",
            "titulo": "Aquisicoes",
            "descricao": "Administre fornecedores, solicitacoes, cotacoes, ordens e compromissos em uma jornada unica.",
            "itens": [
                {
                    "titulo": "Fornecedores",
                    "descricao": "Cadastre e acompanhe os parceiros que participam da cadeia de suprimentos.",
                    "url_name": "fornecedor_list",
                },
                {
                    "titulo": "Solicitacoes",
                    "descricao": "Abra e acompanhe as demandas de compra originadas pela obra.",
                    "url_name": "solicitacao_compra_list",
                },
                {
                    "titulo": "Cotacoes",
                    "descricao": "Compare propostas e consolide o processo de aquisicao.",
                    "url_name": "cotacao_list",
                },
                {
                    "titulo": "Ordens de Compra",
                    "descricao": "Visualize as ordens emitidas a partir das cotacoes aprovadas.",
                    "url_name": "ordem_compra_list",
                },
                {
                    "titulo": "Compras e Contratacoes",
                    "descricao": "Gerencie pedidos, contratos, aditivos e compromissos financeiros da obra.",
                    "url_name": "compromisso_list",
                },
            ],
        },
        "comunicacoes": {
            "slug": "comunicacoes",
            "titulo": "Comunicacoes",
            "descricao": "Conduza reunioes de curto, medio e longo prazo com pauta semi automatica, definicoes e ata formal.",
            "itens": [
                {
                    "titulo": "Reunioes e Atas",
                    "descricao": "Monte pautas automaticamente a partir do contexto da obra, registre respostas e envie a ata para aprovacao.",
                    "url_name": "reuniao_comunicacao_list",
                },
            ],
        },
        "relatorios": {
            "slug": "relatorios",
            "titulo": "Relatorios",
            "descricao": "Reuna os principais documentos gerenciais e de acompanhamento da obra.",
            "itens": [
                {
                    "titulo": "Dossie da Obra",
                    "descricao": "Consulte o documento consolidado de acompanhamento tecnico e gerencial.",
                    "url_name": "dossie_obra",
                },
                {
                    "titulo": "Fechamento Mensal",
                    "descricao": "Analise o fechamento periodico da obra com base em custo e execucao.",
                    "url_name": "fechamento_mensal",
                },
                {
                    "titulo": "Curva ABC",
                    "descricao": "Veja a concentracao de relevancia financeira dos centros de custo da obra.",
                    "url_name": "curva_abc",
                },
            ],
        },
        "juridico": {
            "slug": "juridico",
            "titulo": "Juridico",
            "descricao": "Centralize governanca, transparência e documentos institucionais do sistema.",
            "itens": [
                {
                    "titulo": "LGPD",
                    "descricao": "Acompanhe a governanca de dados pessoais, trilhas e evidencias de conformidade.",
                    "url_name": "lgpd_governanca",
                },
                {
                    "titulo": "Termos de Uso",
                    "descricao": "Consulte as regras de uso e responsabilidade do sistema.",
                    "url_name": "termos_uso",
                },
                {
                    "titulo": "Politica de Privacidade",
                    "descricao": "Visualize os principios de tratamento e protecao de dados adotados.",
                    "url_name": "politica_privacidade",
                },
            ],
        },
        "financeiro": {
            "slug": "financeiro",
            "titulo": "Financeiro",
            "descricao": "Monitore a execucao financeira da obra com foco em notas, medicoes e projecoes.",
            "itens": [
                {
                    "titulo": "Notas Fiscais",
                    "descricao": "Controle emissao, rateio e situacao financeira das notas da obra.",
                    "url_name": "nota_fiscal_list",
                },
                {
                    "titulo": "Medicoes",
                    "descricao": "Acompanhe lancamentos, aprovacoes e valores medidos.",
                    "url_name": "medicao_list",
                },
                {
                    "titulo": "Projecao Financeira",
                    "descricao": "Projete entradas e saidas futuras com base no andamento da obra.",
                    "url_name": "projecao_financeira",
                },
            ],
        },
    }


def _datahora_local(datahora):
    if not datahora:
        return None
    return timezone.localtime(datahora)


def _normalizar_texto_exportacao(valor):
    if valor is None:
        return "-"
    texto = str(valor)
    return corrigir_mojibake(texto)


def _sanear_texto_exportacao_seguro(texto):
    texto = "-" if texto is None else str(texto)
    texto = corrigir_mojibake(texto)
    texto = _sanear_texto_exportacao(texto)
    reparos = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2026": "...",
        "\xa0": " ",
        "\xa2": "\u00e2",
        "\xb3": "\u00f3",
        "\xb4": "\u00f4",
        "\xb5": "\u00f5",
        "\xa3": "\u00e3",
        "\xa9": "\u00e9",
        "\xaa": "\u00ea",
        "\xad": "\u00ed",
        "\xa1": "\u00e1",
        "\xa7": "\u00e7",
        "\xba": "\u00fa",
    }
    for antigo, novo in reparos.items():
        texto = texto.replace(antigo, novo)
    return unicodedata.normalize("NFC", texto)


def _normalizar_linhas_exportacao(linhas):
    normalizadas = []
    for linha in linhas:
        if isinstance(linha, dict):
            normalizadas.append({chave: _sanear_texto_exportacao_seguro(valor) for chave, valor in linha.items()})
        else:
            normalizadas.append(_sanear_texto_exportacao_seguro(linha))
    return normalizadas


def _sanear_texto_exportacao(texto):
    texto = "-" if texto is None else str(texto)
    return corrigir_mojibake(texto)


def _png_scanline_unfilter(filtro, linha_filtrada, anterior, bpp):
    resultado = bytearray(len(linha_filtrada))
    if filtro == 0:
        return bytearray(linha_filtrada)
    if filtro == 1:
        for indice, valor in enumerate(linha_filtrada):
            esquerda = resultado[indice - bpp] if indice >= bpp else 0
            resultado[indice] = (valor + esquerda) & 0xFF
        return resultado
    if filtro == 2:
        for indice, valor in enumerate(linha_filtrada):
            acima = anterior[indice] if anterior else 0
            resultado[indice] = (valor + acima) & 0xFF
        return resultado
    if filtro == 3:
        for indice, valor in enumerate(linha_filtrada):
            esquerda = resultado[indice - bpp] if indice >= bpp else 0
            acima = anterior[indice] if anterior else 0
            resultado[indice] = (valor + ((esquerda + acima) // 2)) & 0xFF
        return resultado
    if filtro == 4:
        def _paeth(a, b, c):
            p = a + b - c
            pa = abs(p - a)
            pb = abs(p - b)
            pc = abs(p - c)
            if pa <= pb and pa <= pc:
                return a
            if pb <= pc:
                return b
            return c

        for indice, valor in enumerate(linha_filtrada):
            esquerda = resultado[indice - bpp] if indice >= bpp else 0
            acima = anterior[indice] if anterior else 0
            diagonal = anterior[indice - bpp] if anterior and indice >= bpp else 0
            resultado[indice] = (valor + _paeth(esquerda, acima, diagonal)) & 0xFF
        return resultado
    raise ValueError("Filtro PNG nao suportado.")


def _carregar_png_para_pdf(caminho_logo):
    if not os.path.exists(caminho_logo):
        return None

    with open(caminho_logo, "rb") as arquivo_logo:
        dados = arquivo_logo.read()

    assinatura = b"\x89PNG\r\n\x1a\n"
    if not dados.startswith(assinatura):
        return None

    cursor = len(assinatura)
    largura = altura = None
    bit_depth = color_type = interlace = None
    idat_chunks = []

    while cursor < len(dados):
        tamanho = struct.unpack(">I", dados[cursor:cursor + 4])[0]
        cursor += 4
        tipo = dados[cursor:cursor + 4]
        cursor += 4
        payload = dados[cursor:cursor + tamanho]
        cursor += tamanho + 4

        if tipo == b"IHDR":
            largura, altura, bit_depth, color_type, _, _, interlace = struct.unpack(">IIBBBBB", payload)
        elif tipo == b"IDAT":
            idat_chunks.append(payload)
        elif tipo == b"IEND":
            break

    if not largura or not altura or bit_depth != 8 or interlace != 0 or color_type not in (2, 6):
        return None

    canais = 4 if color_type == 6 else 3
    bytes_por_pixel = canais
    bytes_por_linha = largura * canais
    dados_descomprimidos = zlib.decompress(b"".join(idat_chunks))

    cursor = 0
    linha_anterior = None
    rgb = bytearray()
    alpha_bytes = bytearray() if canais == 4 else None
    for _ in range(altura):
        filtro = dados_descomprimidos[cursor]
        cursor += 1
        linha_filtrada = dados_descomprimidos[cursor:cursor + bytes_por_linha]
        cursor += bytes_por_linha
        linha = _png_scanline_unfilter(filtro, linha_filtrada, linha_anterior, bytes_por_pixel)
        linha_anterior = linha

        if canais == 3:
            rgb.extend(linha)
            continue

        for indice in range(0, len(linha), 4):
            vermelho, verde, azul, alpha = linha[indice:indice + 4]
            rgb.extend((vermelho, verde, azul))
            alpha_bytes.extend((alpha,))

    return {
        "width": largura,
        "height": altura,
        "stream": zlib.compress(bytes(rgb)),
        "alpha_stream": zlib.compress(bytes(alpha_bytes)) if alpha_bytes is not None else None,
    }

def _obter_contrato_from_request(request, instance=None):
    contrato = None
    contrato_id = request.POST.get("contrato") or request.GET.get("contrato")
    if contrato_id:
        try:
            contrato = Compromisso.objects.get(pk=contrato_id, tipo="CONTRATO", status="APROVADO")
        except Compromisso.DoesNotExist:
            contrato = None
    elif instance and getattr(instance, "contrato_id", None):
        contrato = instance.contrato
    return contrato


def _obter_origem_nota(request, instance=None):
    pedido = None
    medicao = None
    pedido_id = request.POST.get("pedido_compra") or request.GET.get("pedido_compra")
    medicao_id = request.POST.get("medicao") or request.GET.get("medicao")
    if pedido_id:
        try:
            pedido = Compromisso.objects.get(pk=pedido_id, tipo="PEDIDO_COMPRA", status="APROVADO")
        except Compromisso.DoesNotExist:
            pedido = None
    if medicao_id:
        try:
            medicao = Medicao.objects.get(pk=medicao_id, status="APROVADA")
        except Medicao.DoesNotExist:
            medicao = None
    if not pedido and not medicao and instance:
        pedido = getattr(instance, "pedido_compra", None)
        medicao = getattr(instance, "medicao", None)
    return pedido, medicao


def _construir_formset_medicao(*, data=None, instance=None, prefix="itens", contrato=None):
    centros_queryset = obter_centros_do_contrato(contrato)
    return MedicaoItemFormSet(
        data=data,
        instance=instance,
        prefix=prefix,
        centros_queryset=centros_queryset,
    )


def _construir_formset_nota(*, data=None, instance=None, prefix="rateio", pedido=None, medicao=None, obra=None):
    centros_queryset = obter_centros_da_origem_nota(instance, pedido, medicao, obra)
    return NotaFiscalCentroCustoFormSet(
        data=data,
        instance=instance,
        prefix=prefix,
        centros_queryset=centros_queryset,
    )


def _obter_alcada_contexto(user, valor):
    papel = get_papel_aprovacao(user)
    limite = get_limite_aprovacao(user)
    if limite is None:
        limite_label = "ilimitada"
    else:
        limite_label = money_br(limite)
    return {
        "papel_aprovacao": papel,
        "limite_aprovacao": limite,
        "limite_aprovacao_label": limite_label,
        "pode_enviar_para_aprovacao": can_submit_for_approval(user),
        "pode_aprovar": can_approve_value(user, valor),
    }


def _criar_baseline_orcamento(obra, *, descricao, usuario):
    planos = _consolidar_plano_contas(
        PlanoContas.objects.filter(obra=obra).annotate(filhos_count=Count("filhos")).order_by("tree_id", "lft")
    )
    baseline = OrcamentoBaseline.objects.create(
        obra=obra,
        descricao=descricao,
        criado_por=usuario,
    )
    itens = [
        OrcamentoBaselineItem(
            baseline=baseline,
            codigo=plano.codigo,
            descricao=plano.descricao,
            parent_codigo=plano.parent.codigo if plano.parent_id else "",
            level=plano.level,
            unidade=plano.unidade or "",
            quantidade=plano.quantidade,
            valor_unitario=plano.valor_unitario,
            valor_total=plano.valor_total,
            valor_total_consolidado=plano.valor_total_consolidado_calc,
        )
        for plano in planos
    ]
    OrcamentoBaselineItem.objects.bulk_create(itens)
    return baseline


def _enviar_baseline_para_aprovacao(request, baseline):
    if not can_submit_for_approval(request.user):
        messages.error(request, "Seu usuario nao possui funcao operacional para enviar a baseline para aprovacao.")
        return False
    if baseline.status == "EM_APROVACAO":
        messages.info(request, "Esta baseline ja esta em aprovacao.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    baseline.status = "EM_APROVACAO"
    baseline.enviado_para_aprovacao_em = timezone.now()
    baseline.enviado_para_aprovacao_por = request.user
    baseline.parecer_aprovacao = parecer
    baseline.aprovado_em = None
    baseline.aprovado_por = None
    baseline.is_ativa = False
    baseline.save()
    _registrar_historico(
        "BASELINE_ORCAMENTO",
        baseline.obra,
        f"Baseline de orcamento enviada para aprovacao: {baseline.descricao}" + (f" Parecer: {parecer}" if parecer else ""),
        request.user,
    )
    messages.success(request, "Baseline enviada para aprovacao.")
    return True


def _aprovar_baseline(request, baseline):
    valor = baseline.valor_total
    if not can_approve_value(request.user, valor):
        messages.error(request, "Sua funcao nao possui alcada suficiente para aprovar esta baseline.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    before = AuditService.instance_to_dict(baseline)
    OrcamentoBaseline.objects.filter(obra=baseline.obra, is_ativa=True).update(is_ativa=False)
    baseline.status = "APROVADA"
    baseline.aprovado_em = timezone.now()
    baseline.aprovado_por = request.user
    baseline.parecer_aprovacao = parecer
    if not baseline.enviado_para_aprovacao_em:
        baseline.enviado_para_aprovacao_em = timezone.now()
    if not baseline.enviado_para_aprovacao_por:
        baseline.enviado_para_aprovacao_por = request.user
    baseline.is_ativa = True
    baseline.save()
    after = AuditService.instance_to_dict(baseline)
    AuditService.log_event(request, "APPROVE", baseline, before, after)
    _registrar_historico(
        "BASELINE_ORCAMENTO",
        baseline.obra,
        f"Baseline de orcamento aprovada: {baseline.descricao}" + (f" Parecer: {parecer}" if parecer else ""),
        request.user,
    )
    messages.success(request, "Baseline aprovada com sucesso.")
    return True


def _retornar_baseline_para_ajuste(request, baseline):
    valor = baseline.valor_total
    if not can_approve_value(request.user, valor):
        messages.error(request, "Sua funcao nao possui alcada suficiente para devolver esta baseline para ajuste.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    if not parecer:
        messages.error(request, "Informe um parecer para devolver a baseline para ajuste.")
        return False
    before = AuditService.instance_to_dict(baseline)
    baseline.status = "RASCUNHO"
    baseline.parecer_aprovacao = parecer
    baseline.aprovado_em = None
    baseline.aprovado_por = None
    baseline.is_ativa = False
    baseline.save()
    after = AuditService.instance_to_dict(baseline)
    AuditService.log_event(request, "REJECT", baseline, before, after)
    _registrar_historico(
        "BASELINE_ORCAMENTO",
        baseline.obra,
        f"Baseline de orcamento devolvida para ajuste: {baseline.descricao}. Parecer: {parecer}",
        request.user,
    )
    messages.success(request, "Baseline devolvida para ajuste.")
    return True


def _enviar_documento_para_aprovacao(request, objeto, *, status_em_aprovacao, descricao):
    if not can_submit_for_approval(request.user):
        messages.error(request, "Seu usuario nao possui funcao operacional para enviar este registro para aprovacao.")
        return False
    if objeto.status == status_em_aprovacao:
        messages.info(request, "Este registro ja esta em aprovacao.")
        return False

    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    objeto.status = status_em_aprovacao
    objeto.enviado_para_aprovacao_em = timezone.now()
    objeto.enviado_para_aprovacao_por = request.user
    objeto.parecer_aprovacao = parecer
    objeto.aprovado_em = None
    objeto.aprovado_por = None
    objeto.save()
    descricao_historico = descricao if not parecer else f"{descricao} Parecer: {parecer}"
    _registrar_historico("APROVACAO", objeto, descricao_historico, request.user)
    messages.success(request, "Registro enviado para aprovacao.")
    return True


def _aprovar_documento(request, objeto, *, valor, status_aprovado, descricao):
    if not can_approve_value(request.user, valor):
        messages.error(request, "Sua funcao nao possui alcada suficiente para aprovar este valor.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    before = AuditService.instance_to_dict(objeto)
    objeto.status = status_aprovado
    objeto.parecer_aprovacao = parecer
    objeto.aprovado_em = timezone.now()
    objeto.aprovado_por = request.user
    if not objeto.enviado_para_aprovacao_em:
        objeto.enviado_para_aprovacao_em = timezone.now()
    if not objeto.enviado_para_aprovacao_por:
        objeto.enviado_para_aprovacao_por = request.user
    objeto.save()
    after = AuditService.instance_to_dict(objeto)
    AuditService.log_event(request, "APPROVE", objeto, before, after)
    descricao_historico = descricao if not parecer else f"{descricao} Parecer: {parecer}"
    _registrar_historico("APROVACAO", objeto, descricao_historico, request.user)
    messages.success(request, "Registro aprovado com sucesso.")
    return True


def _retornar_documento_para_ajuste(
    request,
    objeto,
    *,
    valor,
    status_ajuste,
    descricao,
):
    if not can_approve_value(request.user, valor):
        messages.error(request, "Sua funcao nao possui alcada suficiente para devolver este valor para ajuste.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    if not parecer:
        messages.error(request, "Informe um parecer para devolver o registro para ajuste.")
        return False
    before = AuditService.instance_to_dict(objeto)
    objeto.status = status_ajuste
    objeto.parecer_aprovacao = parecer
    objeto.aprovado_em = None
    objeto.aprovado_por = None
    objeto.save()
    after = AuditService.instance_to_dict(objeto)
    AuditService.log_event(request, "REJECT", objeto, before, after)
    _registrar_historico("APROVACAO", objeto, f"{descricao} Parecer: {parecer}", request.user)
    messages.success(request, "Registro devolvido para ajuste.")
    return True


def _valor_total_aditivo(aditivo):
    return aditivo.itens.aggregate(total=Sum("valor"))["total"] or Decimal("0.00")


def _aprovar_aditivo_contrato(request, aditivo):
    valor = _valor_total_aditivo(aditivo)
    if not can_approve_aditivo(request.user, valor):
        messages.error(request, "Somente Coordenador de Obras ou Gerente de Obras podem aprovar aditivos, respeitando a alcada aplicavel.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    before = AuditService.instance_to_dict(aditivo)
    aditivo.status = "APROVADO"
    aditivo.parecer_aprovacao = parecer
    aditivo.aprovado_em = timezone.now()
    aditivo.aprovado_por = request.user
    if not aditivo.enviado_para_aprovacao_em:
        aditivo.enviado_para_aprovacao_em = timezone.now()
    if not aditivo.enviado_para_aprovacao_por:
        aditivo.enviado_para_aprovacao_por = request.user
    aditivo.save()
    after = AuditService.instance_to_dict(aditivo)
    AuditService.log_event(request, "APPROVE", aditivo, before, after)
    descricao = f"Aditivo {aditivo.get_tipo_display()} do contrato {aditivo.contrato.numero} aprovado."
    if parecer:
        descricao = f"{descricao} Parecer: {parecer}"
    _registrar_historico("ADITIVO_APROVADO", aditivo.contrato, descricao, request.user)
    messages.success(request, "Aditivo aprovado com sucesso.")
    return True


def _retornar_aditivo_contrato_para_ajuste(request, aditivo):
    valor = _valor_total_aditivo(aditivo)
    if not can_approve_aditivo(request.user, valor):
        messages.error(request, "Somente Coordenador de Obras ou Gerente de Obras podem devolver aditivos para ajuste, respeitando a alcada aplicavel.")
        return False
    parecer = (request.POST.get("parecer_aprovacao") or "").strip()
    if not parecer:
        messages.error(request, "Informe um parecer para devolver o aditivo para ajuste.")
        return False
    before = AuditService.instance_to_dict(aditivo)
    aditivo.status = "RASCUNHO"
    aditivo.parecer_aprovacao = parecer
    aditivo.aprovado_em = None
    aditivo.aprovado_por = None
    aditivo.save()
    after = AuditService.instance_to_dict(aditivo)
    AuditService.log_event(request, "REJECT", aditivo, before, after)
    _registrar_historico(
        "ADITIVO_AJUSTE",
        aditivo.contrato,
        f"Aditivo {aditivo.get_tipo_display()} do contrato {aditivo.contrato.numero} devolvido para ajuste. Parecer: {parecer}",
        request.user,
    )
    messages.success(request, "Aditivo devolvido para ajuste.")
    return True


def _enviar_aditivo_contrato_para_aprovacao(request, aditivo):
    if not can_submit_for_approval(request.user):
        messages.error(request, "Seu usuario nao possui funcao operacional para enviar este aditivo para aprovacao.")
        return False
    if aditivo.status == "EM_APROVACAO":
        messages.info(request, "Este aditivo ja esta em aprovacao.")
        return False
    before = AuditService.instance_to_dict(aditivo)
    aditivo.status = "EM_APROVACAO"
    aditivo.enviado_para_aprovacao_em = timezone.now()
    aditivo.enviado_para_aprovacao_por = request.user
    aditivo.aprovado_em = None
    aditivo.aprovado_por = None
    aditivo.save()
    after = AuditService.instance_to_dict(aditivo)
    AuditService.log_event(request, "SUBMIT", aditivo, before, after)
    _registrar_historico(
        "ADITIVO_ENVIO",
        aditivo.contrato,
        f"Aditivo {aditivo.get_tipo_display()} do contrato {aditivo.contrato.numero} enviado para aprovacao.",
        request.user,
    )
    messages.success(request, "Aditivo enviado para aprovacao.")
    return True


def _mapa_somas_por_centro(modelo, campo_valor):
    return {
        row["centro_custo_id"]: row["total"] or Decimal("0.00")
        for row in modelo.objects.values("centro_custo_id").annotate(total=Sum(campo_valor))
    }


def _filtrar_periodo(queryset, campo_data, data_inicio, data_fim):
    if data_inicio:
        queryset = queryset.filter(**{f"{campo_data}__gte": data_inicio})
    if data_fim:
        queryset = queryset.filter(**{f"{campo_data}__lte": data_fim})
    return queryset


def _filtros_compromissos(request, queryset):
    termo = request.GET.get("q", "").strip()
    obra_id = request.GET.get("obra", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor = request.GET.get("fornecedor", "").strip()
    responsavel = request.GET.get("responsavel", "").strip()
    centro_custo_id = request.GET.get("centro_custo", "").strip()
    data_inicio = request.GET.get("data_inicio", "").strip()
    data_fim = request.GET.get("data_fim", "").strip()

    if termo:
        queryset = queryset.filter(
            Q(numero__icontains=termo)
            | Q(fornecedor__icontains=termo)
            | Q(cnpj__icontains=termo)
            | Q(responsavel__icontains=termo)
            | Q(descricao__icontains=termo)
        ).distinct()
    if obra_id:
        queryset = queryset.filter(obra_id=obra_id)
    if status:
        queryset = queryset.filter(status=status)
    if fornecedor:
        queryset = queryset.filter(fornecedor__icontains=fornecedor)
    if responsavel:
        queryset = queryset.filter(responsavel__icontains=responsavel)
    if centro_custo_id:
        queryset = queryset.filter(Q(centro_custo_id=centro_custo_id) | Q(itens__centro_custo_id=centro_custo_id)).distinct()

    return _filtrar_periodo(queryset, "data_assinatura", data_inicio, data_fim)


def _filtros_medicoes(request, queryset):
    termo = request.GET.get("q", "").strip()
    obra_id = request.GET.get("obra", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor = request.GET.get("fornecedor", "").strip()
    responsavel = request.GET.get("responsavel", "").strip()
    contrato = request.GET.get("contrato", "").strip()
    centro_custo_id = request.GET.get("centro_custo", "").strip()
    data_inicio = request.GET.get("data_inicio", "").strip()
    data_fim = request.GET.get("data_fim", "").strip()

    if termo:
        queryset = queryset.filter(
            Q(numero_da_medicao__icontains=termo)
            | Q(fornecedor__icontains=termo)
            | Q(cnpj__icontains=termo)
        )
    if obra_id:
        queryset = queryset.filter(obra_id=obra_id)
    if status:
        queryset = queryset.filter(status=status)
    if fornecedor:
        queryset = queryset.filter(fornecedor__icontains=fornecedor)
    if responsavel:
        queryset = queryset.filter(responsavel__icontains=responsavel)
    if contrato:
        queryset = queryset.filter(contrato__numero__icontains=contrato)
    if centro_custo_id:
        queryset = queryset.filter(Q(centro_custo_id=centro_custo_id) | Q(itens__centro_custo_id=centro_custo_id)).distinct()

    return _filtrar_periodo(queryset, "data_medicao", data_inicio, data_fim)


def _filtros_notas(request, queryset):
    termo = request.GET.get("q", "").strip()
    obra_id = request.GET.get("obra", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor = request.GET.get("fornecedor", "").strip()
    contrato = request.GET.get("contrato", "").strip()
    centro_custo_id = request.GET.get("centro_custo", "").strip()
    data_inicio = request.GET.get("data_inicio", "").strip()
    data_fim = request.GET.get("data_fim", "").strip()

    if termo:
        queryset = queryset.filter(
            Q(numero__icontains=termo)
            | Q(fornecedor__icontains=termo)
            | Q(cnpj__icontains=termo)
        )
    if obra_id:
        queryset = queryset.filter(obra_id=obra_id)
    if status:
        queryset = queryset.filter(status=status)
    if fornecedor:
        queryset = queryset.filter(fornecedor__icontains=fornecedor)
    if contrato:
        queryset = queryset.filter(Q(medicao__contrato__numero__icontains=contrato) | Q(pedido_compra__numero__icontains=contrato))
    if centro_custo_id:
        queryset = queryset.filter(centros_custo__centro_custo_id=centro_custo_id).distinct()

    return _filtrar_periodo(queryset, "data_emissao", data_inicio, data_fim)


def _anotar_execucao_compromissos(queryset):
    return queryset.annotate(
        total_medicoes_execucao=Coalesce(Sum("medicoes__valor_medido"), Decimal("0.00")),
        total_notas_execucao=Coalesce(Sum("notas_fiscais_material__valor_total"), Decimal("0.00")),
    ).annotate(
        valor_executado_anotado=Case(
            When(tipo="CONTRATO", then=F("total_medicoes_execucao")),
            default=F("total_notas_execucao"),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        )
    ).annotate(
        saldo_anotado=ExpressionWrapper(
            F("valor_contratado") - F("valor_executado_anotado"),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        )
    )


def _consolidar_plano_contas(planos_queryset):
    planos = list(
        planos_queryset.only(
            "id",
            "codigo",
            "descricao",
            "unidade",
            "quantidade",
            "valor_unitario",
            "valor_total",
            "parent_id",
            "tree_id",
            "lft",
            "rght",
            "level",
        )
    )
    planos_por_id = {plano.id: plano for plano in planos}
    comprometido_por_centro = _mapa_somas_por_centro(CompromissoItem, "valor_total")
    medido_por_centro = _mapa_somas_por_centro(MedicaoItem, "valor_total")
    executado_por_centro = _mapa_somas_por_centro(NotaFiscalCentroCusto, "valor")

    for plano in planos:
        plano.valor_total_consolidado_calc = plano.valor_total or Decimal("0.00")
        plano.valor_comprometido_calc = comprometido_por_centro.get(plano.id, Decimal("0.00"))
        plano.valor_medido_calc = medido_por_centro.get(plano.id, Decimal("0.00"))
        plano.valor_executado_calc = executado_por_centro.get(plano.id, Decimal("0.00"))
        plano.nivel_indentacao = getattr(plano, "level", 0)

    for plano in reversed(planos):
        if not plano.parent_id:
            continue
        parent = planos_por_id.get(plano.parent_id)
        if not parent:
            continue
        parent.valor_total_consolidado_calc += plano.valor_total_consolidado_calc
        parent.valor_comprometido_calc += plano.valor_comprometido_calc
        parent.valor_medido_calc += plano.valor_medido_calc
        parent.valor_executado_calc += plano.valor_executado_calc

    for plano in planos:
        plano.saldo_a_comprometer_calc = plano.valor_total_consolidado_calc - plano.valor_comprometido_calc
        plano.saldo_a_medir_calc = plano.valor_comprometido_calc - plano.valor_medido_calc
        plano.saldo_a_executar_calc = plano.valor_total_consolidado_calc - plano.valor_executado_calc

    return planos


def _ajustar_larguras_excel(worksheet):
    # Controla a largura automÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡tica das colunas no Excel.
    for coluna in worksheet.columns:
        indice_coluna = coluna[0].column
        largura = 12
        for celula in coluna:
            try:
                valor = "" if celula.value is None else str(celula.value)
            except Exception:
                valor = ""
            largura = max(largura, min(len(valor) + 2, 36))
        worksheet.column_dimensions[get_column_letter(indice_coluna)].width = largura


def _estilizar_tabela_excel(worksheet, linha_cabecalho, linha_inicio_dados, linha_fim_dados):
    # Controla borda, quebra de texto, alinhamento e altura das linhas no Excel.
    if linha_cabecalho:
        for celula in worksheet[linha_cabecalho]:
            if celula.value is None:
                continue
            celula.fill = _EXCEL_FILL_RED
            celula.font = _EXCEL_FONT_WHITE
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celula.border = _EXCEL_BORDER
        worksheet.row_dimensions[linha_cabecalho].height = 32

    for linha in worksheet.iter_rows(min_row=linha_inicio_dados, max_row=linha_fim_dados):
        altura_linha = 20
        for celula in linha:
            if celula.value is None:
                continue
            celula.border = _EXCEL_BORDER
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            valor = str(celula.value)
            altura_linha = max(altura_linha, min(48, 16 + ((len(valor) // 35) * 12)))
        worksheet.row_dimensions[linha[0].row].height = altura_linha


def _aplicar_layout_excel_relatorio(worksheet, titulo_relatorio, subtitulo=None):
    # Controla o cabeÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â§alho institucional e a estrutura visual padrÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â£o dos relatÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³rios em Excel.
    ultima_coluna = max(worksheet.max_column, 2)
    worksheet.insert_rows(1, amount=3)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_coluna)
    worksheet["A1"] = "RELATORIOS DE ACOMPANHAMENTO DE OBRA"
    worksheet["A1"].font = Font(color="000000", bold=True, size=16)
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_coluna)
    worksheet["A2"] = titulo_relatorio
    worksheet["A2"].fill = _EXCEL_FILL_RED
    worksheet["A2"].font = _EXCEL_FONT_WHITE
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A2"].border = _EXCEL_BORDER

    if subtitulo:
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ultima_coluna)
        worksheet["A3"] = subtitulo
        worksheet["A3"].font = _EXCEL_FONT_BLACK_BOLD
        worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.freeze_panes = "A5"
    _estilizar_tabela_excel(worksheet, 4, 5, worksheet.max_row)
    _ajustar_larguras_excel(worksheet)


def _exportar_excel_response(nome_arquivo, sheet_name, linhas):
    output = BytesIO()
    linhas = _normalizar_linhas_exportacao(linhas)
    dataframe = pd.DataFrame(linhas)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        _aplicar_layout_excel_relatorio(worksheet, sheet_name)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


import re


def _pdf_escape(texto):
    """
    Sanitiza texto antes de escrever conteudo bruto no PDF.
    """
    texto = "-" if texto is None else str(texto)
    texto = _sanear_texto_exportacao_seguro(texto)
    for termo_antigo, termo_novo in {
        "HISTRICO": "HIST\u00d3RICO",
        "APROVAO": "APROVA\u00c7\u00c3O",
        "DESCRIO": "DESCRI\u00c7\u00c3O",
        "CONTRATAES": "CONTRATA\u00c7\u00d5ES",
        "MEDIO": "MEDI\u00c7\u00c3O",
    }.items():
        texto = texto.replace(termo_antigo, termo_novo)
    return (
        texto
        .replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
    )


def _pdf_wrap_text(texto, largura_maxima, tamanho_fonte):
    """
    Quebra texto em multiplas linhas de forma mais conservadora para evitar
    estouro horizontal dentro da celula do PDF.
    """
    texto = _sanear_texto_exportacao_seguro(texto or "-").strip()
    if not texto:
        return ["-"]

    largura_maxima = max(float(largura_maxima or 0), 20.0)
    tamanho_fonte = max(float(tamanho_fonte or 0), 6.0)

    largura_util = max(largura_maxima - 6.0, 10.0)

    paragrafos = texto.splitlines()
    linhas_finais = []

    for paragrafo in paragrafos:
        paragrafo = re.sub(r"\s+", " ", paragrafo).strip()

        if not paragrafo:
            linhas_finais.append("")
            continue

        palavras = paragrafo.split(" ")
        linha_atual = ""

        for palavra in palavras:
            while palavra and _pdf_estimar_largura_texto(palavra, tamanho_fonte) > largura_util:
                if linha_atual:
                    linhas_finais.append(linha_atual)
                    linha_atual = ""
                parte, resto = _pdf_quebrar_palavra_longa(palavra, largura_util, tamanho_fonte)
                linhas_finais.append(parte)
                palavra = resto

            if not linha_atual:
                linha_atual = palavra
            else:
                candidato = f"{linha_atual} {palavra}"
                if _pdf_estimar_largura_texto(candidato, tamanho_fonte) <= largura_util:
                    linha_atual = candidato
                else:
                    linhas_finais.append(linha_atual)
                    linha_atual = palavra

        if linha_atual:
            linhas_finais.append(linha_atual)

    return linhas_finais or ["-"]


def _pdf_estimar_largura_texto(texto, tamanho_fonte):
    texto = _sanear_texto_exportacao_seguro(texto or "")
    largura = 0.0
    for caractere in texto:
        if caractere in "W@%MmQGOD":
            fator = 0.92
        elif caractere in "ABCDEFGHKNOPRSTUVXYZ":
            fator = 0.78
        elif caractere in "mw":
            fator = 0.82
        elif caractere in "ijlI1|.,:;!'` ":
            fator = 0.30
        elif caractere in "-_/\\()[]{}":
            fator = 0.42
        elif caractere.isdigit():
            fator = 0.60
        else:
            fator = 0.56
        largura += tamanho_fonte * fator
    return largura


def _pdf_ajustar_texto_para_largura(texto, largura_maxima, tamanho_fonte, sufixo="..."):
    texto = _sanear_texto_exportacao_seguro(texto or "")
    if _pdf_estimar_largura_texto(texto, tamanho_fonte) <= largura_maxima:
        return texto
    base = texto
    while base:
        candidato = f"{base}{sufixo}"
        if _pdf_estimar_largura_texto(candidato, tamanho_fonte) <= largura_maxima:
            return candidato
        base = base[:-1].rstrip()
    return sufixo


def _pdf_quebrar_palavra_longa(palavra, largura_maxima, tamanho_fonte):
    ponto_quebra = 1
    for indice in range(1, len(palavra) + 1):
        trecho = palavra[:indice]
        if _pdf_estimar_largura_texto(trecho, tamanho_fonte) > largura_maxima:
            break
        ponto_quebra = indice
    if ponto_quebra >= len(palavra):
        return palavra, ""
    if ponto_quebra > 3:
        return f"{palavra[:ponto_quebra - 1]}-", palavra[ponto_quebra - 1:]
    return palavra[:ponto_quebra], palavra[ponto_quebra:]


def _pdf_text_commands(x, y, texto, *, fonte="F1", tamanho=10):
    return [
        "0 0 0 rg",
        "BT",
        f"/{fonte} {tamanho} Tf",
        f"{x:.2f} {y:.2f} Td",
        f"({_pdf_escape(texto)}) Tj",
        "ET",
    ]


def _pdf_text_commands_color(x, y, texto, *, fonte="F1", tamanho=10, rgb=(0, 0, 0)):
    r, g, b = rgb
    return [
        f"{r} {g} {b} rg",
        "BT",
        f"/{fonte} {tamanho} Tf",
        f"{x:.2f} {y:.2f} Td",
        f"({_pdf_escape(texto)}) Tj",
        "ET",
        "0 0 0 rg",
    ]


def _pdf_normalizar_colunas(colunas):
    colunas_normalizadas = []
    for coluna in colunas:
        if isinstance(coluna, dict):
            colunas_normalizadas.append(
                {
                    "chave": coluna.get("chave") or coluna.get("titulo") or "valor",
                    "titulo": coluna.get("titulo") or coluna.get("chave") or "Valor",
                    "largura": float(coluna.get("largura", 80)),
                    "align": coluna.get("align") or _pdf_inferir_alinhamento_coluna(coluna.get("titulo") or coluna.get("chave") or ""),
                }
            )
        else:
            titulo, largura = coluna
            colunas_normalizadas.append(
                {
                    "chave": titulo,
                    "titulo": titulo,
                    "largura": float(largura),
                    "align": _pdf_inferir_alinhamento_coluna(titulo),
                }
            )
    return colunas_normalizadas


def _pdf_inferir_alinhamento_coluna(titulo):
    titulo_normalizado = _sanear_texto_exportacao_seguro(titulo or "").lower()
    if any(chave in titulo_normalizado for chave in ["valor", "saldo", "total", "quantidade", "%", "percent", "nivel"]):
        return "right"
    if any(chave in titulo_normalizado for chave in ["data", "emissao", "vencimento", "validade"]):
        return "center"
    return "left"


def _pdf_x_texto_alinhado(x_cursor, largura, texto, tamanho_fonte, alinhamento, padding_x):
    largura_texto = _pdf_estimar_largura_texto(texto, tamanho_fonte)
    if alinhamento == "right":
        return max(x_cursor + padding_x, x_cursor + largura - padding_x - largura_texto)
    if alinhamento == "center":
        return max(x_cursor + padding_x, x_cursor + ((largura - largura_texto) / 2))
    return x_cursor + padding_x


def _pdf_valor_documento(valor, *, vazio="N\u00e3o informado"):
    if valor is None:
        return vazio
    if isinstance(valor, str):
        texto = _sanear_texto_exportacao_seguro(valor).strip()
        return texto if texto and texto not in {"-", "--", "- - - -"} else vazio
    return str(valor)


def _pdf_normalizar_linhas_documento(linhas, colunas):
    colunas_normalizadas = _pdf_normalizar_colunas(colunas)
    linhas_normalizadas = []
    for linha in list(linhas or []):
        linha_normalizada = {}
        for coluna in colunas_normalizadas:
            linha_normalizada[coluna["chave"]] = _pdf_valor_documento(linha.get(coluna["chave"]))
        linhas_normalizadas.append(linha_normalizada)
    if linhas_normalizadas:
        return linhas_normalizadas
    return [{colunas_normalizadas[0]["chave"]: "Nenhum registro encontrado"}]


def _pdf_obter_metadados_relatorio(titulo, resumo):
    resumo = resumo or {}
    return {
        "sistema": "CONSTRUTASK",
        "relatorio": _pdf_valor_documento(titulo, vazio="Relat\u00f3rio"),
        "codigo_documento": _pdf_valor_documento(
            resumo.get("Numero")
            or resumo.get("C\u00f3digo")
            or resumo.get("C\u00f3digo Documento")
            or resumo.get("Identificador")
            or resumo.get("Identificador da Evid\u00eancia")
        ),
        "obra": _pdf_valor_documento(resumo.get("Obra")),
        "data_emissao": _pdf_valor_documento(resumo.get("Emitido em") or _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")),
        "id_interno": _pdf_valor_documento(
            resumo.get("ID Interno") or resumo.get("Identificador da Evid\u00eancia"),
            vazio="N\u00e3o informado",
        ),
    }


def _pdf_titulo_limpo(titulo):
    texto = _sanear_texto_exportacao_seguro(titulo or "").strip()
    if " - EVD-" in texto:
        return texto.split(" - EVD-", 1)[0].strip()
    return texto


def desenhar_cabecalho_pdf(y_topo, titulo, resumo, logo_pdf):
    metadados = _pdf_obter_metadados_relatorio(_pdf_titulo_limpo(titulo), resumo)
    x_caixa = 50
    largura_caixa = 495
    y_caixa = y_topo - 76
    altura_caixa = 72
    padding_x = 10
    padding_y = 8
    comandos = [
        "0.96 0.96 0.96 rg",
        "0 0 0 RG",
        f"{x_caixa} {y_caixa:.2f} {largura_caixa} {altura_caixa} re",
        "B",
    ]
    if logo_pdf:
        largura_logo = 125
        altura_logo = round((logo_pdf["height"] / logo_pdf["width"]) * largura_logo, 2)
        x_logo = x_caixa + largura_caixa - padding_x - largura_logo
        y_logo = y_caixa + altura_caixa - padding_y - altura_logo
        comandos.extend(
            [
                "q",
                f"{largura_logo} 0 0 {altura_logo} {x_logo:.2f} {y_logo:.2f} cm",
                "/Im1 Do",
                "Q",
            ]
        )
        largura_titulo = max(140, x_logo - (x_caixa + padding_x) - 14)
    else:
        comandos.extend(
            [
                "0.85 0.85 0.85 rg",
                "0 0 0 RG",
                f"{x_caixa + largura_caixa - padding_x - 125:.2f} {y_caixa + altura_caixa - padding_y - 28:.2f} 125 28 re",
                "B",
                *_pdf_text_commands(x_caixa + largura_caixa - padding_x - 105, y_caixa + altura_caixa - padding_y - 18, "LOGO", fonte="F2", tamanho=12),
            ]
        )
        largura_titulo = largura_caixa - (padding_x * 2) - 14
    titulo_cabecalho = _pdf_ajustar_texto_para_largura(metadados["relatorio"], largura_titulo, 11)
    comandos.extend(
        [
            *_pdf_text_commands(x_caixa + padding_x, y_caixa + altura_caixa - 18, titulo_cabecalho, fonte="F2", tamanho=11),
            *_pdf_text_commands(x_caixa + padding_x, y_caixa + altura_caixa - 34, f"C\u00f3digo: {metadados['codigo_documento']}", tamanho=8),
            *_pdf_text_commands(x_caixa + padding_x, y_caixa + altura_caixa - 50, f"Obra: {metadados['obra']}", tamanho=8),
            "0.75 0.75 0.75 RG",
            f"{x_caixa} {y_caixa - 4:.2f} m {x_caixa + largura_caixa} {y_caixa - 4:.2f} l S",
        ]
    )
    return comandos


def desenhar_rodape_pdf(numero_pagina, total_paginas, resumo, y_base=26):
    metadados = _pdf_obter_metadados_relatorio("", resumo)
    texto_geracao = f"Gerado por Construtask em {metadados['data_emissao']}"
    x_geracao = _pdf_x_texto_alinhado(225, 320, texto_geracao, 8, "right", 0)
    return [
        "0.75 0.75 0.75 RG",
        f"40 {y_base + 10:.2f} m 555 {y_base + 10:.2f} l S",
        *_pdf_text_commands(42, y_base, f"P\u00e1gina {numero_pagina} de {total_paginas}", tamanho=8),
        *_pdf_text_commands(x_geracao, y_base, texto_geracao, tamanho=8),
    ]


def desenhar_titulo_secao(y, titulo):
    return _pdf_section_title_commands(y, titulo)


def desenhar_bloco_informacoes(y_topo, titulo, resumo):
    linhas = [{"Campo": campo, "Valor": _pdf_valor_documento(valor)} for campo, valor in (resumo or {}).items()]
    return _pdf_table_commands(
        y_topo,
        [
            {"chave": "Campo", "titulo": "Campo", "largura": 160, "align": "left"},
            {"chave": "Valor", "titulo": "Valor", "largura": 335, "align": "left"},
        ],
        linhas,
        titulo=titulo,
    )


def desenhar_tabela_padrao(y_topo, titulo, colunas, linhas, *, max_linhas=None):
    linhas_normalizadas = _pdf_normalizar_linhas_documento(linhas, colunas)
    return _pdf_table_commands(y_topo, colunas, linhas_normalizadas, titulo=titulo, max_linhas=max_linhas)


def _pdf_section_title_commands(y, titulo):
    titulo_bruto = _sanear_texto_exportacao_seguro(titulo or "")
    titulo_base = titulo_bruto.lower()
    if "histor" in titulo_base and "aprov" in titulo_base:
        titulo_normalizado = "HIST\u00d3RICO DE APROVA\u00c7\u00c3O"
    elif "histor" in titulo_base and "aditivo" in titulo_base:
        titulo_normalizado = "HIST\u00d3RICO DOS ADITIVOS"
    else:
        titulo_normalizado = titulo_bruto.upper()
    for termo_antigo, termo_novo in {
        "HISTRICO": "HIST\u00d3RICO",
        "APROVAO": "APROVA\u00c7\u00c3O",
        "DESCRIO": "DESCRI\u00c7\u00c3O",
        "MEDIO": "MEDI\u00c7\u00c3O",
        "CONTRATAES": "CONTRATA\u00c7\u00d5ES",
    }.items():
        titulo_normalizado = titulo_normalizado.replace(termo_antigo, termo_novo)
    return [
        "0.87 0.87 0.87 rg",
        "0 0 0 RG",
        f"50 {y - 20:.2f} 495 20 re",
        "B",
        * _pdf_text_commands_color(56, y - 14, titulo_normalizado, fonte="F2", tamanho=10, rgb=(0, 0, 0)),
        "0 0 0 rg",
    ]


def _pdf_table_commands(y_topo, colunas, linhas, *, titulo=None, max_linhas=None):
    colunas = _pdf_normalizar_colunas(colunas)
    # Controla a construÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â§ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â£o de cada tabela no PDF: larguras, altura de linha, cabeÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â§alhos e conteÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âºdo.
    comandos = []
    y_atual = y_topo
    if titulo:
        comandos.extend(_pdf_section_title_commands(y_atual, titulo))
        y_atual -= 30

    if max_linhas is not None:
        linhas = list(linhas)[:max_linhas]

    altura_linha = 22
    padding_x = 5
    tamanho_fonte_cabecalho = 7
    tamanho_fonte_corpo = 7
    espacamento_linha = 9

    if not linhas:
        linhas = [{colunas[0]["chave"]: "-"}]

    header_wraps = []
    for coluna in colunas:
        header_wraps.append(
            _pdf_wrap_text(coluna["titulo"], coluna["largura"] - (padding_x * 2), tamanho_fonte_cabecalho)
        )
    header_height = max(20, 10 + (max(len(item) for item in header_wraps) * espacamento_linha))
    comandos.extend(
        [
            "0.90 0.90 0.90 rg",
            "0 0 0 RG",
            f"50 {y_atual - header_height:.2f} 495 {header_height} re",
            "B",
        ]
    )
    x_cursor = 50
    for indice_coluna, coluna in enumerate(colunas):
        largura = coluna["largura"]
        comandos.extend(
            [
                "0.90 0.90 0.90 rg",
                "0 0 0 RG",
                f"{x_cursor:.2f} {y_atual - header_height:.2f} {largura:.2f} {header_height} re",
                "B",
            ]
        )
        y_header = y_atual - 12
        for sublinha in header_wraps[indice_coluna]:
            comandos.extend(
                _pdf_text_commands_color(
                    _pdf_x_texto_alinhado(
                        x_cursor,
                        largura,
                        sublinha,
                        tamanho_fonte_cabecalho,
                        coluna.get("align", "left"),
                        padding_x,
                    ),
                    y_header,
                    sublinha,
                    fonte="F2",
                    tamanho=tamanho_fonte_cabecalho,
                    rgb=(0, 0, 0),
                )
            )
            y_header -= espacamento_linha
        x_cursor += largura
    y_atual -= header_height

    for linha in linhas:
        alturas = []
        conteudos = []
        for coluna in colunas:
            valor = _pdf_valor_documento(linha.get(coluna["chave"]))
            quebrado = _pdf_wrap_text(valor, coluna["largura"] - (padding_x * 2), tamanho_fonte_corpo)
            conteudos.append(quebrado)
            alturas.append(max(altura_linha, 10 + (len(quebrado) * espacamento_linha)))
        altura = max(alturas)
        x_cursor = 50
        for indice, coluna in enumerate(colunas):
            largura = coluna["largura"]
            comandos.append(f"{x_cursor:.2f} {y_atual - altura:.2f} {largura:.2f} {altura:.2f} re")
            comandos.append("S")
            y_texto = y_atual - 13
            for sublinha in conteudos[indice]:
                comandos.extend(
                    _pdf_text_commands(
                        _pdf_x_texto_alinhado(
                            x_cursor,
                            largura,
                            sublinha,
                            tamanho_fonte_corpo,
                            coluna.get("align", "left"),
                            padding_x,
                        ),
                        y_texto,
                        sublinha,
                        fonte="F1",
                        tamanho=tamanho_fonte_corpo,
                    )
                )
                y_texto -= espacamento_linha
            x_cursor += largura
        y_atual -= altura

    return comandos, y_atual


def _pdf_estimar_altura_tabela(colunas, linhas, *, titulo=None):
    colunas = _pdf_normalizar_colunas(colunas)
    padding_x = 5
    tamanho_fonte_cabecalho = 7
    tamanho_fonte_corpo = 7
    espacamento_linha = 9
    altura_total = 0
    if titulo:
        altura_total += 30
    header_wraps = [
        _pdf_wrap_text(coluna["titulo"], coluna["largura"] - (padding_x * 2), tamanho_fonte_cabecalho)
        for coluna in colunas
    ]
    altura_total += max(20, 10 + (max(len(item) for item in header_wraps) * espacamento_linha))
    for linha in linhas:
        alturas = []
        for coluna in colunas:
            quebrado = _pdf_wrap_text(
                _pdf_valor_documento(linha.get(coluna["chave"])),
                coluna["largura"] - (padding_x * 2),
                tamanho_fonte_corpo,
            )
            alturas.append(max(22, 10 + (len(quebrado) * espacamento_linha)))
        altura_total += max(alturas)
    return altura_total


def _pdf_relatorio_tabelas_response(nome_arquivo, titulo, resumo, secoes):
    logo_pdf = _carregar_png_para_pdf(_PDF_LOGO_PATH)
    altura_pagina = 842
    margem_superior = 40
    margem_inferior = 60
    y_inicial_pagina = altura_pagina - margem_superior - 102
    paginas = []

    def _novo_conteudo_pagina():
        comandos_pagina = desenhar_cabecalho_pdf(790, titulo, resumo, logo_pdf)
        return comandos_pagina, y_inicial_pagina

    def _fechar_pagina(comandos_pagina):
        paginas.append(comandos_pagina)

    conteudo, y_atual = _novo_conteudo_pagina()
    comandos_resumo, y_atual = desenhar_bloco_informacoes(y_atual, "RESUMO", resumo)
    conteudo.extend(comandos_resumo)

    for secao in secoes:
        colunas_secao = _pdf_normalizar_colunas(secao["colunas"])
        linhas = list(secao["linhas"] or [{colunas_secao[0]["chave"]: "-"}])
        max_linhas = secao.get("max_linhas")
        if max_linhas is not None:
            linhas = linhas[:max_linhas]
        primeiro_bloco = True
        while linhas:
            y_bloco = y_atual - 18
            titulo_secao = secao["titulo"] if primeiro_bloco else f'{secao["titulo"]} (continua)'
            linhas_bloco = []
            for linha in linhas:
                candidato = linhas_bloco + [linha]
                altura_prevista = _pdf_estimar_altura_tabela(colunas_secao, candidato, titulo=titulo_secao)
                if y_bloco - altura_prevista < margem_inferior:
                    break
                linhas_bloco = candidato
            if not linhas_bloco:
                _fechar_pagina(conteudo)
                conteudo, y_atual = _novo_conteudo_pagina()
                continue
            comandos_secao, y_atual = desenhar_tabela_padrao(y_bloco, titulo_secao, colunas_secao, linhas_bloco)
            conteudo.extend(comandos_secao)
            linhas = linhas[len(linhas_bloco):]
            primeiro_bloco = False
            if linhas:
                _fechar_pagina(conteudo)
                conteudo, y_atual = _novo_conteudo_pagina()

    _fechar_pagina(conteudo)

    page_count = len(paginas)
    paginas_stream = []
    for indice_pagina, comandos_pagina in enumerate(paginas, start=1):
        comandos_pagina.extend(desenhar_rodape_pdf(indice_pagina, page_count, resumo))
        paginas_stream.append("\n".join(comandos_pagina).encode("cp1252", "replace"))
    font1_id = 3 + (page_count * 2)
    font2_id = font1_id + 1
    image_id = font2_id + 1 if logo_pdf else None
    alpha_image_id = image_id + 1 if logo_pdf and logo_pdf.get("alpha_stream") else None

    recursos_pagina = f"/Font << /F1 {font1_id} 0 R /F2 {font2_id} 0 R >>".encode("ascii")
    if image_id:
        recursos_pagina += f" /XObject << /Im1 {image_id} 0 R >>".encode("ascii")

    objetos = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        (
            b"2 0 obj << /Type /Pages /Kids ["
            + b" ".join(f"{3 + (indice * 2)} 0 R".encode("ascii") for indice in range(page_count))
            + b"] /Count "
            + str(page_count).encode("ascii")
            + b" >> endobj\n"
        ),
    ]
    for indice_pagina, stream in enumerate(paginas_stream):
        page_obj_id = 3 + (indice_pagina * 2)
        content_obj_id = page_obj_id + 1
        objetos.append(
            f"{page_obj_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << ".encode("ascii")
            + recursos_pagina
            + f" >> /Contents {content_obj_id} 0 R >> endobj\n".encode("ascii")
        )
        objetos.append(
            f"{content_obj_id} 0 obj << /Length {len(stream)} >> stream\n".encode("ascii")
            + stream
            + b"\nendstream endobj\n"
        )
    objetos.extend(
        [
            f"{font1_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >> endobj\n".encode("ascii"),
            f"{font2_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >> endobj\n".encode("ascii"),
        ]
    )
    if logo_pdf and image_id:
        objetos.append(
            f"{image_id} 0 obj << /Type /XObject /Subtype /Image /Width ".encode("ascii")
            + str(logo_pdf["width"]).encode("ascii")
            + b" /Height "
            + str(logo_pdf["height"]).encode("ascii")
            + b" /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length "
            + str(len(logo_pdf["stream"])).encode("ascii")
            + (f" /SMask {alpha_image_id} 0 R".encode("ascii") if alpha_image_id else b"")
            + b" >> stream\n"
            + logo_pdf["stream"]
            + b"\nendstream endobj\n"
        )
    if logo_pdf and alpha_image_id:
        objetos.append(
            f"{alpha_image_id} 0 obj << /Type /XObject /Subtype /Image /Width ".encode("ascii")
            + str(logo_pdf["width"]).encode("ascii")
            + b" /Height "
            + str(logo_pdf["height"]).encode("ascii")
            + b" /ColorSpace /DeviceGray /BitsPerComponent 8 /Filter /FlateDecode /Length "
            + str(len(logo_pdf["alpha_stream"])).encode("ascii")
            + b" >> stream\n"
            + logo_pdf["alpha_stream"]
            + b"\nendstream endobj\n"
        )

    pdf = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objetos:
        offsets.append(len(pdf))
        pdf += obj
    xref = len(pdf)
    pdf += f"xref\n0 {len(offsets)}\n".encode("ascii")
    pdf += b"0000000000 65535 f \n"
    for offset in offsets[1:]:
        pdf += f"{offset:010d} 00000 n \n".encode("ascii")
    pdf += f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode("ascii")

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _pdf_relatorio_probatorio_response(
    nome_arquivo,
    titulo,
    resumo,
    historico,
    extras,
    *,
    extras_titulo,
    extras_colunas,
    incluir_historico=True,
    extras_max_linhas=None,
    secoes_extras=None,
):
    secoes = []
    if incluir_historico:
        secoes.append(
            {
                "titulo": "Hist\u00f3rico de Aprova\u00e7\u00e3o",
                "colunas": [("Data", 75), ("A\u00e7\u00e3o", 75), ("Usu\u00e1rio", 105), ("Descri\u00e7\u00e3o", 240)],
                "linhas": [
                    {
                        "Data": linha.get("Data", "-"),
                        "A\u00e7\u00e3o": linha.get("Acao", linha.get("A\u00e7\u00e3o", "-")),
                        "Usu\u00e1rio": linha.get("Usuario", linha.get("Usu\u00e1rio", "-")),
                        "Descri\u00e7\u00e3o": linha.get("Descricao", linha.get("Descri\u00e7\u00e3o", "-")),
                    }
                    for linha in (historico or [{"Data": "-", "Acao": "-", "Usuario": "-", "Descricao": "-"}])
                ],
            }
        )
    secoes.append(
        {
            "titulo": extras_titulo,
            "colunas": extras_colunas,
            "linhas": extras or [{extras_colunas[0][0]: "-"}],
            **({"max_linhas": extras_max_linhas} if extras_max_linhas is not None else {}),
        }
    )
    for secao_extra in secoes_extras or []:
        secoes.append(secao_extra)
    return _pdf_relatorio_tabelas_response(nome_arquivo, titulo, resumo, secoes)


def _pdf_simples_response(nome_arquivo, titulo, linhas):
    resumo = []
    for linha in list(linhas):
        texto = str(linha)
        if ":" in texto:
            campo, valor = texto.split(":", 1)
            resumo.append({"Campo": campo.strip(), "Valor": valor.strip() or "-"})
        else:
            resumo.append({"Campo": "Informação", "Valor": texto})

    return _pdf_relatorio_probatorio_response(
        nome_arquivo,
        titulo,
        {"Documento": titulo, "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")},
        [],
        resumo,
        extras_titulo="Dados Exportados",
        extras_colunas=[("Campo", 165), ("Valor", 330)],
    )


def _exportar_relatorio_probatorio_excel_response(
    nome_arquivo,
    sheet_resumo,
    resumo,
    historico_linhas,
    *,
    extras_sheet_name=None,
    extras_linhas=None,
):
    output = BytesIO()
    resumo_linhas = _normalizar_linhas_exportacao([{"Campo": chave, "Valor": valor} for chave, valor in resumo.items()])
    historico_normalizado = _normalizar_linhas_exportacao(
        historico_linhas or [{"Data": "-", "Acao": "-", "Usuario": "-", "Descricao": "-"}]
    )
    extras_normalizados = _normalizar_linhas_exportacao(extras_linhas or [{"Informação": "-"}])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(resumo_linhas).to_excel(writer, index=False, sheet_name=sheet_resumo)
        _aplicar_layout_excel_relatorio(writer.book[sheet_resumo], "RELATÓRIO PROBATÓRIO DE APROVAÇÃO", sheet_resumo)
        pd.DataFrame(historico_normalizado).to_excel(
            writer,
            index=False,
            sheet_name="Histórico",
        )
        _aplicar_layout_excel_relatorio(writer.book["Histórico"], "RELATÓRIO PROBATÓRIO DE APROVAÇÃO", "Histórico")
        if extras_sheet_name:
            pd.DataFrame(extras_normalizados).to_excel(
                writer,
                index=False,
                sheet_name=extras_sheet_name,
            )
            _aplicar_layout_excel_relatorio(
                writer.book[extras_sheet_name],
                "RELATÓRIO PROBATÓRIO DE APROVAÇÃO",
                extras_sheet_name,
            )
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _formatar_usuario_data(usuario, datahora):
    if not usuario and not datahora:
        return "-"
    usuario_label = getattr(usuario, "username", None) or "-"
    data_local = _datahora_local(datahora)
    data_label = data_local.strftime("%d/%m/%Y %H:%M") if data_local else "-"
    return f"{usuario_label} em {data_label}"


def _identificador_evidencia(tipo, objeto, data_referencia):
    data_base = _datahora_local(data_referencia) or _datahora_local(timezone.now())
    return f"EVD-{tipo}-{objeto.pk:06d}-{data_base:%Y%m%d%H%M}"


def _historico_probatorio(historicos_queryset):
    linhas = []
    for item in historicos_queryset.order_by("-criado_em")[:12]:
        data_local = _datahora_local(item.criado_em)
        linhas.append(
            {
                "Data": data_local.strftime("%d/%m/%Y %H:%M") if data_local else "-",
                "Acao": item.acao,
                "Usuario": getattr(item.usuario, "username", "-") or "-",
                "Descricao": item.descricao,
            }
        )
    return linhas


def _linhas_aditivos_compromisso(compromisso):
    linhas = []
    for aditivo in compromisso.aditivos.all():
        total_aditivo = aditivo.itens.aggregate(total=Sum("valor"))["total"] or Decimal("0.00")
        centros = ", ".join(
            f"{item.centro_custo.codigo} - {item.centro_custo.descricao}"
            for item in aditivo.itens.select_related("centro_custo").all()
        )
        if aditivo.tipo == "PRAZO":
            impacto = f"{aditivo.delta_dias or 0:+d} dias"
        else:
            impacto = money_br(total_aditivo) if total_aditivo else "Não informado"

        linhas.append(
            {
                "Tipo": aditivo.get_tipo_display(),
                "Status": aditivo.get_status_display(),
                "Descrição": aditivo.descricao or "Não informado",
                "Impacto": impacto,
                "Centro de Custo": centros or "Não informado",
                "Data": _datahora_local(aditivo.criado_em).strftime("%d/%m/%Y") if _datahora_local(aditivo.criado_em) else "Não informado",
                "Motivo": aditivo.motivo_mudanca or "Não informado",
            }
        )
    return linhas


def _historico_aditivos_compromisso(compromisso):
    linhas = []
    for item in compromisso.historicos.filter(Q(descricao__icontains="Aditivo") | Q(acao__startswith="ADITIVO")).order_by("-criado_em"):
        data_local = _datahora_local(item.criado_em)
        if item.acao == "ADITIVO_ENVIO":
            acao_label = "Enviado para Aprova\u00e7\u00e3o"
        elif item.acao == "ADITIVO_APROVADO":
            acao_label = "Aprovado"
        elif item.acao == "ADITIVO_AJUSTE":
            acao_label = "Devolvido para Ajuste"
        elif item.acao == "ADITIVO":
            acao_label = "Inclu\u00eddo"
        else:
            descricao_normalizada = _sanear_texto_exportacao_seguro(item.descricao or "")
            if "devolvido para ajuste" in descricao_normalizada.lower():
                acao_label = "Devolvido para Ajuste"
            elif "aprovad" in descricao_normalizada.lower():
                acao_label = "Aprovado"
            elif "enviado para aprovacao" in descricao_normalizada.lower():
                acao_label = "Enviado para Aprova\u00e7\u00e3o"
            else:
                acao_label = _sanear_texto_exportacao_seguro(item.acao or "N\u00e3o informado")
        linhas.append(
            {
                "Data": data_local.strftime("%d/%m/%Y %H:%M") if data_local else "N\u00e3o informado",
                "Aditivo": compromisso.numero,
                "A\u00e7\u00e3o": acao_label,
                "Usu\u00e1rio": _sanear_texto_exportacao_seguro(getattr(item.usuario, "username", "-") or "-"),
                "Descri\u00e7\u00e3o": _sanear_texto_exportacao_seguro(item.descricao or "N\u00e3o informado"),
            }
        )
    if linhas:
        return linhas

    for aditivo in compromisso.aditivos.all().order_by("-criado_em"):
        if aditivo.status == "APROVADO":
            acao_label = "Aprovado"
        elif aditivo.status == "EM_APROVACAO":
            acao_label = "Enviado para Aprova\u00e7\u00e3o"
        elif aditivo.parecer_aprovacao:
            acao_label = "Devolvido para Ajuste"
        else:
            acao_label = "Inclu\u00eddo"
        usuario = getattr(aditivo.aprovado_por, "username", None) or getattr(aditivo.enviado_para_aprovacao_por, "username", None) or "N\u00e3o informado"
        data_referencia = aditivo.aprovado_em or aditivo.enviado_para_aprovacao_em or aditivo.criado_em
        data_local = _datahora_local(data_referencia)
        linhas.append(
            {
                "Data": data_local.strftime("%d/%m/%Y %H:%M") if data_local else "N\u00e3o informado",
                "Aditivo": compromisso.numero,
                "A\u00e7\u00e3o": acao_label,
                "Usu\u00e1rio": _sanear_texto_exportacao_seguro(usuario),
                "Descri\u00e7\u00e3o": _sanear_texto_exportacao_seguro(aditivo.parecer_aprovacao or aditivo.descricao or "N\u00e3o informado"),
            }
        )
    return linhas


def _dados_relatorio_aprovacao_compromisso(compromisso):
    evidencia_id = _identificador_evidencia(
        "COMPROMISSO",
        compromisso,
        compromisso.aprovado_em or compromisso.enviado_para_aprovacao_em or compromisso.criado_em,
    )
    resumo = {
        "Identificador da Evid\u00eancia": evidencia_id,
        "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        "N\u00famero": compromisso.numero,
        "Tipo": compromisso.get_tipo_display(),
        "Obra": f"{compromisso.obra.codigo if compromisso.obra else '-'} - {compromisso.obra.nome if compromisso.obra else '-'}",
        "Status atual": compromisso.get_status_display(),
        "Valor": money_br(compromisso.valor_contratado),
        "Fornecedor": compromisso.fornecedor,
        "Respons\u00e1vel operacional": compromisso.responsavel,
        "Enviado para aprova\u00e7\u00e3o por": _formatar_usuario_data(compromisso.enviado_para_aprovacao_por, compromisso.enviado_para_aprovacao_em),
        "Aprovado por": _formatar_usuario_data(compromisso.aprovado_por, compromisso.aprovado_em),
        "Parecer": compromisso.parecer_aprovacao or "-",
    }
    historico = _historico_probatorio(compromisso.historicos)
    extras = [
        {
            "Centro de Custo": f"{item.centro_custo.codigo} - {item.centro_custo.descricao}",
            "Descri\u00e7\u00e3o do Item": item.descricao_tecnica or item.centro_custo.descricao,
            "Quantidade": item.quantidade,
            "Unidade": item.unidade or "-",
            "Valor Unit\u00e1rio": money_br(item.valor_unitario),
            "Valor Total": money_br(item.valor_total),
        }
        for item in compromisso.itens.all()
    ]
    secoes_extras = []
    if compromisso.tipo == "CONTRATO":
        secoes_extras.append(
            {
                "titulo": "Aditivos do Contrato",
                "colunas": [
                    ("Tipo", 55),
                    ("Status", 65),
                    ("Descri\u00e7\u00e3o", 140),
                    ("Impacto", 70),
                    ("Centro de Custo", 110),
                    ("Data", 55),
                ],
                "linhas": _linhas_aditivos_compromisso(compromisso)
                or [{"Tipo": "Nenhum registro encontrado", "Status": "", "Descri\u00e7\u00e3o": "", "Impacto": "", "Centro de Custo": "", "Data": ""}],
            }
        )
        secoes_extras.append(
            {
                "titulo": "Hist\u00f3rico dos Aditivos",
                "colunas": [
                    ("Data", 75),
                    ("Aditivo", 75),
                    ("A\u00e7\u00e3o", 80),
                    ("Usu\u00e1rio", 85),
                    ("Descri\u00e7\u00e3o", 180),
                ],
                "linhas": _historico_aditivos_compromisso(compromisso)
                or [{"Data": "Nenhum registro encontrado", "Aditivo": "", "A\u00e7\u00e3o": "", "Usu\u00e1rio": "", "Descri\u00e7\u00e3o": ""}],
            }
        )
    linhas_pdf = [f"{campo}: {valor}" for campo, valor in resumo.items()] + ["", "Hist\u00f3rico recente:"] + [
        f"- {linha['Data']} | {linha['Acao']} | {linha['Usuario']} | {linha['Descricao']}" for linha in historico
    ]
    return evidencia_id, resumo, historico, extras, linhas_pdf, secoes_extras


def _dados_relatorio_aprovacao_medicao(medicao):
    evidencia_id = _identificador_evidencia(
        "MEDICAO",
        medicao,
        medicao.aprovado_em or medicao.enviado_para_aprovacao_em or medicao.criado_em,
    )
    resumo = {
        "Identificador da Evidencia": evidencia_id,
        "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        "Numero": medicao.numero_da_medicao,
        "Contrato": medicao.contrato.numero,
        "Obra": f"{medicao.obra.codigo if medicao.obra else '-'} - {medicao.obra.nome if medicao.obra else '-'}",
        "Status atual": medicao.get_status_display(),
        "Valor medido": money_br(medicao.valor_medido),
        "Fornecedor": medicao.fornecedor,
        "Responsável operacional": medicao.responsavel,
        "Enviado para aprovação por": _formatar_usuario_data(medicao.enviado_para_aprovacao_por, medicao.enviado_para_aprovacao_em),
        "Aprovado por": _formatar_usuario_data(medicao.aprovado_por, medicao.aprovado_em),
        "Parecer": medicao.parecer_aprovacao or "-",
    }
    historico = _historico_probatorio(medicao.historicos)
    extras = [
        {
            "Centro de Custo": f"{item.centro_custo.codigo} - {item.centro_custo.descricao}",
            "Quantidade": item.quantidade,
            "Unidade": item.unidade or "-",
            "Valor Total": money_br(item.valor_total),
        }
        for item in medicao.itens.all()
    ]
    linhas_pdf = [f"{campo}: {valor}" for campo, valor in resumo.items()] + ["", "Histórico recente:"] + [
        f"- {linha['Data']} | {linha['Acao']} | {linha['Usuario']} | {linha['Descricao']}" for linha in historico
    ]
    return evidencia_id, resumo, historico, extras, linhas_pdf


def _dados_relatorio_aprovacao_baseline(baseline):
    evidencia_id = _identificador_evidencia(
        "BASELINE",
        baseline,
        baseline.aprovado_em or baseline.enviado_para_aprovacao_em or baseline.criado_em,
    )
    resumo = {
        "Identificador da Evidência": evidencia_id,
        "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        "Descricao": baseline.descricao,
        "Obra": f"{baseline.obra.codigo if baseline.obra else '-'} - {baseline.obra.nome if baseline.obra else '-'}",
        "Status atual": baseline.get_status_display(),
        "Versao ativa da obra": "Sim" if baseline.is_ativa else "Nao",
        "Valor baseline": money_br(baseline.valor_total),
        "Criado por": _formatar_usuario_data(baseline.criado_por, baseline.criado_em),
        "Enviado para aprovação por": _formatar_usuario_data(baseline.enviado_para_aprovacao_por, baseline.enviado_para_aprovacao_em),
        "Aprovado por": _formatar_usuario_data(baseline.aprovado_por, baseline.aprovado_em),
        "Parecer": baseline.parecer_aprovacao or "-",
    }
    historico = []
    extras = [
        {
            "Codigo": item.codigo,
            "Descricao": item.descricao,
            "Nivel": item.level,
            "Valor Consolidado": money_br(item.valor_total_consolidado),
        }
        for item in baseline.itens.filter(level__lte=2)[:200]
    ]
    linhas_pdf = [f"{campo}: {valor}" for campo, valor in resumo.items()] + ["", "Resumo de itens snapshot:"] + [
        f"- {item['Codigo']} | {item['Descricao']} | {item['Valor Consolidado']}" for item in extras[:20]
    ]
    return evidencia_id, resumo, historico, extras, linhas_pdf


def _listar_evidencias_obra(request, obra_contexto, *, tipo=""):
    evidencias = []
    if not obra_contexto:
        return evidencias

    contratos = _filtrar_por_obra_contexto(
        request,
        Compromisso.objects.select_related("obra", "aprovado_por", "enviado_para_aprovacao_por").filter(
            Q(enviado_para_aprovacao_em__isnull=False)
            | Q(aprovado_em__isnull=False)
            | Q(enviado_para_aprovacao_por__isnull=False)
            | Q(aprovado_por__isnull=False)
        ),
    )
    medicoes = _filtrar_por_obra_contexto(
        request,
        Medicao.objects.select_related("obra", "contrato", "aprovado_por", "enviado_para_aprovacao_por").filter(
            Q(enviado_para_aprovacao_em__isnull=False)
            | Q(aprovado_em__isnull=False)
            | Q(enviado_para_aprovacao_por__isnull=False)
            | Q(aprovado_por__isnull=False)
        ),
    )
    baselines = _filtrar_por_obra_contexto(
        request,
        OrcamentoBaseline.objects.select_related("obra", "aprovado_por", "enviado_para_aprovacao_por").filter(
            Q(enviado_para_aprovacao_em__isnull=False)
            | Q(aprovado_em__isnull=False)
            | Q(enviado_para_aprovacao_por__isnull=False)
            | Q(aprovado_por__isnull=False)
        ),
    )

    if tipo in ("", "contrato"):
        for contrato in contratos.order_by("-aprovado_em", "-enviado_para_aprovacao_em", "-criado_em"):
            evidencia_id, resumo, _, _, _, _ = _dados_relatorio_aprovacao_compromisso(contrato)
            evidencias.append(
                {
                    "tipo": "Contrato",
                    "identificador": evidencia_id,
                    "referencia": contrato.numero,
                    "status": resumo["Status atual"],
                    "valor": resumo["Valor"],
                    "responsavel": resumo["Aprovado por"],
                    "parecer": resumo["Parecer"],
                    "pdf_url": reverse_lazy("compromisso_aprovacao_pdf", args=[contrato.pk]),
                    "excel_url": reverse_lazy("compromisso_aprovacao_excel", args=[contrato.pk]),
                }
            )

    if tipo in ("", "medicao"):
        for medicao in medicoes.order_by("-aprovado_em", "-enviado_para_aprovacao_em", "-criado_em"):
            evidencia_id, resumo, _, _, _ = _dados_relatorio_aprovacao_medicao(medicao)
            evidencias.append(
                    {
                        "tipo": "Medição",
                    "identificador": evidencia_id,
                    "referencia": medicao.numero_da_medicao,
                    "status": resumo["Status atual"],
                    "valor": resumo["Valor medido"],
                    "responsavel": resumo["Aprovado por"],
                    "parecer": resumo["Parecer"],
                    "pdf_url": reverse_lazy("medicao_aprovacao_pdf", args=[medicao.pk]),
                    "excel_url": reverse_lazy("medicao_aprovacao_excel", args=[medicao.pk]),
                }
            )

    if tipo in ("", "baseline"):
        for baseline in baselines.order_by("-aprovado_em", "-enviado_para_aprovacao_em", "-criado_em"):
            evidencia_id, resumo, _, _, _ = _dados_relatorio_aprovacao_baseline(baseline)
            evidencias.append(
                {
                    "tipo": "Baseline",
                    "identificador": evidencia_id,
                    "referencia": baseline.descricao,
                    "status": resumo["Status atual"],
                    "valor": resumo["Valor baseline"],
                    "responsavel": resumo["Aprovado por"],
                    "parecer": resumo["Parecer"],
                    "pdf_url": reverse_lazy("plano_contas_baseline_aprovacao_pdf", args=[baseline.pk]),
                    "excel_url": reverse_lazy("plano_contas_baseline_aprovacao_excel", args=[baseline.pk]),
                }
            )

    return evidencias


class CentralEvidenciasView(TemplateView):
    template_name = "app/central_evidencias.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        tipo = self.request.GET.get("tipo", "").strip()
        evidencias = _listar_evidencias_obra(self.request, obra_contexto, tipo=tipo)
        context["obra_contexto"] = obra_contexto
        context["tipo_filtro"] = tipo
        context["evidencias"] = evidencias
        return context


def _dados_dossie_obra(request):
    obra_contexto = _obter_obra_contexto(request)
    dados = {
        "obra_contexto": obra_contexto,
        "analise_plano": [],
        "evidencias": [],
        "dados_cabecalho": {
            "cliente": "-",
            "obra": "-",
            "endereco": "-",
            "responsavel": "-",
        },
        "relatorio_gerencial": {
            "valor_orcado": Decimal("0.00"),
            "valor_comprometido": Decimal("0.00"),
            "valor_medido": Decimal("0.00"),
            "valor_executado": Decimal("0.00"),
            "percentual_comprometido": Decimal("0.00"),
            "percentual_medido": Decimal("0.00"),
            "percentual_executado": Decimal("0.00"),
            "saldo_a_comprometer": Decimal("0.00"),
            "saldo_a_executar": Decimal("0.00"),
        },
    }
    if not obra_contexto:
        return dados

    folhas = PlanoContas.objects.annotate(filhos_count=Count("filhos")).filter(filhos_count=0, obra=obra_contexto)
    valor_orcado = folhas.aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00")
    valor_comprometido = Compromisso.objects.filter(obra=obra_contexto).aggregate(total=Sum("valor_contratado"))["total"] or Decimal("0.00")
    valor_medido = Medicao.objects.filter(obra=obra_contexto).aggregate(total=Sum("valor_medido"))["total"] or Decimal("0.00")
    valor_executado = NotaFiscal.objects.filter(obra=obra_contexto).aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00")

    dados["relatorio_gerencial"] = {
        "valor_orcado": valor_orcado,
        "valor_comprometido": valor_comprometido,
        "valor_medido": valor_medido,
        "valor_executado": valor_executado,
        "percentual_comprometido": _calcular_percentual(valor_comprometido, valor_orcado),
        "percentual_medido": _calcular_percentual(valor_medido, valor_orcado),
        "percentual_executado": _calcular_percentual(valor_executado, valor_orcado),
        "saldo_a_comprometer": arredondar_moeda(valor_orcado - valor_comprometido),
        "saldo_a_executar": arredondar_moeda(valor_orcado - valor_executado),
    }

    planos = _consolidar_plano_contas(
        PlanoContas.objects.filter(obra=obra_contexto, level=4).annotate(filhos_count=Count("filhos")).order_by("tree_id", "lft")
    )
    dados["analise_plano"] = [
        {
            "indice": f"{indice:02d}",
            "codigo": plano.codigo,
            "descricao": plano.descricao,
            "valor_orcado": plano.valor_total_consolidado_calc,
            "valor_contratado": plano.valor_comprometido_calc,
            "valor_pago": plano.valor_executado_calc,
            "saldo_a_pagar": arredondar_moeda(plano.valor_comprometido_calc - plano.valor_executado_calc),
        }
        for indice, plano in enumerate(planos[:14], start=1)
    ]
    dados["evidencias"] = _listar_evidencias_obra(request, obra_contexto)[:12]
    dados["dados_cabecalho"] = {
        "cliente": obra_contexto.cliente or "-",
        "obra": obra_contexto.nome or "-",
        "endereco": "-",
        "responsavel": obra_contexto.responsavel or "-",
    }
    return dados


class DossieObraView(TemplateView):
    template_name = "app/dossie_obra.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_dados_dossie_obra(self.request))
        return context


@login_required
def dossie_obra_pdf_view(request):
    dados = _dados_dossie_obra(request)
    obra_contexto = dados["obra_contexto"]
    resumo = {
        "Cliente": dados["dados_cabecalho"]["cliente"],
        "Obra": dados["dados_cabecalho"]["obra"],
        "Responsavel": dados["dados_cabecalho"]["responsavel"],
        "Valor Total Orcado": money_br(dados["relatorio_gerencial"]["valor_orcado"]),
        "Valor Total Comprometido": money_br(dados["relatorio_gerencial"]["valor_comprometido"]),
        "Valor Total Medido": money_br(dados["relatorio_gerencial"]["valor_medido"]),
        "Valor Total Executado": money_br(dados["relatorio_gerencial"]["valor_executado"]),
        "Saldo a Comprometer": money_br(dados["relatorio_gerencial"]["saldo_a_comprometer"]),
        "Saldo a Executar": money_br(dados["relatorio_gerencial"]["saldo_a_executar"]),
    }
    secoes = [
        {
            "titulo": "Analise de Plano de Contas",
            "colunas": [
                ("Indice", 40),
                ("Codigo", 65),
                ("Descricao", 170),
                ("Valor Orcado", 70),
                ("Valor Contratado", 80),
                ("Valor Pago", 70),
            ],
            "linhas": [
                {
                    "Indice": item["indice"],
                    "Codigo": item["codigo"],
                    "Descricao": item["descricao"],
                    "Valor Orcado": money_br(item["valor_orcado"]),
                    "Valor Contratado": money_br(item["valor_contratado"]),
                    "Valor Pago": money_br(item["valor_pago"]),
                }
                for item in dados["analise_plano"]
            ] or [{"Indice": "-", "Codigo": "-", "Descricao": "-", "Valor Orcado": "-", "Valor Contratado": "-", "Valor Pago": "-"}],
        },
        {
            "titulo": "Evidencias de Aprovacao",
            "colunas": [
                ("Tipo", 60),
                ("Identificador", 110),
                ("Referencia", 120),
                ("Status", 70),
                ("Valor", 60),
                ("Responsavel", 75),
            ],
            "linhas": [
                {
                    "Tipo": evidencia["tipo"],
                    "Identificador": evidencia["identificador"],
                    "Referencia": evidencia["referencia"],
                    "Status": evidencia["status"],
                    "Valor": evidencia["valor"] or "-",
                    "Responsavel": evidencia["responsavel"] or "-",
                }
                for evidencia in dados["evidencias"]
            ] or [{"Tipo": "-", "Identificador": "-", "Referencia": "-", "Status": "-", "Valor": "-", "Responsavel": "-"}],
        },
    ]
    nome_arquivo = f"dossie_obra_{obra_contexto.codigo if obra_contexto else 'sem_obra'}.pdf"
    titulo = f"Dossie da Obra - {obra_contexto.codigo if obra_contexto else 'Sem Obra'}"
    return _pdf_relatorio_tabelas_response(nome_arquivo, titulo, resumo, secoes)


def _apagar_objeto(request, queryset, success_url):
    if request.method != "POST":
        raise Http404()
    objeto_id = _coletar_post_int(request, "id")
    if not objeto_id:
        messages.error(request, "Nao foi possivel identificar o registro para exclusao.")
        return redirect(success_url)
    objeto = queryset.filter(pk=objeto_id).first()
    if not objeto:
        messages.error(request, "Nao foi possivel identificar o registro para exclusao.")
        return redirect(success_url)
    try:
        historico = _registrar_historico("EXCLUSAO", objeto, f"Exclusao de {objeto}", getattr(request, "user", None))
        objeto.delete()
        messages.success(request, "Registro excluido com sucesso.")
    except ProtectedError:
        # Se a exclusao for protegida, removemos o historico criado para nao registrar operacao inexistente.
        try:
            if "historico" in locals() and historico and getattr(historico, "pk", None):
                historico.delete()
        except Exception:
            pass
        messages.error(
            request,
            "Este registro nao pode ser excluido porque possui vinculos em outras operacoes do sistema.",
        )
    return redirect(success_url)


def _registrar_historico(acao, objeto, descricao, usuario=None):
    payload = {"acao": acao, "descricao": descricao, "usuario": usuario}
    if isinstance(objeto, Obra):
        if getattr(objeto, "pk", None):
            payload["obra_id"] = objeto.pk
        else:
            payload["obra"] = objeto
    elif isinstance(objeto, Compromisso):
        if getattr(objeto, "obra_id", None):
            payload["obra_id"] = objeto.obra_id
        if getattr(objeto, "pk", None):
            payload["compromisso_id"] = objeto.pk
        else:
            payload["compromisso"] = objeto
    elif isinstance(objeto, Medicao):
        if getattr(objeto, "obra_id", None):
            payload["obra_id"] = objeto.obra_id
        if getattr(objeto, "pk", None):
            payload["medicao_id"] = objeto.pk
        else:
            payload["medicao"] = objeto
    elif isinstance(objeto, NotaFiscal):
        if getattr(objeto, "obra_id", None):
            payload["obra_id"] = objeto.obra_id
        if getattr(objeto, "pk", None):
            payload["nota_fiscal_id"] = objeto.pk
        else:
            payload["nota_fiscal"] = objeto
    return HistoricoOperacional.objects.create(**payload)


@login_required
def selecionar_obra_contexto_view(request):
    proxima_url = request.POST.get("next") or reverse_lazy("home")
    obra_id = request.POST.get("obra_contexto")
    if obra_id:
        obra = _get_obras_permitidas(request.user).filter(pk=obra_id).first()
        if obra:
            request.session["obra_contexto_id"] = obra.pk
        else:
            messages.error(request, "Voce nao tem acesso a obra selecionada.")
    else:
        request.session.pop("obra_contexto_id", None)
    return redirect(proxima_url)


class HomeView(TemplateView):
    template_name = "app/home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        
        # Se nao tem obra selecionada, nao calcula indicadores (dados vao ficar zerados)
        if not obra_contexto:
            context["sem_obra_selecionada"] = True
            context["cronograma_ativo_url"] = reverse("plano_fisico_dashboard")
            context["indicadores"] = {
                "valor_orcado": Decimal("0.00"),
                "valor_comprometido": Decimal("0.00"),
                "valor_medido": Decimal("0.00"),
                "valor_pago": Decimal("0.00"),
            }
            context["indicadores_exec"] = {
                "orcado": Decimal("0.00"),
                "comprometido": Decimal("0.00"),
                "medido": Decimal("0.00"),
                "executado": Decimal("0.00"),
                "planejado": Decimal("0.00"),
                "planejado_total": Decimal("0.00"),
                "custo_real": Decimal("0.00"),
            }
            context["score_operacional"] = {
                "pontuacao": Decimal("0.00"),
                "faixa": "Sem leitura",
                "componentes": [],
                "total_alertas_ativos": 0,
                "total_alertas_pendentes_score": 0,
                "total_riscos_ativos": 0,
                "total_ncs_abertas": 0,
            }
            context["score_operacional_grafico"] = {"gradiente": "#d1d5db 0 100%", "fatias": []}
            context["eva"] = {
                "PV": Decimal("0.00"),
                "EV": Decimal("0.00"),
                "AC": Decimal("0.00"),
                "CPI": Decimal("0.00"),
                "SPI": Decimal("0.00"),
                "CV": Decimal("0.00"),
                "SV": Decimal("0.00"),
            }
            context["grafico_geral"] = []
            context["cards_percentuais"] = []
            context["top_itens_orcamento"] = []
            context["pendencias"] = []
            context["resumo_riscos"] = {
                "total": 0,
                "criticos": 0,
                "altos": 0,
                "medios": 0,
                "baixos": 0,
                "em_tratamento": 0,
                "fechados": 0,
            }
            context["alertas"] = []
            context["alertas_operacionais"] = []
            context["alertas_planejamento_suprimentos"] = []
            context["prioridades_executivas"] = []
            context["correlacoes_operacionais"] = []
            context["execucoes_regras_operacionais"] = []
            context["nao_conformidades_abertas"] = []
            context["pipeline_aquisicoes"] = []
            context["ultimas_ordens_compra"] = []
            return context
        folhas = PlanoContas.objects.annotate(filhos_count=Count("filhos")).filter(filhos_count=0)
        folhas = _filtrar_por_obra_contexto(self.request, folhas)
        valor_orcado = folhas.aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00")
        valor_comprometido = _filtrar_por_obra_contexto(self.request, Compromisso.objects.all()).aggregate(total=Sum("valor_contratado"))["total"] or Decimal("0.00")
        valor_medido = _filtrar_por_obra_contexto(self.request, Medicao.objects.all()).aggregate(total=Sum("valor_medido"))["total"] or Decimal("0.00")
        valor_pago = _filtrar_por_obra_contexto(self.request, NotaFiscal.objects.all()).aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00")

        context["indicadores"] = {
            "valor_orcado": valor_orcado,
            "valor_comprometido": valor_comprometido,
            "valor_medido": valor_medido,
            "valor_pago": valor_pago,
        }
        indicadores_exec = _cache_get_or_set_local(
            f"home:consolidado:{obra_contexto.pk}",
            lambda: IntegracaoService.consolidar_obra(obra_contexto),
            request=self.request,
        )
        eva = _cache_get_or_set_local(
            f"home:eva:{obra_contexto.pk}",
            lambda: EVAService.calcular(obra_contexto),
            request=self.request,
        )
        context["indicadores_exec"] = indicadores_exec
        plano_referencia = IntegracaoService.obter_plano_referencia(obra_contexto)
        context["cronograma_ativo_url"] = (
            reverse("plano_fisico_detail", kwargs={"pk": plano_referencia.pk})
            if plano_referencia
            else reverse("plano_fisico_dashboard")
        )
        context["eva"] = eva
        context["indicadores_dashboard"] = IndicadoresService.resumo_obra(
            obra_contexto,
            include_curva_s=False,
            consolidado=indicadores_exec,
            eva=eva,
        )
        context["score_operacional"] = context["indicadores_dashboard"]["score_operacional"]
        context["score_operacional_grafico"] = _grafico_score_operacional(context["score_operacional"])
        context["grafico_geral"] = [
            {"label": "Orçado", "valor": valor_orcado, "percentual": 100 if valor_orcado else 0, "cor": "#2563eb", "mostrar_percentual": False},
            {"label": "Comprometido", "valor": valor_comprometido, "percentual": _calcular_percentual(valor_comprometido, valor_orcado), "cor": "#b91c1c", "mostrar_percentual": True},
            {"label": "Medido", "valor": valor_medido, "percentual": _calcular_percentual(valor_medido, valor_orcado), "cor": "#f59e0b", "mostrar_percentual": True},
            {"label": "Valor pago", "valor": valor_pago, "percentual": _calcular_percentual(valor_pago, valor_orcado), "cor": "#16a34a", "mostrar_percentual": True},
        ]
        context["cards_percentuais"] = [
            {"label": "% Comprometido", "valor": _calcular_percentual(valor_comprometido, valor_orcado), "cor": "#8f2020"},
            {"label": "% Medido", "valor": _calcular_percentual(valor_medido, valor_orcado), "cor": "#4b5563"},
            {"label": "% Pago", "valor": _calcular_percentual(valor_pago, valor_orcado), "cor": "#315fc0"},
        ]

        # Cronograma de orÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â§amento por mÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âªs (barras mensais + linha do acumulado).
        # Distribui o valor total orÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â§ado uniformemente entre data_inicio..data_fim da obra.
        cronograma_orcado_meses = []
        cronograma_orcado_svg_width = 1
        cronograma_orcado_polyline_points = ""
        if obra_contexto and obra_contexto.data_inicio and obra_contexto.data_fim and valor_orcado:
            start = obra_contexto.data_inicio.replace(day=1)
            end = obra_contexto.data_fim.replace(day=1)
            if end < start:
                start, end = end, start

            total_meses = ((end.year - start.year) * 12 + (end.month - start.month) + 1)
            total_meses = max(1, total_meses)

            # DistribuiÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â§ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â£o estÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡vel (sem "estourar" o ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âºltimo mÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âªs por arredondamento):
            # base = floor(valor/meses, 2 casas) e distribui o resto (centavos) nos primeiros meses.
            cent = Decimal("0.01")
            base = (valor_orcado / Decimal(total_meses)).quantize(cent, rounding="ROUND_DOWN")
            resto = (valor_orcado - (base * Decimal(total_meses))).quantize(cent)
            # Quantidade de centavos a distribuir (sempre >= 0).
            extra_cents = int((resto / cent).to_integral_value())

            # Garante que a soma feche no valor total no ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âºltimo mÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âªs.
            month_starts = []
            cursor = start
            while cursor <= end:
                month_starts.append(cursor)
                if cursor.month == 12:
                    cursor = cursor.replace(year=cursor.year + 1, month=1, day=1)
                else:
                    cursor = cursor.replace(month=cursor.month + 1, day=1)

            acumulado = Decimal("0.00")
            max_val = Decimal("0.00")
            for i, _ in enumerate(month_starts):
                valor_mes = base + (cent if i < extra_cents else Decimal("0.00"))
                acumulado += valor_mes
                max_val = max(max_val, acumulado)

            # Evita divisÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â£o por zero.
            max_val = max(max_val, base)

            bar_max_height = 130
            chart_top = 20
            chart_bottom = chart_top + bar_max_height
            chart_width_step = 70
            bar_width = 22
            padding_left = 16

            polyline_points = []
            acumulado = Decimal("0.00")
            for i, ms in enumerate(month_starts):
                valor_mes = base + (cent if i < extra_cents else Decimal("0.00"))

                acumulado += valor_mes
                bar_h = int((valor_mes / max_val) * bar_max_height) if max_val else 0
                line_y = int(chart_bottom - (acumulado / max_val) * bar_max_height) if max_val else chart_bottom

                x = padding_left + i * chart_width_step
                bar_x = x
                bar_y = chart_bottom - bar_h
                line_x = x + (bar_width // 2)

                cronograma_orcado_meses.append(
                    {
                        "label": ms.strftime("%m/%Y"),
                        "valor_mes": valor_mes,
                        "acumulado": acumulado,
                        "bar_x": bar_x,
                        "bar_y": bar_y,
                        "bar_h": bar_h,
                        "line_x": line_x,
                        "line_y": line_y,
                    }
                )
                polyline_points.append(f"{line_x},{line_y}")

            cronograma_orcado_svg_width = max(1, len(cronograma_orcado_meses) * chart_width_step + padding_left + 10)
            cronograma_orcado_polyline_points = " ".join(polyline_points)

        context["cronograma_orcado_meses"] = cronograma_orcado_meses
        context["cronograma_orcado_svg_width"] = cronograma_orcado_svg_width
        context["cronograma_orcado_polyline_points"] = cronograma_orcado_polyline_points

        context["ultimos_compromissos"] = (
            _anotar_execucao_compromissos(
                _filtrar_por_obra_contexto(self.request, Compromisso.objects.select_related("centro_custo", "obra"))
            )
            .prefetch_related(Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo")))
            .order_by("-id")[:5]
        )
        context["obras_ativas"] = Obra.objects.exclude(status="CONCLUIDA").count() if not obra_contexto else 1
        
        # EstatÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â­sticas de Riscos ISO 6.1
        riscos_qs = _filtrar_por_obra_contexto(self.request, Risco.objects.all())
        resumo_riscos = _cache_get_or_set_local(
            f"home:riscos:{obra_contexto.pk}",
            lambda: riscos_qs.aggregate(
                total=Count("id"),
                criticos=Count("id", filter=Q(nivel__gt=15)),
                altos=Count("id", filter=Q(nivel__gte=10, nivel__lte=15)),
                medios=Count("id", filter=Q(nivel__gte=5, nivel__lt=10)),
                baixos=Count("id", filter=Q(nivel__lt=5)),
                em_tratamento=Count("id", filter=Q(status="EM_TRATAMENTO")),
                fechados=Count("id", filter=Q(status="FECHADO")),
            ),
            request=self.request,
        )
        context["resumo_riscos"] = resumo_riscos
        
        # Substituir pendÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âªncias por resumo de riscos
        context["pendencias"] = [
            {"label": "Riscos CrÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â­ticos", "valor": resumo_riscos["criticos"], "cor": "red"},
            {"label": "Riscos Altos", "valor": resumo_riscos["altos"], "cor": "orange"},
            {"label": "Riscos MÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â©dios", "valor": resumo_riscos["medios"], "cor": "yellow"},
            {"label": "Em Tratamento", "valor": resumo_riscos["em_tratamento"], "cor": "blue"},
            {"label": "Fechados", "valor": resumo_riscos["fechados"], "cor": "green"},
        ]
        context["alertas"] = list(
            _anotar_execucao_compromissos(
                _filtrar_por_obra_contexto(
                    self.request,
                    Compromisso.objects.filter(tipo="CONTRATO", valor_contratado__gt=Decimal("0.00")).select_related("obra"),
                )
            )
            .annotate(
                saldo_percentual_alerta=ExpressionWrapper(
                    F("saldo_anotado") / F("valor_contratado"),
                    output_field=DecimalField(max_digits=15, decimal_places=4),
                )
            )
            .filter(saldo_percentual_alerta__lt=Decimal("0.10"))
            .order_by("numero")[:8]
        )
        _sincronizar_alertas_operacionais_rate_limited(obra_contexto)
        painel_alertas = _cache_get_or_set_local(
            f"home:painel_alertas:{obra_contexto.pk}",
            lambda: resumo_executivo_alertas_operacionais(obra_contexto),
            request=self.request,
        )
        resumo_alertas = painel_alertas["resumo_alertas"]
        alertas_recentes = _cache_get_or_set_local(
            f"home:alertas_recentes:{obra_contexto.pk}",
            lambda: listar_alertas_operacionais_ativos(obra_contexto, limit=8),
            request=self.request,
        )
        alertas_planejamento_suprimentos = _cache_get_or_set_local(
            f"home:alertas_planejamento_suprimentos:{obra_contexto.pk}",
            lambda: listar_alertas_planejamento_suprimentos(obra_contexto, limit=6),
            request=self.request,
        )
        context["alertas_operacionais"] = [
            {
                "label": "Contratos sem medicao",
                "valor": resumo_alertas["contrato_sem_medicao"],
                "nivel": "alto" if resumo_alertas["contrato_sem_medicao"] else "medio",
            },
            {
                "label": "Medicoes sem nota",
                "valor": resumo_alertas["medicao_sem_nota"],
                "nivel": "alto",
            },
            {
                "label": "Notas sem apropriacao completa",
                "valor": resumo_alertas["nota_sem_rateio"],
                "nivel": "critico" if resumo_alertas["nota_sem_rateio"] else "alto",
            },
            {
                "label": "Contratos com saldo crítico",
                "valor": len(context["alertas"]),
                "nivel": "critico",
            },
            {
                "label": "Atividades futuras sem solicitacao",
                "valor": resumo_alertas["planejamento_suprimentos"],
                "nivel": "critico" if any(alerta.severidade == "CRITICA" for alerta in alertas_planejamento_suprimentos) else "alto",
            },
            {
                "label": "Atividades sem avanco",
                "valor": resumo_alertas["atividade_sem_avanco"],
                "nivel": "alto",
            },
            {
                "label": "Desvios de prazo",
                "valor": resumo_alertas["desvio_prazo"] + resumo_alertas["estouro_prazo"],
                "nivel": "critico" if (resumo_alertas["desvio_prazo"] + resumo_alertas["estouro_prazo"]) else "alto",
            },
            {
                "label": "Desvios de custo",
                "valor": resumo_alertas["desvio_custo"] + resumo_alertas["custo_sem_avanco"] + resumo_alertas["compromisso_acima_orcado"],
                "nivel": "critico" if (resumo_alertas["desvio_custo"] + resumo_alertas["custo_sem_avanco"] + resumo_alertas["compromisso_acima_orcado"]) else "alto",
            },
            {
                "label": "Desvio combinado custo + prazo",
                "valor": resumo_alertas["desvio_combinado"],
                "nivel": "critico",
            },
            {
                "label": "Riscos vencidos sem tratamento",
                "valor": resumo_alertas["risco_vencido"] + resumo_alertas["acumulo_riscos"],
                "nivel": "critico" if resumo_alertas["risco_vencido"] else "alto",
            },
            {
                "label": "NCs sem evolucao",
                "valor": resumo_alertas["nc_sem_evolucao"],
                "nivel": "alto",
            },
        ]
        context["alertas_planejamento_suprimentos"] = alertas_planejamento_suprimentos
        context["alertas_operacionais_recentes"] = alertas_recentes
        context["prioridades_executivas"] = painel_alertas["prioridades_executivas"]
        context["correlacoes_operacionais"] = painel_alertas["correlacoes_operacionais"]
        context["execucoes_regras_operacionais"] = painel_alertas["execucoes_recentes"]
        context["nao_conformidades_abertas"] = list(
            NaoConformidade.objects.filter(obra=obra_contexto).exclude(status__in=["ENCERRADA", "CANCELADA"]).select_related("responsavel").order_by("-criado_em")[:5]
        )
        context["pipeline_aquisicoes"] = [
            {"label": "Solicitacoes abertas", "valor": SolicitacaoCompra.objects.filter(obra=obra_contexto).exclude(status__in=["ENCERRADA", "CANCELADA"]).count()},
            {"label": "Cotacoes aprovadas", "valor": Cotacao.objects.filter(obra=obra_contexto, status="APROVADA").count()},
            {"label": "Ordens emitidas", "valor": OrdemCompra.objects.filter(obra=obra_contexto).count()},
        ]
        context["ultimas_ordens_compra"] = list(
            OrdemCompra.objects.filter(obra=obra_contexto).select_related("fornecedor", "compromisso_relacionado").order_by("-data_emissao", "-id")[:5]
        )
        context["ultimos_fechamentos"] = _filtrar_por_obra_contexto(self.request, FechamentoMensal.objects.select_related("obra")).order_by("-ano", "-mes")[:5]
        context["obra_contexto"] = obra_contexto

        # Top 10 itens mais importantes do Orcamento (Nivel 5 do Plano de Contas)
        centros_nivel5_qs = PlanoContas.objects.filter(level=4).only(
            "id", "codigo", "descricao", "tree_id", "lft", "rght", "level", "obra_id"
        )
        centros_nivel5_qs = _filtrar_por_obra_contexto(self.request, centros_nivel5_qs)
        centros_nivel5 = list(centros_nivel5_qs)

        top_itens = []
        if centros_nivel5:
            desc_filter = Q()
            for c in centros_nivel5:
                desc_filter |= Q(tree_id=c.tree_id, lft__gte=c.lft, rght__lte=c.rght)
            all_desc = list(PlanoContas.objects.filter(desc_filter).values("id", "tree_id", "lft", "rght", "valor_total"))

            node_to_nivel5 = {}
            for node in all_desc:
                for c in centros_nivel5:
                    if node["tree_id"] == c.tree_id and node["lft"] >= c.lft and node["rght"] <= c.rght:
                        node_to_nivel5[node["id"]] = c.id
                        break
            all_desc_ids = list(node_to_nivel5.keys())

            orcado_by_nivel5 = defaultdict(Decimal)
            for node in all_desc:
                n5_id = node_to_nivel5.get(node["id"])
                if n5_id:
                    orcado_by_nivel5[n5_id] += node["valor_total"] or Decimal("0.00")

            comprometido_by_nivel5 = defaultdict(Decimal)
            for row in CompromissoItem.objects.filter(centro_custo_id__in=all_desc_ids).values("centro_custo_id").annotate(total=Sum("valor_total")):
                n5_id = node_to_nivel5.get(row["centro_custo_id"])
                if n5_id:
                    comprometido_by_nivel5[n5_id] += row["total"] or Decimal("0.00")

            medido_by_nivel5 = defaultdict(Decimal)
            for row in MedicaoItem.objects.filter(centro_custo_id__in=all_desc_ids).values("centro_custo_id").annotate(total=Sum("valor_total")):
                n5_id = node_to_nivel5.get(row["centro_custo_id"])
                if n5_id:
                    medido_by_nivel5[n5_id] += row["total"] or Decimal("0.00")

            pago_by_nivel5 = defaultdict(Decimal)
            for row in NotaFiscalCentroCusto.objects.filter(centro_custo_id__in=all_desc_ids).values("centro_custo_id").annotate(total=Sum("valor")):
                n5_id = node_to_nivel5.get(row["centro_custo_id"])
                if n5_id:
                    pago_by_nivel5[n5_id] += row["total"] or Decimal("0.00")

            for centro in sorted(centros_nivel5, key=lambda c: -(orcado_by_nivel5.get(c.id, Decimal("0.00")))):
                valor_orcado_centro = orcado_by_nivel5.get(centro.id, Decimal("0.00"))
                valor_comprometido_centro = comprometido_by_nivel5.get(centro.id, Decimal("0.00"))
                valor_medido_centro = medido_by_nivel5.get(centro.id, Decimal("0.00"))
                valor_pago_centro = pago_by_nivel5.get(centro.id, Decimal("0.00"))

                saldo_comprometer = arredondar_moeda(valor_orcado_centro - valor_comprometido_centro)
                saldo_medir = arredondar_moeda(valor_comprometido_centro - valor_medido_centro)
                saldo_executar = arredondar_moeda(valor_orcado_centro - valor_pago_centro)

                if saldo_executar <= Decimal("0.00"):
                    situacao = "Concluido"
                elif valor_comprometido_centro > Decimal("0.00") and saldo_medir > Decimal("0.00"):
                    situacao = "Em Medicao"
                elif valor_comprometido_centro > Decimal("0.00"):
                    situacao = "Apropriacao em andamento"
                else:
                    situacao = "Sem Compromisso"

                top_itens.append({
                    "centro": centro,
                    "situacao": situacao,
                    "valor_orcado": valor_orcado_centro,
                    "valor_comprometido": valor_comprometido_centro,
                    "valor_medido": valor_medido_centro,
                    "valor_pago": valor_pago_centro,
                    "saldo_a_comprometer": saldo_comprometer,
                    "saldo_a_medir": saldo_medir,
                    "saldo_a_executar": saldo_executar,
                })
        context["top_itens_orcamento"] = top_itens[:10]
        return context


class CentralAlertasOperacionaisView(LoginRequiredMixin, TemplateView):
    template_name = "app/alerta_operacional_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        context["obra_contexto"] = obra_contexto
        context["sem_obra_selecionada"] = obra_contexto is None
        context["grupos_navegacao"] = list(_obter_grupos_navegacao().values())
        context["alertas"] = []
        context["regras_disponiveis"] = []
        context["resumo_status"] = []
        context["resumo_severidade"] = []
        context["execucoes_recentes"] = []
        context["catalogo_regras"] = []
        context["filtros"] = {
            "status": (self.request.GET.get("status") or "").strip(),
            "severidade": (self.request.GET.get("severidade") or "").strip(),
            "regra": (self.request.GET.get("regra") or "").strip(),
            "responsavel": (self.request.GET.get("responsavel") or "").strip(),
            "atraso": (self.request.GET.get("atraso") or "").strip(),
        }
        context.update(acoes_alerta_permitidas(self.request.user))
        context.update(obter_contexto_central_alertas(obra_contexto, context["filtros"]))
        return context


class PainelExecutivoAlertasView(LoginRequiredMixin, TemplateView):
    template_name = "app/alerta_operacional_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["grupos_navegacao"] = list(_obter_grupos_navegacao().values())
        context.update(obter_dados_painel_executivo_alertas(self.request))
        context["score_operacional_grafico"] = _grafico_score_operacional(context["score_operacional"])
        return context


class AlertaOperacionalDetailView(LoginRequiredMixin, DetailView):
    model = AlertaOperacional
    template_name = "app/alerta_operacional_detail.html"
    context_object_name = "alerta"

    def get_queryset(self):
        obra_contexto = _obter_obra_contexto(self.request)
        queryset = (
            AlertaOperacional.objects.select_related("obra", "responsavel", "ultima_acao_por")
            .prefetch_related("historico__usuario")
        )
        if obra_contexto:
            queryset = queryset.filter(obra=obra_contexto)
        else:
            queryset = queryset.none()
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["historico"] = self.object.historico.select_related("usuario").all()
        context["execucoes_regra"] = self.object.execucoes_automacao.all()[:10]
        context["regra_catalogo"] = obter_regra_operacional(self.object.codigo_regra, self.object.obra.empresa)
        context["today"] = timezone.localdate()
        context["grupos_navegacao"] = list(_obter_grupos_navegacao().values())
        context.update(acoes_alerta_permitidas(self.request.user))
        return context


@login_required
def alerta_operacional_workflow_view(request, pk):
    alerta = get_object_or_404(AlertaOperacional.objects.select_related("obra"), pk=pk)
    obra_contexto = _obter_obra_contexto(request)
    if not obra_contexto or alerta.obra_id != obra_contexto.id:
        raise Http404("Alerta operacional nao encontrado para a obra selecionada.")
    if request.method != "POST":
        return redirect("alerta_operacional_detail", pk=alerta.pk)

    acao = (request.POST.get("acao") or "").strip()
    observacao = (request.POST.get("observacao") or "").strip()
    prazo_solucao = (request.POST.get("prazo_solucao_em") or "").strip()
    next_url = (request.POST.get("next") or "").strip() or reverse("alerta_operacional_detail", args=[alerta.pk])

    if acao == "assumir":
        if not can_assume_alert(request.user):
            messages.error(request, "Seu perfil nao pode assumir alertas para tratamento.")
            return redirect(next_url)
        if not prazo_solucao:
            messages.error(request, "Informe o prazo para solucao ao assumir o alerta.")
            return redirect(next_url)
        try:
            prazo_solucao_em = date.fromisoformat(prazo_solucao)
        except ValueError:
            messages.error(request, "Informe um prazo de solucao valido.")
            return redirect(next_url)
        if prazo_solucao_em < timezone.localdate():
            messages.error(request, "O prazo para solucao nao pode estar no passado.")
            return redirect(next_url)
        atualizar_status_alerta(
            alerta,
            novo_status="EM_TRATAMENTO",
            usuario=request.user,
            observacao=observacao or "Alerta assumido para tratamento.",
            responsavel=request.user,
            acao_historico="TRATAMENTO",
            prazo_solucao_em=prazo_solucao_em,
        )
        messages.success(request, "Alerta colocado em tratamento.")
    elif acao == "justificar":
        if not can_justify_alert(request.user):
            messages.error(request, "Seu perfil nao pode justificar alertas operacionais.")
            return redirect(next_url)
        if not observacao:
            messages.error(request, "Informe a justificativa para registrar o alerta.")
            return redirect(next_url)
        atualizar_status_alerta(
            alerta,
            novo_status="JUSTIFICADO",
            usuario=request.user,
            observacao=observacao,
            responsavel=alerta.responsavel or request.user,
            acao_historico="JUSTIFICATIVA",
        )
        messages.success(request, "Justificativa registrada com sucesso.")
    elif acao == "encerrar":
        if not can_close_alert(request.user):
            messages.error(request, "Seu perfil nao pode encerrar alertas operacionais.")
            return redirect(next_url)
        if not observacao:
            messages.error(request, "Informe a evidencia ou comentario de encerramento.")
            return redirect(next_url)
        atualizar_status_alerta(
            alerta,
            novo_status="ENCERRADO",
            usuario=request.user,
            observacao=observacao,
            responsavel=alerta.responsavel or request.user,
            acao_historico="ENCERRAMENTO",
        )
        messages.success(request, "Alerta encerrado com sucesso.")
    elif acao == "reabrir":
        if not can_close_alert(request.user):
            messages.error(request, "Seu perfil nao pode reabrir alertas operacionais.")
            return redirect(next_url)
        atualizar_status_alerta(
            alerta,
            novo_status="ABERTO",
            usuario=request.user,
            observacao=observacao or "Alerta reaberto para acompanhamento.",
            responsavel=request.user,
            acao_historico="REABERTURA",
        )
        messages.success(request, "Alerta reaberto.")
    else:
        messages.error(request, "Acao do alerta nao reconhecida.")

    return redirect(next_url)


@login_required
def alerta_operacional_dashboard_export_view(request):
    dados = obter_dados_painel_executivo_alertas(request)
    if dados["sem_obra_selecionada"]:
        messages.error(request, "Selecione uma obra para exportar o painel executivo de alertas.")
        return redirect("alerta_operacional_dashboard")

    linhas = []
    for item in dados["prioridades_executivas"]:
        linhas.append(
            {
                "Secao": "Prioridades Executivas",
                "Item": item["frente"],
                "Nivel": item["nivel"].upper(),
                "Quantidade": item["total"],
                "Detalhe": item["acao"],
            }
        )
    for item in dados["correlacoes_operacionais"]:
        linhas.append(
            {
                "Secao": "Correlacoes Operacionais",
                "Item": item["titulo"],
                "Nivel": item["nivel"].upper(),
                "Quantidade": item["quantidade"],
                "Detalhe": item["descricao"],
            }
        )
    for alerta in dados["alertas_em_atraso"]:
        linhas.append(
            {
                "Secao": "Alertas em Atraso",
                "Item": alerta.codigo_regra,
                "Nivel": alerta.severidade,
                "Quantidade": 1,
                "Detalhe": f"{alerta.titulo} | Responsavel: {getattr(alerta.responsavel, 'username', '-') or '-'} | Prazo: {alerta.prazo_solucao_em.strftime('%d/%m/%Y') if alerta.prazo_solucao_em else '-'}",
            }
        )
    if not linhas:
        linhas.append({"Secao": "Resumo", "Item": "Sem dados", "Nivel": "-", "Quantidade": 0, "Detalhe": "Nenhum alerta executivo consolidado."})
    return _exportar_excel_response("painel_alertas_operacionais.xlsx", "Painel Alertas Operacionais", linhas)


@login_required
def alerta_operacional_dashboard_pdf_view(request):
    dados = obter_dados_painel_executivo_alertas(request)
    if dados["sem_obra_selecionada"]:
        messages.error(request, "Selecione uma obra para exportar o painel executivo de alertas.")
        return redirect("alerta_operacional_dashboard")

    obra = dados["obra_contexto"]
    score = dados["score_operacional"]
    resumo = {
        "Obra": str(obra),
        "Score Operacional": f"{score.get('pontuacao', Decimal('0.00'))}",
        "Faixa": score.get("faixa", "-"),
        "Alertas criticos": len(dados["alertas_criticos"]),
        "Alertas em atraso": len(dados["alertas_em_atraso"]),
        "Execucoes recentes": len(dados["execucoes_recentes"]),
    }
    secoes = [
        {
            "titulo": "Prioridades Executivas",
            "colunas": [
                {"chave": "Frente", "titulo": "Frente"},
                {"chave": "Nivel", "titulo": "Nivel"},
                {"chave": "Total", "titulo": "Total"},
                {"chave": "Acao", "titulo": "Acao"},
            ],
            "linhas": [
                {
                    "Frente": item["frente"],
                    "Nivel": item["nivel"].upper(),
                    "Total": item["total"],
                    "Acao": item["acao"],
                }
                for item in dados["prioridades_executivas"]
            ],
        },
        {
            "titulo": "Correlacoes Operacionais",
            "colunas": [
                {"chave": "Correlacao", "titulo": "Correlacao"},
                {"chave": "Nivel", "titulo": "Nivel"},
                {"chave": "Ocorrencias", "titulo": "Ocorrencias"},
                {"chave": "Leitura", "titulo": "Leitura"},
            ],
            "linhas": [
                {
                    "Correlacao": item["titulo"],
                    "Nivel": item["nivel"].upper(),
                    "Ocorrencias": item["quantidade"],
                    "Leitura": item["descricao"],
                }
                for item in dados["correlacoes_operacionais"]
            ],
        },
        {
            "titulo": "Alertas Criticos",
            "colunas": [
                {"chave": "Regra", "titulo": "Regra"},
                {"chave": "Titulo", "titulo": "Titulo"},
                {"chave": "Status", "titulo": "Status"},
                {"chave": "Responsavel", "titulo": "Responsavel"},
                {"chave": "Prazo", "titulo": "Prazo"},
            ],
            "linhas": [
                {
                    "Regra": alerta.codigo_regra,
                    "Titulo": alerta.titulo,
                    "Status": alerta.get_status_display(),
                    "Responsavel": getattr(alerta.responsavel, "username", "-") or "-",
                    "Prazo": alerta.prazo_solucao_em.strftime("%d/%m/%Y") if alerta.prazo_solucao_em else "-",
                }
                for alerta in dados["alertas_criticos"]
            ],
        },
        {
            "titulo": "Alertas em Atraso",
            "colunas": [
                {"chave": "Regra", "titulo": "Regra"},
                {"chave": "Titulo", "titulo": "Titulo"},
                {"chave": "Prazo vencido", "titulo": "Prazo vencido"},
                {"chave": "SLA", "titulo": "SLA"},
                {"chave": "Responsavel", "titulo": "Responsavel"},
            ],
            "linhas": [
                {
                    "Regra": alerta.codigo_regra,
                    "Titulo": alerta.titulo,
                    "Prazo vencido": "SIM" if alerta.em_atraso_prazo else "NAO",
                    "SLA": "SIM" if alerta.em_atraso_sla else "NAO",
                    "Responsavel": getattr(alerta.responsavel, "username", "-") or "-",
                }
                for alerta in dados["alertas_em_atraso"]
            ],
        },
        {
            "titulo": "Execucoes Automaticas",
            "colunas": [
                {"chave": "Data", "titulo": "Data"},
                {"chave": "Regra", "titulo": "Regra"},
                {"chave": "Resultado", "titulo": "Resultado"},
                {"chave": "Referencia", "titulo": "Referencia"},
            ],
            "linhas": [
                {
                    "Data": _datahora_local(execucao.executado_em).strftime("%d/%m/%Y %H:%M") if _datahora_local(execucao.executado_em) else "-",
                    "Regra": execucao.codigo_regra,
                    "Resultado": execucao.get_resultado_display(),
                    "Referencia": execucao.referencia or "-",
                }
                for execucao in dados["execucoes_recentes"]
            ],
        },
    ]
    return _pdf_relatorio_tabelas_response(
        "painel_alertas_operacionais.pdf",
        "Painel Executivo de Alertas Operacionais",
        resumo,
        secoes,
    )


class ModuloGrupoView(LoginRequiredMixin, TemplateView):
    template_name = "app/modulo_grupo.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grupos = _obter_grupos_navegacao()
        grupo = grupos.get(self.kwargs["slug"])
        if not grupo:
            raise Http404("Grupo de navegacao nao encontrado.")
        context["grupo"] = grupo
        context["grupos_navegacao"] = list(grupos.values())
        return context


def _dados_curva_abc(request):
    obra_contexto = _obter_obra_contexto(request)
    qs = PlanoContas.objects.filter(level=4)
    if obra_contexto:
        qs = qs.filter(obra=obra_contexto)
    qs = qs.order_by("-codigo")

    valores = []
    for plano in qs:
        valores.append((plano, plano.valor_total_consolidado or Decimal("0.00")))
    valores.sort(key=lambda t: (t[1], t[0].codigo), reverse=True)

    total_geral = sum((v for _, v in valores), start=Decimal("0.00")) or Decimal("0.00")
    acumulado_perc = Decimal("0.00")
    dados = []

    for plano, valor in valores:
        percentual = (valor / total_geral * Decimal("100")) if total_geral else Decimal("0.00")
        acumulado_perc += percentual

        if acumulado_perc <= Decimal("80.00"):
            classe = "A"
        elif acumulado_perc <= Decimal("95.00"):
            classe = "B"
        else:
            classe = "C"

        dados.append(
            {
                "codigo": plano.codigo,
                "descricao": plano.descricao,
                "valor_total": valor,
                "percentual": round(float(percentual), 1),
                "acumulado": round(float(acumulado_perc), 1),
                "classe": classe,
            }
        )
    return {"dados": dados, "obra_contexto": obra_contexto, "total_geral": total_geral}


class CurvaABCView(TemplateView):
    template_name = "app/curva_abc.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)

        # 5ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âº nÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â­vel do plano de contas (raiz=0 => level=4).
        qs = PlanoContas.objects.filter(level=4)
        if obra_contexto:
            qs = qs.filter(obra=obra_contexto)
        qs = qs.order_by("-codigo")

        valores = []
        for plano in qs:
            valores.append((plano, plano.valor_total_consolidado or Decimal("0.00")))
        valores.sort(key=lambda t: (t[1], t[0].codigo), reverse=True)

        total_geral = sum((v for _, v in valores), start=Decimal("0.00")) or Decimal("0.00")
        acumulado_perc = Decimal("0.00")
        dados = []

        for plano, valor in valores:
            percentual = (valor / total_geral * Decimal("100")) if total_geral else Decimal("0.00")
            acumulado_perc += percentual

            if acumulado_perc <= Decimal("80.00"):
                classe = "A"
            elif acumulado_perc <= Decimal("95.00"):
                classe = "B"
            else:
                classe = "C"

            dados.append(
                {
                    "codigo": plano.codigo,
                    "descricao": plano.descricao,
                    "valor_total": valor,
                    "percentual": round(float(percentual), 1),
                    "acumulado": round(float(acumulado_perc), 1),
                    "classe": classe,
                }
            )

        context["dados"] = dados
        context["obra_contexto"] = obra_contexto
        return context


class CurvaABCView(TemplateView):
    template_name = "app/curva_abc.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_dados_curva_abc(self.request))
        return context


@login_required
def curva_abc_export_view(request):
    dados = _dados_curva_abc(request)
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="TERCEIRO",
        entidade="CurvaABC",
        identificador=f"Obra {dados['obra_contexto']}" if dados["obra_contexto"] else "Curva ABC consolidada",
        acao="EXPORT",
        finalidade="Exportacao gerencial de dados analiticos da obra",
        detalhes="Exportacao Excel da Curva ABC.",
    )
    linhas = [
        {
            "Codigo": item["codigo"],
            "Descricao": item["descricao"],
            "Valor Total": item["valor_total"],
            "Percentual": f'{item["percentual"]}%',
            "Percentual Acumulado": f'{item["acumulado"]}%',
            "Classe": item["classe"],
        }
        for item in dados["dados"]
    ]
    return _exportar_excel_response("curva_abc.xlsx", "Curva ABC", linhas)


@login_required
def curva_abc_pdf_view(request):
    dados = _dados_curva_abc(request)
    resumo = {
        "Obra": f'{dados["obra_contexto"].codigo} - {dados["obra_contexto"].nome}' if dados["obra_contexto"] else "Todas",
        "Total Geral": money_br(dados["total_geral"]),
        "Quantidade de Itens": len(dados["dados"]),
    }
    extras = [
        {
            "Codigo": item["codigo"],
            "Descricao": item["descricao"],
            "Valor Total": money_br(item["valor_total"]),
            "%": f'{item["percentual"]}%',
            "% Acum.": f'{item["acumulado"]}%',
            "Classe": item["classe"],
        }
        for item in dados["dados"]
    ]
    return _pdf_relatorio_probatorio_response(
        "curva_abc.pdf",
        "Curva ABC",
        resumo,
        [],
        extras,
        extras_titulo="Classificacao ABC",
        extras_colunas=[("Codigo", 55), ("Descricao", 180), ("Valor Total", 90), ("%", 45), ("% Acum.", 55), ("Classe", 70)],
        incluir_historico=False,
    )


class ObraListView(ListView):
    model = Obra
    template_name = "app/obra_list.html"
    context_object_name = "obras"

    def get_queryset(self):
        queryset = _get_obras_permitidas(self.request.user).order_by("codigo")
        termo = self.request.GET.get("q", "").strip()
        status = self.request.GET.get("status", "").strip()
        if termo:
            queryset = queryset.filter(Q(codigo__icontains=termo) | Q(nome__icontains=termo) | Q(cliente__icontains=termo))
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["busca"] = self.request.GET.get("q", "").strip()
        context["status_filtro"] = self.request.GET.get("status", "").strip()
        context["status_choices"] = Obra._meta.get_field("status").choices
        return context


class ObraCreateView(CreateView):
    model = Obra
    form_class = ObraForm
    template_name = "app/form.html"
    success_url = reverse_lazy("obra_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nova Obra"
        context["voltar_url"] = reverse_lazy("obra_list")
        return context

    def form_valid(self, form):
        empresa = _get_empresa_operacional(self.request)
        if not empresa:
            form.add_error(None, "Nao foi possivel identificar a empresa da obra.")
            return self.form_invalid(form)
        self.object = form.save(commit=False)
        self.object.empresa = empresa
        self.object.save()
        _registrar_historico("CRIACAO", self.object, f"Obra criada: {self.object}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class ObraUpdateView(UpdateView):
    model = Obra
    form_class = ObraForm
    template_name = "app/form.html"
    success_url = reverse_lazy("obra_list")

    def get_queryset(self):
        return _get_obras_permitidas(self.request.user).order_by("codigo")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = f"Editar Obra {self.object.codigo}"
        context["voltar_url"] = reverse_lazy("obra_list")
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        _registrar_historico("ATUALIZACAO", self.object, f"Obra atualizada: {self.object}", self.request.user)
        return response


class PlanoContasConsultaView(ListView):
    model = PlanoContas
    template_name = "app/plano_contas_list.html"
    context_object_name = "planos_contas"

    def get_queryset(self):
        queryset = PlanoContas.objects.annotate(filhos_count=Count("filhos")).order_by("tree_id", "lft")
        queryset = _filtrar_por_obra_contexto(self.request, queryset)
        termo = self.request.GET.get("q", "").strip()
        if termo:
            queryset = queryset.filter(Q(descricao__icontains=termo) | Q(codigo__icontains=termo))
        return _consolidar_plano_contas(queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        baseline_ativa = None
        baseline_map = {}
        baselines_recentes = []
        baseline_selecionada = None
        if obra_contexto:
            baseline_id = self.request.GET.get("baseline")
            baselines_qs = OrcamentoBaseline.objects.filter(obra=obra_contexto).select_related("criado_por", "aprovado_por", "enviado_para_aprovacao_por").prefetch_related("itens")
            baseline_ativa = baselines_qs.filter(is_ativa=True).first() or baselines_qs.first()
            if baseline_id:
                baseline_selecionada = baselines_qs.filter(pk=baseline_id).first()
            if not baseline_selecionada:
                baseline_selecionada = baseline_ativa
            baselines_recentes = list(baselines_qs[:5])
            if baseline_selecionada:
                baseline_map = {
                    item.codigo: item.valor_total_consolidado
                    for item in baseline_selecionada.itens.all()
                }

        for plano in context["planos_contas"]:
            valor_baseline = baseline_map.get(plano.codigo, Decimal("0.00"))
            plano.valor_baseline_calc = valor_baseline
            plano.desvio_baseline_calc = arredondar_moeda(plano.valor_total_consolidado_calc - valor_baseline)

        context["busca"] = self.request.GET.get("q", "").strip()
        context["obra_contexto"] = obra_contexto
        context["baseline_ativa"] = baseline_ativa
        context["baseline_selecionada"] = baseline_selecionada
        context["baselines_recentes"] = baselines_recentes
        if baseline_selecionada:
            context.update(_obter_alcada_contexto(self.request.user, baseline_selecionada.valor_total))
        return context


class PlanoContasUpdateView(UpdateView):
    model = PlanoContas
    form_class = PlanoContasForm
    template_name = "app/form.html"
    success_url = reverse_lazy("plano_contas_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = f"Editar Plano de Contas {self.object.codigo}"
        context["voltar_url"] = reverse_lazy("plano_contas_list")
        return context


@login_required
def plano_contas_delete_view(request):
    return _apagar_objeto(request, PlanoContas.objects.all(), "plano_contas_list")


@login_required
def plano_contas_export_view(request):
    queryset = PlanoContas.objects.annotate(filhos_count=Count("filhos")).order_by("tree_id", "lft")
    queryset = _filtrar_por_obra_contexto(request, queryset)
    termo = request.GET.get("q", "").strip()
    if termo:
        queryset = queryset.filter(Q(descricao__icontains=termo) | Q(codigo__icontains=termo))
    planos = _consolidar_plano_contas(queryset)
    linhas = [
        {
            "Codigo": plano.codigo,
            "Descricao": plano.descricao,
            "Unidade": plano.unidade or "",
            "Quantidade": plano.quantidade,
            "Valor Unitario": plano.valor_unitario,
            "Valor Total": plano.valor_total_consolidado_calc,
            "Comprometido": plano.valor_comprometido_calc,
            "Medido": plano.valor_medido_calc,
            "Valor Executado": plano.valor_executado_calc,
            "Saldo a Comprometer": plano.saldo_a_comprometer_calc,
            "Saldo a Medir": plano.saldo_a_medir_calc,
            "Saldo a Executar": plano.saldo_a_executar_calc,
        }
        for plano in planos
    ]
    return _exportar_excel_response("plano_de_contas.xlsx", "Plano de Contas", linhas)


@login_required
def plano_contas_criar_baseline_view(request):
    obra_contexto = _obter_obra_contexto(request)
    if request.method != "POST":
        return redirect("plano_contas_list")
    if not obra_contexto:
        messages.error(request, "Selecione uma obra antes de criar uma baseline do orcamento.")
        return redirect("plano_contas_list")
    if not PlanoContas.objects.filter(obra=obra_contexto).exists():
        messages.error(request, "Nao ha plano de contas para gerar baseline nesta obra.")
        return redirect("plano_contas_list")

    descricao = (request.POST.get("descricao_baseline") or "").strip()
    if not descricao:
        descricao = f"Baseline de orcamento {timezone.now():%d/%m/%Y %H:%M}"

    baseline = _criar_baseline_orcamento(obra_contexto, descricao=descricao, usuario=request.user)
    _registrar_historico(
        "BASELINE_ORCAMENTO",
        obra_contexto,
        f"Baseline de orcamento criada: {baseline.descricao}",
        request.user,
    )
    messages.success(request, "Baseline de orcamento criada com sucesso.")
    return redirect("plano_contas_list")


@login_required
def plano_contas_baseline_workflow_view(request, pk):
    if request.method != "POST":
        return redirect("plano_contas_list")
    baseline = get_object_or_404(
        OrcamentoBaseline.objects.select_related("obra"),
        pk=pk,
    )
    baseline = get_object_or_404(
        _filtrar_por_obra_contexto(request, OrcamentoBaseline.objects.select_related("obra")),
        pk=pk,
    )
    acao = request.POST.get("acao")
    if acao == "enviar_para_aprovacao":
        _enviar_baseline_para_aprovacao(request, baseline)
    elif acao == "aprovar":
        _aprovar_baseline(request, baseline)
    elif acao == "retornar_para_ajuste":
        _retornar_baseline_para_ajuste(request, baseline)
    return redirect(f"{reverse_lazy('plano_contas_list')}?baseline={baseline.pk}")


@login_required
def plano_contas_baseline_aprovacao_pdf_view(request, pk):
    baseline = get_object_or_404(
        _filtrar_por_obra_contexto(
            request,
            OrcamentoBaseline.objects.select_related(
                "obra",
                "criado_por",
                "enviado_para_aprovacao_por",
                "aprovado_por",
            ).prefetch_related("itens"),
        ),
        pk=pk,
    )
    evidencia_id, resumo, historico, extras, _ = _dados_relatorio_aprovacao_baseline(baseline)
    return _pdf_relatorio_probatorio_response(
        f"baseline_orcamento_{baseline.pk}_aprovacao.pdf",
        f"Relatório Probatório de Aprovação - {baseline.descricao}",
        resumo,
        historico,
        extras,
        extras_titulo="Snapshot da Baseline",
        extras_colunas=[
            ("Codigo", 70),
            ("Descricao", 245),
            ("Nivel", 45),
            ("Valor Consolidado", 135),
        ],
    )


@login_required
def plano_contas_baseline_aprovacao_excel_view(request, pk):
    baseline = get_object_or_404(
        _filtrar_por_obra_contexto(
            request,
            OrcamentoBaseline.objects.select_related(
                "obra",
                "criado_por",
                "enviado_para_aprovacao_por",
                "aprovado_por",
            ).prefetch_related("itens"),
        ),
        pk=pk,
    )
    _, resumo, historico, extras, _ = _dados_relatorio_aprovacao_baseline(baseline)
    return _exportar_relatorio_probatorio_excel_response(
        f"baseline_orcamento_{baseline.pk}_aprovacao.xlsx",
        "Resumo",
        resumo,
        historico,
        extras_sheet_name="Snapshot",
        extras_linhas=extras,
    )


def plano_contas_notas_view(request, pk):
    try:
        plano = PlanoContas.objects.get(pk=pk)
    except PlanoContas.DoesNotExist as exc:
        raise Http404("Centro de custo não encontrado.") from exc

    centros_ids = list(plano.get_descendants(include_self=True).values_list("id", flat=True))
    rateios = (
        NotaFiscalCentroCusto.objects
        .filter(centro_custo_id__in=centros_ids)
        .select_related("nota_fiscal", "centro_custo")
        .order_by("-nota_fiscal__data_emissao", "-nota_fiscal_id", "centro_custo__codigo")
    )
    notas = [
        {
            "id": rateio.nota_fiscal_id,
            "numero": rateio.nota_fiscal.numero,
            "fornecedor": rateio.nota_fiscal.fornecedor,
            "cnpj": rateio.nota_fiscal.cnpj,
            "descricao": rateio.nota_fiscal.descricao,
            "centro_custo": f"{rateio.centro_custo.codigo} - {rateio.centro_custo.descricao}",
            "valor": money_br(rateio.valor or Decimal("0.00")),
            "data": rateio.nota_fiscal.data_emissao.strftime("%d/%m/%Y"),
        }
        for rateio in rateios
    ]
    return JsonResponse(
        {
            "centro_custo": f"{plano.codigo} - {plano.descricao}",
            "quantidade_notas": len(notas),
            "notas": notas,
        }
    )


class CompromissoListView(ListView):
    model = Compromisso
    template_name = "app/compromisso_list.html"
    context_object_name = "compromissos"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            _anotar_execucao_compromissos(Compromisso.objects.select_related("centro_custo", "obra"))
            .prefetch_related(Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo")))
            .order_by("-id")
        )
        queryset = _filtrar_por_obra_contexto(self.request, queryset)
        return _filtros_compromissos(self.request, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtros = self.request.GET.copy()
        filtros.pop("page", None)
        context["querystring_sem_pagina"] = filtros.urlencode()
        context["busca"] = self.request.GET.get("q", "").strip()
        context["obra_filtro"] = self.request.GET.get("obra", "").strip()
        context["status_filtro"] = self.request.GET.get("status", "").strip()
        context["fornecedor_filtro"] = self.request.GET.get("fornecedor", "").strip()
        context["responsavel_filtro"] = self.request.GET.get("responsavel", "").strip()
        context["centro_custo_filtro"] = self.request.GET.get("centro_custo", "").strip()
        context["data_inicio"] = self.request.GET.get("data_inicio", "").strip()
        context["data_fim"] = self.request.GET.get("data_fim", "").strip()
        context["obras"] = Obra.objects.order_by("codigo")
        context["centros_custo"] = _filtrar_por_obra_contexto(self.request, PlanoContas.objects.order_by("tree_id", "lft"))
        context["status_choices"] = Compromisso._meta.get_field("status").choices
        return context


class CompromissoCreateView(CreateView):
    model = Compromisso
    form_class = CompromissoForm
    template_name = "app/compromisso_form.html"
    success_url = reverse_lazy("compromisso_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["obra_contexto"] = _obter_obra_contexto(self.request)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        context["titulo"] = "Nova Compra ou Contratação"
        context["voltar_url"] = reverse_lazy("compromisso_list")
        context["item_formset"] = kwargs.get("item_formset") or CompromissoItemFormSet(prefix="itens", form_kwargs={"obra_contexto": obra_contexto})
        return context

    def form_valid(self, form):
        obra_contexto = _obter_obra_contexto(self.request)
        item_formset = CompromissoItemFormSet(self.request.POST, prefix="itens", form_kwargs={"obra_contexto": obra_contexto})
        if not item_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, item_formset=item_formset))

        self.object = form.save(commit=False)
        self.object.status = "RASCUNHO"
        self.object.save()
        item_formset.instance = self.object
        item_formset.save()
        self.object.recalcular_totais_por_itens()
        _registrar_historico("CRIACAO", self.object, f"Compromisso criado: {self.object.numero}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class CompromissoUpdateView(UpdateView):
    model = Compromisso
    form_class = CompromissoForm
    template_name = "app/compromisso_form.html"
    success_url = reverse_lazy("compromisso_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["obra_contexto"] = _obter_obra_contexto(self.request)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        context["titulo"] = f"Editar Compra ou Contratação {self.object.numero}"
        context["voltar_url"] = reverse_lazy("compromisso_list")
        context["item_formset"] = kwargs.get("item_formset") or CompromissoItemFormSet(instance=self.object, prefix="itens", form_kwargs={"obra_contexto": obra_contexto})
        return context

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status != "RASCUNHO":
            messages.error(request, "Somente contratos ou pedidos em rascunho podem ser editados.")
            return redirect("contrato_detail", pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        obra_contexto = _obter_obra_contexto(self.request)
        item_formset = CompromissoItemFormSet(self.request.POST, instance=self.object, prefix="itens", form_kwargs={"obra_contexto": obra_contexto})
        if not item_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, item_formset=item_formset))

        self.object = form.save(commit=False)
        self.object.status = "RASCUNHO"
        self.object.save()
        item_formset.instance = self.object
        item_formset.save()
        self.object.recalcular_totais_por_itens()
        _registrar_historico("ATUALIZACAO", self.object, f"Compromisso atualizado: {self.object.numero}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class ContratoDetailView(DetailView):
    model = Compromisso
    template_name = "app/contrato_detail.html"
    context_object_name = "contrato"

    def get_queryset(self):
        queryset = (
            _anotar_execucao_compromissos(Compromisso.objects)
            .select_related("obra")
            .prefetch_related(
                Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo")),
                "medicoes",
                "anexos",
                "historicos",
                "ordens_compra_estruturadas__solicitacao",
                "ordens_compra_estruturadas__cotacao_aprovada",
            )
        )
        return _filtrar_por_obra_contexto(self.request, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contrato = self.object
        context["medicoes_contrato"] = contrato.medicoes.order_by("-data_medicao")
        if contrato.tipo == "CONTRATO":
            context["notas_contrato"] = (
                NotaFiscal.objects.filter(medicao__contrato=contrato)
                .select_related("medicao", "obra")
                .order_by("-data_emissao")
            )
        else:
            context["notas_contrato"] = (
                NotaFiscal.objects.filter(pedido_compra=contrato)
                .select_related("pedido_compra", "obra")
                .order_by("-data_emissao")
            )
        context["anexo_form"] = kwargs.get("anexo_form") or AnexoOperacionalForm()
        context["saldo_percentual"] = _calcular_percentual(contrato.valor_executado, contrato.valor_contratado) if contrato.valor_contratado else 0

        context["aditivos"] = contrato.aditivos.select_related(
            "enviado_para_aprovacao_por",
            "aprovado_por",
        ).prefetch_related("itens__centro_custo").order_by("-criado_em")

        centros_queryset = obter_centros_do_contrato(contrato)
        aditivo_form = kwargs.get("aditivo_form")
        if not aditivo_form:
            aditivo_form = AditivoContratoForm(initial={"tipo": "VALOR"})

        tipo_formset = aditivo_form.data.get("tipo") if getattr(aditivo_form, "data", None) else None
        if not tipo_formset:
            tipo_formset = (getattr(aditivo_form, "initial", None) or {}).get("tipo") or "VALOR"

        aditivo_item_formset = kwargs.get("aditivo_item_formset")
        if not aditivo_item_formset:
            aditivo_instance = AditivoContrato(contrato=contrato, tipo=tipo_formset)
            aditivo_item_formset = AditivoContratoItemFormSet(
                prefix="aditivos_itens",
                instance=aditivo_instance,
                centros_queryset=centros_queryset,
            )

        context["aditivo_form"] = aditivo_form
        context["aditivo_item_formset"] = aditivo_item_formset
        context["origem_aquisicao"] = contrato.ordens_compra_estruturadas.select_related("solicitacao", "cotacao_aprovada").first()
        context.update(_obter_alcada_contexto(self.request.user, contrato.valor_contratado))
        context["pode_aprovar_aditivo_workflow"] = get_papel_aprovacao(self.request.user) in {"COORDENADOR_OBRAS", "GERENTE_OBRAS"}
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        acao = request.POST.get("acao")
        if acao == "enviar_para_aprovacao":
            _enviar_documento_para_aprovacao(
                request,
                self.object,
                status_em_aprovacao="EM_APROVACAO",
                descricao=f"{self.object.numero} enviado para aprovacao.",
            )
            return redirect("contrato_detail", pk=self.object.pk)
        if acao == "aprovar":
            _aprovar_documento(
                request,
                self.object,
                valor=self.object.valor_contratado,
                status_aprovado="APROVADO",
                descricao=f"{self.object.numero} aprovado.",
            )
            return redirect("contrato_detail", pk=self.object.pk)
        if acao == "retornar_para_ajuste":
            _retornar_documento_para_ajuste(
                request,
                self.object,
                valor=self.object.valor_contratado,
                status_ajuste="RASCUNHO",
                descricao=f"{self.object.numero} devolvido para ajuste.",
            )
            return redirect("contrato_detail", pk=self.object.pk)
        form = AnexoOperacionalForm(request.POST, request.FILES)
        if form.is_valid():
            anexo = form.save(commit=False)
            anexo.compromisso = self.object
            anexo.obra = self.object.obra
            anexo.save()
            _registrar_historico("ANEXO", self.object, f"Anexo incluido no contrato {self.object.numero}", self.request.user)
            messages.success(request, "Anexo incluido com sucesso.")
            return redirect("contrato_detail", pk=self.object.pk)
        return self.render_to_response(self.get_context_data(anexo_form=form))


class AditivoContratoCreateView(CreateView):
    model = AditivoContrato
    form_class = AditivoContratoForm

    def post(self, request, *args, **kwargs):
        contrato = get_object_or_404(
            _filtrar_por_obra_contexto(
                request,
                Compromisso.objects.filter(tipo="CONTRATO"),
                vazio_quando_sem_obra=True,
            ),
            pk=kwargs.get("pk"),
        )

        aditivo_form = self.form_class(request.POST)

        tipo_post = request.POST.get("tipo") or "VALOR"
        aditivo_instance = AditivoContrato(contrato=contrato, tipo=tipo_post)
        if aditivo_form.is_valid():
            aditivo_instance = aditivo_form.save(commit=False)
            aditivo_instance.contrato = contrato

        centros_queryset = obter_centros_do_contrato(contrato)
        aditivo_item_formset = AditivoContratoItemFormSet(
            request.POST,
            instance=aditivo_instance,
            prefix="aditivos_itens",
            centros_queryset=centros_queryset,
        )

        if aditivo_form.is_valid() and aditivo_item_formset.is_valid():
            aditivo_instance.save()
            if not aditivo_instance.solicitado_por_id:
                aditivo_instance.solicitado_por = request.user
            if not aditivo_instance.solicitado_em:
                aditivo_instance.solicitado_em = timezone.now()
            aditivo_item_formset.instance = aditivo_instance
            aditivo_item_formset.save()
            if not aditivo_instance.impacto_resumido:
                if aditivo_instance.tipo == "PRAZO":
                    aditivo_instance.impacto_resumido = f"Prazo ajustado em {aditivo_instance.delta_dias or 0} dia(s)"
                else:
                    aditivo_instance.impacto_resumido = f"Impacto financeiro de {money_br(_valor_total_aditivo(aditivo_instance))}"
            aditivo_instance.save(update_fields=["solicitado_por", "solicitado_em", "impacto_resumido"])
            _registrar_historico(
                "ADITIVO",
                contrato,
                f"Aditivo {aditivo_instance.get_tipo_display()} incluído no contrato {contrato.numero}. Motivo: {aditivo_instance.motivo_mudanca or '-'} | Impacto: {aditivo_instance.impacto_resumido or '-'}",
                request.user,
            )
            messages.success(request, "Aditivo incluído com sucesso.")
            return redirect("contrato_detail", pk=contrato.pk)

        # Erros: renderiza na mesma tela do contrato.
        detail_view = ContratoDetailView()
        detail_view.request = request
        detail_view.object = contrato
        context = detail_view.get_context_data(aditivo_form=aditivo_form, aditivo_item_formset=aditivo_item_formset)
        return render(request, detail_view.template_name, context)


@login_required
def aditivo_contrato_workflow_view(request, pk):
    aditivo = get_object_or_404(
        _filtrar_por_obra_contexto(
            request,
            AditivoContrato.objects.select_related(
                "contrato",
                "contrato__obra",
                "enviado_para_aprovacao_por",
                "aprovado_por",
            ).prefetch_related("itens__centro_custo"),
            campo="contrato__obra",
            vazio_quando_sem_obra=True,
        ),
        pk=pk,
    )
    acao = request.POST.get("acao")
    if acao == "enviar_para_aprovacao":
        _enviar_aditivo_contrato_para_aprovacao(request, aditivo)
    elif acao == "aprovar":
        _aprovar_aditivo_contrato(request, aditivo)
    elif acao == "retornar_para_ajuste":
        _retornar_aditivo_contrato_para_ajuste(request, aditivo)
    return redirect("contrato_detail", pk=aditivo.contrato_id)


def compromisso_delete_view(request):
    return _apagar_objeto(request, Compromisso.objects.all(), "compromisso_list")


def compromisso_export_view(request):
    queryset = (
        _anotar_execucao_compromissos(Compromisso.objects.select_related("centro_custo"))
        .prefetch_related(Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo")))
        .order_by("-id")
    )
    queryset = _filtrar_por_obra_contexto(request, queryset)
    queryset = _filtros_compromissos(request, queryset)
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="FORNECEDOR",
        entidade="Compromisso",
        identificador="Compras e Contratacoes",
        acao="EXPORT",
        finalidade="Exportacao de dados contratuais e cadastrais de fornecedores",
        detalhes="Exportacao Excel da lista de compras e contratacoes.",
    )
    linhas = [
        {
            "Numero": compromisso.numero,
            "Tipo": compromisso.get_tipo_display(),
            "CNPJ": compromisso.cnpj,
            "Fornecedor": compromisso.fornecedor,
            "Descricao": compromisso.descricao,
            "Itens do Pedido/Contrato": " | ".join(
                f"{item.centro_custo.codigo} - {item.centro_custo.descricao}" for item in compromisso.itens.all()
            ),
            "Quantidade": compromisso.quantidade_total,
            "Valor Unitario": compromisso.valor_unitario_medio,
            "Valor Total": compromisso.valor_contratado,
            "Valor Executado": compromisso.valor_executado,
            "Saldo": compromisso.saldo,
            "Responsavel": compromisso.responsavel,
            "Data": compromisso.data_assinatura.strftime("%d/%m/%Y"),
        }
        for compromisso in queryset
    ]
    return _exportar_excel_response("compras_contratacoes.xlsx", "Compras", linhas)


@login_required
def compromisso_lista_pdf_view(request):
    queryset = (
        _anotar_execucao_compromissos(Compromisso.objects.select_related("obra"))
        .prefetch_related(Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo")))
        .order_by("-id")
    )
    queryset = _filtrar_por_obra_contexto(request, queryset)
    queryset = _filtros_compromissos(request, queryset)
    resumo = {"Quantidade de Registros": queryset.count(), "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")}
    extras = [
        {
            "Numero": compromisso.numero,
            "Tipo": compromisso.get_tipo_display(),
            "Fornecedor": compromisso.fornecedor,
            "Status": compromisso.get_status_display(),
            "Valor Total": money_br(compromisso.valor_contratado),
        }
        for compromisso in queryset
    ]
    return _pdf_relatorio_probatorio_response(
        "compras_contratacoes_lista.pdf",
        "Compras e Contratações",
        resumo,
        [],
        extras,
        extras_titulo="Lista de Pedidos e Contratos",
        extras_colunas=[("Numero", 80), ("Tipo", 85), ("Fornecedor", 180), ("Status", 70), ("Valor Total", 80)],
        incluir_historico=False,
    )


def compromisso_pdf_view(request, pk):
    queryset = (
        Compromisso.objects.select_related("obra", "centro_custo")
        .prefetch_related(
            Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo")),
            Prefetch("aditivos", queryset=AditivoContrato.objects.prefetch_related("itens__centro_custo")),
        )
    )
    compromisso = get_object_or_404(_filtrar_por_obra_contexto(request, queryset), pk=pk)
    resumo = {
        "Numero": compromisso.numero,
        "Tipo": compromisso.get_tipo_display(),
        "Obra": f"{compromisso.obra.codigo if compromisso.obra else '-'} - {compromisso.obra.nome if compromisso.obra else '-'}",
        "Fornecedor": compromisso.fornecedor,
        "CNPJ": compromisso.cnpj,
        "Responsável operacional": compromisso.responsavel,
        "Status": compromisso.get_status_display(),
        "Data": compromisso.data_assinatura.strftime("%d/%m/%Y") if compromisso.data_assinatura else "-",
        "Descricao": compromisso.descricao,
        "Valor total": money_br(compromisso.valor_contratado),
    }
    extras = [
        {
            "Centro de Custo": f"{item.centro_custo.codigo} - {item.centro_custo.descricao}",
            "Descrição do Item": item.descricao_tecnica or item.centro_custo.descricao,
            "Quantidade": item.quantidade,
            "Unidade": item.unidade or "-",
            "Valor Unitário": money_br(item.valor_unitario),
            "Valor Total": money_br(item.valor_total),
        }
        for item in compromisso.itens.all()
    ]
    secoes_extras = []
    if compromisso.tipo == "CONTRATO":
        secoes_extras.append(
            {
                "titulo": "Aditivos do Contrato",
                "colunas": [
                    ("Tipo", 55),
                    ("Status", 65),
                    ("Descrição", 140),
                    ("Impacto", 70),
                    ("Centro de Custo", 110),
                    ("Data", 55),
                ],
                "linhas": _linhas_aditivos_compromisso(compromisso)
                or [{"Tipo": "Nenhum registro encontrado", "Status": "", "Descrição": "", "Impacto": "", "Centro de Custo": "", "Data": ""}],
            }
        )
        secoes_extras.append(
            {
                "titulo": "Histórico dos Aditivos",
                "colunas": [
                    ("Data", 75),
                    ("Aditivo", 75),
                    ("Ação", 80),
                    ("Usuário", 85),
                    ("Descrição", 180),
                ],
                "linhas": _historico_aditivos_compromisso(compromisso)
                or [{"Data": "Nenhum registro encontrado", "Aditivo": "", "Ação": "", "Usuário": "", "Descrição": ""}],
            }
        )
    return _pdf_relatorio_probatorio_response(
        f"{compromisso.numero}.pdf",
        f"Compras e Contratações {compromisso.numero}",
        resumo,
        [],
        extras,
        extras_titulo="Itens do Pedido/Contrato",
        extras_colunas=[
            ("Centro de Custo", 120),
            ("Descrição do Item", 145),
            ("Quantidade", 55),
            ("Unidade", 45),
            ("Valor Unitário", 60),
            ("Valor Total", 70),
        ],
        secoes_extras=secoes_extras,
    )


@login_required
def compromisso_aprovacao_pdf_view(request, pk):
    queryset = (
        Compromisso.objects.select_related("obra", "enviado_para_aprovacao_por", "aprovado_por")
        .prefetch_related("historicos", Prefetch("aditivos", queryset=AditivoContrato.objects.prefetch_related("itens__centro_custo")))
    )
    compromisso = get_object_or_404(_filtrar_por_obra_contexto(request, queryset), pk=pk)
    evidencia_id, resumo, historico, extras, _, secoes_extras = _dados_relatorio_aprovacao_compromisso(compromisso)
    return _pdf_relatorio_probatorio_response(
        f"{compromisso.numero}_aprovacao.pdf",
        f"Relatório Probatório de Aprovação - {compromisso.numero}",
        resumo,
        historico,
        extras,
        extras_titulo="Itens do Pedido/Contrato",
        extras_colunas=[
            ("Centro de Custo", 120),
            ("Descrição do Item", 145),
            ("Quantidade", 55),
            ("Unidade", 45),
            ("Valor Unitário", 60),
            ("Valor Total", 70),
        ],
        secoes_extras=secoes_extras,
    )


@login_required
def compromisso_aprovacao_excel_view(request, pk):
    queryset = (
        Compromisso.objects.select_related("obra", "enviado_para_aprovacao_por", "aprovado_por")
        .prefetch_related("historicos", "itens__centro_custo")
    )
    compromisso = get_object_or_404(_filtrar_por_obra_contexto(request, queryset), pk=pk)
    _, resumo, historico, extras, _, _ = _dados_relatorio_aprovacao_compromisso(compromisso)
    return _exportar_relatorio_probatorio_excel_response(
        f"{compromisso.numero}_aprovacao.xlsx",
        "Resumo",
        resumo,
        historico,
        extras_sheet_name="Itens Pedido Contrato",
        extras_linhas=extras,
    )


class MedicaoListView(ListView):
    model = Medicao
    template_name = "app/medicao_list.html"
    context_object_name = "medicoes"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            Medicao.objects.select_related("contrato", "centro_custo", "obra")
            .prefetch_related(Prefetch("itens", queryset=MedicaoItem.objects.select_related("centro_custo")))
            .order_by("-data_medicao", "-id")
        )
        queryset = _filtrar_por_obra_contexto(self.request, queryset)
        return _filtros_medicoes(self.request, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtros = self.request.GET.copy()
        filtros.pop("page", None)
        context["querystring_sem_pagina"] = filtros.urlencode()
        context["busca"] = self.request.GET.get("q", "").strip()
        context["obra_filtro"] = self.request.GET.get("obra", "").strip()
        context["status_filtro"] = self.request.GET.get("status", "").strip()
        context["fornecedor_filtro"] = self.request.GET.get("fornecedor", "").strip()
        context["responsavel_filtro"] = self.request.GET.get("responsavel", "").strip()
        context["contrato_filtro"] = self.request.GET.get("contrato", "").strip()
        context["centro_custo_filtro"] = self.request.GET.get("centro_custo", "").strip()
        context["data_inicio"] = self.request.GET.get("data_inicio", "").strip()
        context["data_fim"] = self.request.GET.get("data_fim", "").strip()
        context["obras"] = Obra.objects.order_by("codigo")
        context["centros_custo"] = _filtrar_por_obra_contexto(self.request, PlanoContas.objects.order_by("tree_id", "lft"))
        context["status_choices"] = Medicao._meta.get_field("status").choices
        return context


class MedicaoCreateView(CreateView):
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["obra_contexto"] = _obter_obra_contexto(self.request)
        return kwargs

    model = Medicao
    form_class = MedicaoForm
    template_name = "app/medicao_form.html"
    success_url = reverse_lazy("medicao_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contrato = _obter_contrato_from_request(self.request)
        context["titulo"] = "Nova Medição"
        context["voltar_url"] = reverse_lazy("medicao_list")
        context["item_formset"] = kwargs.get("item_formset") or _construir_formset_medicao(prefix="itens", contrato=contrato)
        return context

    def form_valid(self, form):
        contrato = form.cleaned_data.get("contrato")
        if contrato and contrato.status != "APROVADO":
            form.add_error("contrato", "Só é possível emitir medição para contratos aprovados.")
            return self.render_to_response(self.get_context_data(form=form, item_formset=_construir_formset_medicao(data=self.request.POST, prefix="itens", contrato=contrato)))
        item_formset = _construir_formset_medicao(data=self.request.POST, prefix="itens", contrato=contrato)
        if not item_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, item_formset=item_formset))

        obra_contexto = _obter_obra_contexto(self.request)
        self.object = form.save(commit=False)
        self.object.obra = obra_contexto or getattr(contrato, "obra", None)
        self.object.status = "EM_ELABORACAO"
        self.object.save()
        item_formset.instance = self.object
        item_formset.save()
        self.object.recalcular_totais_por_itens()
        _registrar_historico("CRIACAO", self.object, f"Medicao criada: {self.object.numero_da_medicao}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class MedicaoUpdateView(UpdateView):
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["obra_contexto"] = _obter_obra_contexto(self.request)
        return kwargs

    model = Medicao
    form_class = MedicaoForm
    template_name = "app/medicao_form.html"
    success_url = reverse_lazy("medicao_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contrato = _obter_contrato_from_request(self.request, self.object)
        context["titulo"] = f"Editar Medição {self.object.numero_da_medicao}"
        context["voltar_url"] = reverse_lazy("medicao_list")
        context["item_formset"] = kwargs.get("item_formset") or _construir_formset_medicao(instance=self.object, prefix="itens", contrato=contrato)
        return context

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status != "EM_ELABORACAO":
            messages.error(request, "Somente medicoes em elaboracao podem ser editadas.")
            return redirect("medicao_detail", pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        contrato = form.cleaned_data.get("contrato")
        item_formset = _construir_formset_medicao(data=self.request.POST, instance=self.object, prefix="itens", contrato=contrato)
        if not item_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, item_formset=item_formset))

        obra_contexto = _obter_obra_contexto(self.request)
        self.object = form.save(commit=False)
        self.object.obra = obra_contexto or getattr(contrato, "obra", None)
        self.object.status = "EM_ELABORACAO"
        self.object.save()
        item_formset.instance = self.object
        item_formset.save()
        self.object.recalcular_totais_por_itens()
        _registrar_historico("ATUALIZACAO", self.object, f"Medicao atualizada: {self.object.numero_da_medicao}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class MedicaoDetailView(DetailView):
    model = Medicao
    template_name = "app/medicao_detail.html"
    context_object_name = "medicao"

    def get_queryset(self):
        queryset = (
            Medicao.objects.select_related("contrato", "obra")
            .prefetch_related(
                Prefetch("itens", queryset=MedicaoItem.objects.select_related("centro_custo")),
                "notas_fiscais",
                "anexos",
                "historicos",
            )
        )
        return _filtrar_por_obra_contexto(self.request, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        medicao = self.object
        context["notas_medicao"] = medicao.notas_fiscais.order_by("-data_emissao")
        context["anexo_form"] = kwargs.get("anexo_form") or AnexoOperacionalForm()
        context["saldo_percentual"] = _calcular_percentual(medicao.valor_medido, medicao.contrato.valor_contratado) if medicao.contrato.valor_contratado else 0
        context.update(_obter_alcada_contexto(self.request.user, medicao.valor_medido))
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        acao = request.POST.get("acao")
        if acao == "enviar_para_aprovacao":
            _enviar_documento_para_aprovacao(
                request,
                self.object,
                status_em_aprovacao="EM_APROVACAO",
                descricao=f"{self.object.numero_da_medicao} enviada para aprovacao.",
            )
            return redirect("medicao_detail", pk=self.object.pk)
        if acao == "aprovar":
            _aprovar_documento(
                request,
                self.object,
                valor=self.object.valor_medido,
                status_aprovado="APROVADA",
                descricao=f"{self.object.numero_da_medicao} aprovada.",
            )
            return redirect("medicao_detail", pk=self.object.pk)
        if acao == "retornar_para_ajuste":
            _retornar_documento_para_ajuste(
                request,
                self.object,
                valor=self.object.valor_medido,
                status_ajuste="EM_ELABORACAO",
                descricao=f"{self.object.numero_da_medicao} devolvida para ajuste.",
            )
            return redirect("medicao_detail", pk=self.object.pk)
        form = AnexoOperacionalForm(request.POST, request.FILES)
        if form.is_valid():
            anexo = form.save(commit=False)
            anexo.medicao = self.object
            anexo.obra = self.object.obra
            anexo.save()
            _registrar_historico("ANEXO", self.object, f"Anexo incluido na medicao {self.object.numero_da_medicao}", self.request.user)
            messages.success(request, "Anexo incluido com sucesso.")
            return redirect("medicao_detail", pk=self.object.pk)
        return self.render_to_response(self.get_context_data(anexo_form=form))


def medicao_delete_view(request):
    return _apagar_objeto(request, Medicao.objects.all(), "medicao_list")


@login_required
def medicao_aprovacao_pdf_view(request, pk):
    queryset = (
        Medicao.objects.select_related("contrato", "obra", "enviado_para_aprovacao_por", "aprovado_por")
        .prefetch_related("historicos")
    )
    medicao = get_object_or_404(_filtrar_por_obra_contexto(request, queryset), pk=pk)
    evidencia_id, resumo, historico, extras, _ = _dados_relatorio_aprovacao_medicao(medicao)
    return _pdf_relatorio_probatorio_response(
        f"{medicao.numero_da_medicao}_aprovacao.pdf",
        f"Relatório Probatório de Aprovação - {medicao.numero_da_medicao}",
        resumo,
        historico,
        extras,
        extras_titulo="Itens da Medicao",
        extras_colunas=[
            ("Centro de Custo", 245),
            ("Quantidade", 70),
            ("Unidade", 55),
            ("Valor Total", 125),
        ],
    )


@login_required
def medicao_aprovacao_excel_view(request, pk):
    queryset = (
        Medicao.objects.select_related("contrato", "obra", "enviado_para_aprovacao_por", "aprovado_por")
        .prefetch_related("historicos", "itens__centro_custo")
    )
    medicao = get_object_or_404(_filtrar_por_obra_contexto(request, queryset), pk=pk)
    _, resumo, historico, extras, _ = _dados_relatorio_aprovacao_medicao(medicao)
    return _exportar_relatorio_probatorio_excel_response(
        f"{medicao.numero_da_medicao}_aprovacao.xlsx",
        "Resumo",
        resumo,
        historico,
        extras_sheet_name="Itens",
        extras_linhas=extras,
    )


def medicao_export_view(request):
    queryset = (
        Medicao.objects.select_related("contrato", "centro_custo")
        .prefetch_related(Prefetch("itens", queryset=MedicaoItem.objects.select_related("centro_custo")))
        .order_by("-data_medicao", "-id")
    )
    queryset = _filtrar_por_obra_contexto(request, queryset)
    queryset = _filtros_medicoes(request, queryset)
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="FORNECEDOR",
        entidade="Medicao",
        identificador="Medicoes",
        acao="EXPORT",
        finalidade="Exportacao de dados financeiros e cadastrais vinculados a contratos",
        detalhes="Exportacao Excel da lista de medicoes.",
    )
    linhas = [
        {
            "Numero": medicao.numero_da_medicao,
            "Contrato": medicao.contrato.numero,
            "CNPJ": medicao.cnpj,
            "Fornecedor": medicao.fornecedor,
            "Descricao": medicao.descricao,
            "Itens Medidos": " | ".join(
                f"{item.centro_custo.codigo} - {item.centro_custo.descricao}" for item in medicao.itens.all()
            ),
            "Quantidade": medicao.quantidade_total,
            "Valor Unitario": medicao.valor_unitario_medio,
            "Valor Total": medicao.valor_medido,
            "Responsavel": medicao.responsavel,
            "Data": medicao.data_medicao.strftime("%d/%m/%Y"),
        }
        for medicao in queryset
    ]
    return _exportar_excel_response("medicoes.xlsx", "Medicoes", linhas)


@login_required
def medicao_lista_pdf_view(request):
    queryset = (
        Medicao.objects.select_related("contrato")
        .prefetch_related(Prefetch("itens", queryset=MedicaoItem.objects.select_related("centro_custo")))
        .order_by("-data_medicao", "-id")
    )
    queryset = _filtrar_por_obra_contexto(request, queryset)
    queryset = _filtros_medicoes(request, queryset)
    resumo = {"Quantidade de Registros": queryset.count(), "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")}
    extras = [
        {
            "Numero": medicao.numero_da_medicao,
            "Contrato": medicao.contrato.numero,
            "Fornecedor": medicao.fornecedor,
            "Status": medicao.get_status_display(),
            "Valor Total": money_br(medicao.valor_medido),
        }
        for medicao in queryset
    ]
    return _pdf_relatorio_probatorio_response(
        "medicoes_lista.pdf",
        "Lista de Medicoes",
        resumo,
        [],
        extras,
        extras_titulo="Lista de Medicoes",
        extras_colunas=[("Numero", 85), ("Contrato", 85), ("Fornecedor", 180), ("Status", 70), ("Valor Total", 75)],
        incluir_historico=False,
    )


class NotaFiscalListView(ListView):
    model = NotaFiscal
    template_name = "app/nota_fiscal_list.html"
    context_object_name = "notas_fiscais"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            NotaFiscal.objects.select_related("medicao", "pedido_compra", "obra")
            .prefetch_related("centros_custo__centro_custo")
            .order_by("-data_emissao", "-id")
        )
        queryset = _filtrar_por_obra_contexto(self.request, queryset)
        return _filtros_notas(self.request, queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtros = self.request.GET.copy()
        filtros.pop("page", None)
        context["querystring_sem_pagina"] = filtros.urlencode()
        context["busca"] = self.request.GET.get("q", "").strip()
        context["obra_filtro"] = self.request.GET.get("obra", "").strip()
        context["status_filtro"] = self.request.GET.get("status", "").strip()
        context["fornecedor_filtro"] = self.request.GET.get("fornecedor", "").strip()
        context["contrato_filtro"] = self.request.GET.get("contrato", "").strip()
        context["centro_custo_filtro"] = self.request.GET.get("centro_custo", "").strip()
        context["data_inicio"] = self.request.GET.get("data_inicio", "").strip()
        context["data_fim"] = self.request.GET.get("data_fim", "").strip()
        context["obras"] = Obra.objects.order_by("codigo")
        context["centros_custo"] = _filtrar_por_obra_contexto(self.request, PlanoContas.objects.order_by("tree_id", "lft"))
        context["status_choices"] = NotaFiscal._meta.get_field("status").choices
        return context


class NotaFiscalCreateView(CreateView):
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["obra_contexto"] = _obter_obra_contexto(self.request)
        return kwargs

    model = NotaFiscal
    form_class = NotaFiscalForm
    template_name = "app/nota_fiscal_form.html"
    success_url = reverse_lazy("nota_fiscal_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido, medicao = _obter_origem_nota(self.request)
        obra_contexto = _obter_obra_contexto(self.request)
        context["titulo"] = "Nova Nota Fiscal"
        context["voltar_url"] = reverse_lazy("nota_fiscal_list")
        context["rateio_formset"] = kwargs.get("rateio_formset") or _construir_formset_nota(prefix="rateio", pedido=pedido, medicao=medicao, obra=obra_contexto)
        return context

    def form_valid(self, form):
        pedido = form.cleaned_data.get("pedido_compra")
        medicao = form.cleaned_data.get("medicao")
        if medicao and medicao.status != "APROVADA":
            form.add_error("medicao", "Só é possível emitir nota fiscal para medições aprovadas.")
        if pedido and pedido.status != "APROVADO":
            form.add_error("pedido_compra", "Só é possível emitir nota fiscal para pedidos aprovados.")
        if form.errors:
            rateio_formset = _construir_formset_nota(
                data=self.request.POST,
                instance=form.instance,
                prefix="rateio",
                pedido=pedido,
                medicao=medicao,
                obra=_obter_obra_contexto(self.request),
            )
            return self.render_to_response(self.get_context_data(form=form, rateio_formset=rateio_formset))
        obra_contexto = _obter_obra_contexto(self.request)
        self.object = form.save(commit=False)
        self.object.obra = obra_contexto or getattr(medicao, "obra", None) or getattr(pedido, "obra", None)
        rateio_formset = _construir_formset_nota(
            data=self.request.POST,
            instance=self.object,
            prefix="rateio",
            pedido=pedido,
            medicao=medicao,
            obra=obra_contexto,
        )
        if not rateio_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, rateio_formset=rateio_formset))

        self.object.save()
        rateio_formset.instance = self.object
        rateio_formset.save()
        _registrar_historico("CRIACAO", self.object, f"Nota fiscal criada: {self.object.numero}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class NotaFiscalUpdateView(UpdateView):
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["obra_contexto"] = _obter_obra_contexto(self.request)
        return kwargs

    model = NotaFiscal
    form_class = NotaFiscalForm
    template_name = "app/nota_fiscal_form.html"
    success_url = reverse_lazy("nota_fiscal_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido, medicao = _obter_origem_nota(self.request, self.object)
        obra_contexto = _obter_obra_contexto(self.request)
        context["titulo"] = f"Editar Nota Fiscal {self.object.numero}"
        context["voltar_url"] = reverse_lazy("nota_fiscal_list")
        context["rateio_formset"] = kwargs.get("rateio_formset") or _construir_formset_nota(
            instance=self.object,
            prefix="rateio",
            pedido=pedido,
            medicao=medicao,
            obra=obra_contexto,
        )
        return context

    def form_valid(self, form):
        pedido = form.cleaned_data.get("pedido_compra")
        medicao = form.cleaned_data.get("medicao")
        if medicao and medicao.status != "APROVADA":
            form.add_error("medicao", "Só é possível emitir nota fiscal para medições aprovadas.")
        if pedido and pedido.status != "APROVADO":
            form.add_error("pedido_compra", "Só é possível emitir nota fiscal para pedidos aprovados.")
        if form.errors:
            rateio_formset = _construir_formset_nota(
                data=self.request.POST,
                instance=self.object,
                prefix="rateio",
                pedido=pedido,
                medicao=medicao,
                obra=_obter_obra_contexto(self.request),
            )
            return self.render_to_response(self.get_context_data(form=form, rateio_formset=rateio_formset))
        obra_contexto = _obter_obra_contexto(self.request)
        self.object = form.save(commit=False)
        self.object.obra = obra_contexto or getattr(medicao, "obra", None) or getattr(pedido, "obra", None)
        rateio_formset = _construir_formset_nota(
            data=self.request.POST,
            instance=self.object,
            prefix="rateio",
            pedido=pedido,
            medicao=medicao,
            obra=obra_contexto,
        )
        if not rateio_formset.is_valid():
            return self.render_to_response(self.get_context_data(form=form, rateio_formset=rateio_formset))

        self.object.save()
        rateio_formset.instance = self.object
        rateio_formset.save()
        _registrar_historico("ATUALIZACAO", self.object, f"Nota fiscal atualizada: {self.object.numero}", self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class FechamentoMensalView(TemplateView):
    template_name = "app/fechamento_mensal.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obra_contexto = _obter_obra_contexto(self.request)
        obra_id = self.request.GET.get("obra", "").strip()
        hoje = date.today()
        ano = int(self.request.GET.get("ano") or hoje.year)
        mes = int(self.request.GET.get("mes") or hoje.month)
        obras = Obra.objects.order_by("codigo")
        if obra_id:
            obra = obras.filter(pk=obra_id).first()
        elif obra_contexto:
            obra = obra_contexto
        else:
            obra = obras.first()
        compromissos = Compromisso.objects.filter(data_assinatura__year=ano, data_assinatura__month=mes)
        medicoes = Medicao.objects.filter(data_medicao__year=ano, data_medicao__month=mes)
        notas = NotaFiscal.objects.filter(data_emissao__year=ano, data_emissao__month=mes)
        if obra:
            compromissos = compromissos.filter(obra=obra)
            medicoes = medicoes.filter(obra=obra)
            notas = notas.filter(obra=obra)
        itens_compromisso = CompromissoItem.objects.select_related("centro_custo", "compromisso").filter(
            compromisso__data_assinatura__year=ano,
            compromisso__data_assinatura__month=mes,
        )
        itens_medicao = MedicaoItem.objects.select_related("centro_custo", "medicao").filter(
            medicao__data_medicao__year=ano,
            medicao__data_medicao__month=mes,
        )
        rateios_nota = NotaFiscalCentroCusto.objects.select_related("centro_custo", "nota_fiscal").filter(
            nota_fiscal__data_emissao__year=ano,
            nota_fiscal__data_emissao__month=mes,
        )
        if obra:
            itens_compromisso = itens_compromisso.filter(compromisso__obra=obra)
            itens_medicao = itens_medicao.filter(medicao__obra=obra)
            rateios_nota = rateios_nota.filter(nota_fiscal__obra=obra)

        plano_qs = PlanoContas.objects.filter(obra=obra) if obra else PlanoContas.objects.all()
        nodes_by_id = {n.id: n for n in plano_qs.only("id", "parent_id", "level", "codigo", "descricao")}

        def get_nivel5_ancestor(node_id):
            node = nodes_by_id.get(node_id)
            while node:
                if node.level == 4:
                    return node
                if node.level < 4:
                    return None
                node = nodes_by_id.get(node.parent_id)
            return None

        resumo_nivel5 = defaultdict(lambda: {"comprometido": Decimal("0.00"), "medido": Decimal("0.00"), "notas": Decimal("0.00")})
        for item in itens_compromisso:
            anc = get_nivel5_ancestor(item.centro_custo_id)
            if anc:
                resumo_nivel5[anc.id]["centro"] = anc
                resumo_nivel5[anc.id]["comprometido"] += item.valor_total or Decimal("0.00")
        for item in itens_medicao:
            anc = get_nivel5_ancestor(item.centro_custo_id)
            if anc:
                resumo_nivel5[anc.id]["centro"] = anc
                resumo_nivel5[anc.id]["medido"] += item.valor_total or Decimal("0.00")
        for item in rateios_nota:
            anc = get_nivel5_ancestor(item.centro_custo_id)
            if anc:
                resumo_nivel5[anc.id]["centro"] = anc
                resumo_nivel5[anc.id]["notas"] += item.valor or Decimal("0.00")

        centros_fechamento = []
        for payload in resumo_nivel5.values():
            if not payload.get("centro"):
                continue
            centros_fechamento.append(
                {
                    "centro": payload["centro"],
                    "comprometido": payload["comprometido"],
                    "medido": payload["medido"],
                    "notas": payload["notas"],
                    "saldo_a_medir": payload["comprometido"] - payload["medido"],
                    "saldo_a_executar": payload["comprometido"] - payload["notas"],
                }
            )
        centros_fechamento.sort(key=lambda item: item["centro"].codigo if item.get("centro") else "")
        context["obras"] = obras
        context["obra_atual"] = obra
        context["ano"] = ano
        context["mes"] = mes
        context["resumo"] = {
            "valor_comprometido": compromissos.aggregate(total=Sum("valor_contratado"))["total"] or Decimal("0.00"),
            "valor_medido": medicoes.aggregate(total=Sum("valor_medido"))["total"] or Decimal("0.00"),
            "valor_notas": notas.aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00"),
            "qtd_compromissos": compromissos.count(),
            "qtd_medicoes": medicoes.count(),
            "qtd_notas": notas.count(),
        }
        context["resumo_centros"] = centros_fechamento
        context["fechamentos"] = FechamentoMensal.objects.select_related("obra").order_by("-ano", "-mes")[:12]
        return context

    def post(self, request, *args, **kwargs):
        def _parse_int_br(value, default=None):
            if value is None:
                return default
            raw = str(value).strip()
            if not raw:
                return default
            # Aceita "2.026" (separador de milhar pt-BR) e tambÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â©m "2,026" por engano.
            raw = raw.replace(".", "").replace(",", "")
            return int(raw)

        obra = get_object_or_404(Obra, pk=request.POST.get("obra"))
        ano = _parse_int_br(request.POST.get("ano"))
        mes = _parse_int_br(request.POST.get("mes"))
        fechamento, _ = FechamentoMensal.objects.update_or_create(
            obra=obra,
            ano=ano,
            mes=mes,
            defaults={
                "valor_comprometido": Compromisso.objects.filter(obra=obra, data_assinatura__year=ano, data_assinatura__month=mes).aggregate(total=Sum("valor_contratado"))["total"] or Decimal("0.00"),
                "valor_medido": Medicao.objects.filter(obra=obra, data_medicao__year=ano, data_medicao__month=mes).aggregate(total=Sum("valor_medido"))["total"] or Decimal("0.00"),
                "valor_notas": NotaFiscal.objects.filter(obra=obra, data_emissao__year=ano, data_emissao__month=mes).aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00"),
            },
        )
        _registrar_historico("FECHAMENTO", obra, f"Fechamento mensal registrado: {fechamento}", request.user)
        messages.success(request, "Fechamento mensal registrado com sucesso.")
        return redirect(f"{reverse_lazy('fechamento_mensal')}?obra={obra.pk}&ano={ano}&mes={mes}")


class ProjecaoFinanceiraView(TemplateView):
    template_name = "app/projecao_financeira.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        meses_opcoes = [6, 12]
        meses_qtd = int(self.request.GET.get("meses") or 12)
        if meses_qtd not in meses_opcoes:
            meses_qtd = 12

        hoje = date.today()
        inicio = date(hoje.year, hoje.month, 1)

        def add_one_month(d):
            if d.month == 12:
                return d.replace(year=d.year + 1, month=1, day=1)
            return d.replace(month=d.month + 1, day=1)

        month_starts = [inicio]
        for _ in range(meses_qtd - 1):
            month_starts.append(add_one_month(month_starts[-1]))

        fim_exclusivo = add_one_month(month_starts[-1])
        idx_by_month = {m: i for i, m in enumerate(month_starts)}

        entradas = [Decimal("0.00") for _ in month_starts]
        saidas = [Decimal("0.00") for _ in month_starts]

        notas_qs = NotaFiscal.objects.filter(data_emissao__gte=inicio, data_emissao__lt=fim_exclusivo)
        notas_qs = _filtrar_por_obra_contexto(self.request, notas_qs)
        for nota in notas_qs:
            m = date(nota.data_emissao.year, nota.data_emissao.month, 1)
            idx = idx_by_month.get(m)
            if idx is not None:
                entradas[idx] += nota.valor_total or Decimal("0.00")

        medicoes_qs = Medicao.objects.select_related("contrato").all()
        medicoes_qs = _filtrar_por_obra_contexto(self.request, medicoes_qs)
        medicoes = list(medicoes_qs)
        contrato_ids = {m.contrato_id for m in medicoes if getattr(m, "contrato_id", None)}

        prazo_deltas = {}
        if contrato_ids:
            prazo_totais = (
                AditivoContrato.objects.filter(contrato_id__in=contrato_ids, tipo="PRAZO")
                .values("contrato_id")
                .annotate(total=Sum("delta_dias"))
            )
            for row in prazo_totais:
                prazo_deltas[row["contrato_id"]] = row["total"] or 0

        # Distribui o valor medido pelos meses "previstos" (shiftados por PRAZO).
        for m in medicoes:
            valor_medido = m.valor_medido or Decimal("0.00")
            if not valor_medido:
                continue

            med_start_raw = m.data_prevista_inicio or m.data_medicao
            med_end_raw = m.data_prevista_fim or med_start_raw
            delta = prazo_deltas.get(m.contrato_id, 0) or 0
            med_start = med_start_raw + timedelta(days=delta)
            med_end = med_end_raw + timedelta(days=delta)

            med_start_m = med_start.replace(day=1)
            med_end_m = med_end.replace(day=1)

            total_meses_intervalo = ((med_end_m.year - med_start_m.year) * 12 + (med_end_m.month - med_start_m.month) + 1)
            total_meses_intervalo = max(1, total_meses_intervalo)

            share = valor_medido / Decimal(total_meses_intervalo)
            for i, ms in enumerate(month_starts):
                if ms >= med_start_m and ms <= med_end_m:
                    saidas[i] += share

        series = []
        for i, ms in enumerate(month_starts):
            entrada = entradas[i]
            saida = saidas[i]
            saldo = arredondar_moeda(entrada - saida)
            series.append(
                {
                    "label": ms.strftime("%m/%Y"),
                    "entrada": arredondar_moeda(entrada),
                    "saida": arredondar_moeda(saida),
                    "saldo": saldo,
                }
            )

        context["meses_opcoes"] = meses_opcoes
        context["meses_qtd"] = meses_qtd
        context["series"] = series
        context["total_entradas"] = arredondar_moeda(sum(s["entrada"] for s in series))
        context["total_saidas"] = arredondar_moeda(sum(s["saida"] for s in series))
        context["total_saldo"] = arredondar_moeda(context["total_entradas"] - context["total_saidas"])
        return context


def _adicionar_um_mes(data_base):
    if data_base.month == 12:
        return data_base.replace(year=data_base.year + 1, month=1, day=1)
    return data_base.replace(month=data_base.month + 1, day=1)


def _parse_int_query_param(valor, default):
    if valor in (None, ""):
        return default
    bruto = str(valor).strip().replace(".", "").replace(",", "")
    try:
        return int(bruto)
    except (TypeError, ValueError):
        return default


def _dados_fechamento_mensal(request):
    obra_contexto = _obter_obra_contexto(request)
    obra_id = request.GET.get("obra", "").strip()
    hoje = date.today()
    ano = _parse_int_query_param(request.GET.get("ano"), hoje.year)
    mes = _parse_int_query_param(request.GET.get("mes"), hoje.month)
    obras = Obra.objects.order_by("codigo")
    if obra_id:
        obra = obras.filter(pk=obra_id).first()
    elif obra_contexto:
        obra = obra_contexto
    else:
        obra = obras.first()

    compromissos = Compromisso.objects.filter(data_assinatura__year=ano, data_assinatura__month=mes)
    medicoes = Medicao.objects.filter(data_medicao__year=ano, data_medicao__month=mes)
    notas = NotaFiscal.objects.filter(data_emissao__year=ano, data_emissao__month=mes)
    if obra:
        compromissos = compromissos.filter(obra=obra)
        medicoes = medicoes.filter(obra=obra)
        notas = notas.filter(obra=obra)

    itens_compromisso = CompromissoItem.objects.select_related("centro_custo", "compromisso").filter(
        compromisso__data_assinatura__year=ano,
        compromisso__data_assinatura__month=mes,
    )
    itens_medicao = MedicaoItem.objects.select_related("centro_custo", "medicao").filter(
        medicao__data_medicao__year=ano,
        medicao__data_medicao__month=mes,
    )
    rateios_nota = NotaFiscalCentroCusto.objects.select_related("centro_custo", "nota_fiscal").filter(
        nota_fiscal__data_emissao__year=ano,
        nota_fiscal__data_emissao__month=mes,
    )
    if obra:
        itens_compromisso = itens_compromisso.filter(compromisso__obra=obra)
        itens_medicao = itens_medicao.filter(medicao__obra=obra)
        rateios_nota = rateios_nota.filter(nota_fiscal__obra=obra)

    plano_qs = PlanoContas.objects.filter(obra=obra) if obra else PlanoContas.objects.all()
    nodes_by_id = {n.id: n for n in plano_qs.only("id", "parent_id", "level", "codigo", "descricao")}

    def get_nivel5_ancestor(node_id):
        node = nodes_by_id.get(node_id)
        while node:
            if node.level == 4:
                return node
            if node.level < 4:
                return None
            node = nodes_by_id.get(node.parent_id)
        return None

    resumo_nivel5 = defaultdict(lambda: {"comprometido": Decimal("0.00"), "medido": Decimal("0.00"), "notas": Decimal("0.00")})
    for item in itens_compromisso:
        anc = get_nivel5_ancestor(item.centro_custo_id)
        if anc:
            resumo_nivel5[anc.id]["centro"] = anc
            resumo_nivel5[anc.id]["comprometido"] += item.valor_total or Decimal("0.00")
    for item in itens_medicao:
        anc = get_nivel5_ancestor(item.centro_custo_id)
        if anc:
            resumo_nivel5[anc.id]["centro"] = anc
            resumo_nivel5[anc.id]["medido"] += item.valor_total or Decimal("0.00")
    for item in rateios_nota:
        anc = get_nivel5_ancestor(item.centro_custo_id)
        if anc:
            resumo_nivel5[anc.id]["centro"] = anc
            resumo_nivel5[anc.id]["notas"] += item.valor or Decimal("0.00")

    centros_fechamento = []
    for payload in resumo_nivel5.values():
        if not payload.get("centro"):
            continue
        centros_fechamento.append(
            {
                "centro": payload["centro"],
                "comprometido": payload["comprometido"],
                "medido": payload["medido"],
                "notas": payload["notas"],
                "saldo_a_medir": payload["comprometido"] - payload["medido"],
                "saldo_a_executar": payload["comprometido"] - payload["notas"],
            }
        )
    centros_fechamento.sort(key=lambda item: item["centro"].codigo if item.get("centro") else "")

    return {
        "obras": obras,
        "obra_atual": obra,
        "ano": ano,
        "mes": mes,
        "resumo": {
            "valor_comprometido": compromissos.aggregate(total=Sum("valor_contratado"))["total"] or Decimal("0.00"),
            "valor_medido": medicoes.aggregate(total=Sum("valor_medido"))["total"] or Decimal("0.00"),
            "valor_notas": notas.aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00"),
            "qtd_compromissos": compromissos.count(),
            "qtd_medicoes": medicoes.count(),
            "qtd_notas": notas.count(),
        },
        "resumo_centros": centros_fechamento,
        "fechamentos": FechamentoMensal.objects.select_related("obra").order_by("-ano", "-mes")[:12],
    }


class FechamentoMensalView(TemplateView):
    template_name = "app/fechamento_mensal.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(dados_fechamento_mensal_request(self.request))
        context["jobs_recentes"] = listar_jobs_recentes(
            empresa=getattr(context.get("obra_atual"), "empresa", _get_empresa_operacional(self.request)),
            obra=context.get("obra_atual"),
            limite=8,
        )
        return context

    def post(self, request, *args, **kwargs):
        def _parse_int_br(value, default=None):
            if value is None:
                return default
            raw = str(value).strip()
            if not raw:
                return default
            raw = raw.replace(".", "").replace(",", "")
            return int(raw)

        obra = get_object_or_404(Obra, pk=request.POST.get("obra"))
        ano = _parse_int_br(request.POST.get("ano"))
        mes = _parse_int_br(request.POST.get("mes"))
        fechamento = registrar_fechamento_mensal(obra=obra, ano=ano, mes=mes)
        _registrar_historico("FECHAMENTO", obra, f"Fechamento mensal registrado: {fechamento}", request.user)
        messages.success(request, "Fechamento mensal registrado com sucesso.")
        return redirect(f"{reverse_lazy('fechamento_mensal')}?obra={obra.pk}&ano={ano}&mes={mes}")


@login_required
def fechamento_mensal_export_view(request):
    dados = dados_fechamento_mensal_request(request)
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="TERCEIRO",
        entidade="FechamentoMensal",
        identificador=f"{dados['mes']:02d}/{dados['ano']}",
        acao="EXPORT",
        finalidade="Exportacao de consolidacao financeira da obra",
        detalhes="Exportacao Excel do fechamento mensal.",
    )
    linhas = [
        {
            "Centro de Custo": f'{linha["centro"].codigo} - {linha["centro"].descricao}',
            "Comprometido": linha["comprometido"],
            "Medido": linha["medido"],
            "Notas": linha["notas"],
            "Saldo a Medir": linha["saldo_a_medir"],
            "Saldo a Executar": linha["saldo_a_executar"],
        }
        for linha in dados["resumo_centros"]
    ]
    return _exportar_excel_response("fechamento_mensal.xlsx", "Fechamento Mensal", linhas)


@login_required
def fechamento_mensal_pdf_view(request):
    dados = dados_fechamento_mensal_request(request)
    resumo = {
        "Obra": f'{dados["obra_atual"].codigo} - {dados["obra_atual"].nome}' if dados["obra_atual"] else "-",
        "Periodo": f'{dados["mes"]:02d}/{dados["ano"]}',
        "Comprometido": money_br(dados["resumo"]["valor_comprometido"]),
        "Medido": money_br(dados["resumo"]["valor_medido"]),
        "Notas": money_br(dados["resumo"]["valor_notas"]),
    }
    extras = [
        {
            "Centro de Custo": f'{linha["centro"].codigo} - {linha["centro"].descricao}',
            "Comprometido": money_br(linha["comprometido"]),
            "Medido": money_br(linha["medido"]),
            "Notas": money_br(linha["notas"]),
            "Saldo": money_br(linha["saldo_a_executar"]),
        }
        for linha in dados["resumo_centros"]
    ]
    return _pdf_relatorio_probatorio_response(
        "fechamento_mensal.pdf",
        "Fechamento Mensal",
        resumo,
        [],
        extras,
        extras_titulo="Consolidado por Centro de Custo",
        extras_colunas=[("Centro de Custo", 215), ("Comprometido", 70), ("Medido", 70), ("Notas", 70), ("Saldo", 70)],
        incluir_historico=False,
    )


def _dados_projecao_financeira(request):
    meses_opcoes = [6, 12]
    meses_qtd = int(request.GET.get("meses") or 12)
    if meses_qtd not in meses_opcoes:
        meses_qtd = 12

    hoje = date.today()
    inicio = date(hoje.year, hoje.month, 1)
    month_starts = [inicio]
    for _ in range(meses_qtd - 1):
        month_starts.append(_adicionar_um_mes(month_starts[-1]))
    fim_exclusivo = _adicionar_um_mes(month_starts[-1])
    idx_by_month = {m: i for i, m in enumerate(month_starts)}

    entradas = [Decimal("0.00") for _ in month_starts]
    saidas = [Decimal("0.00") for _ in month_starts]

    planos_qs = PlanoContas.objects.annotate(filhos_count=Count("filhos")).filter(filhos_count=0)
    planos_qs = _filtrar_por_obra_contexto(request, planos_qs)
    total_orcado = planos_qs.aggregate(total=Sum("valor_total"))["total"] or Decimal("0.00")

    medicoes_qs = Medicao.objects.filter(data_medicao__gte=inicio, data_medicao__lt=fim_exclusivo)
    medicoes_qs = _filtrar_por_obra_contexto(request, medicoes_qs)
    for medicao in medicoes_qs:
        m = date(medicao.data_medicao.year, medicao.data_medicao.month, 1)
        idx = idx_by_month.get(m)
        if idx is not None:
            percentual_medido = (medicao.valor_medido / total_orcado) if total_orcado else Decimal("0.00")
            entradas[idx] += arredondar_moeda(percentual_medido * total_orcado)

    notas_qs = NotaFiscal.objects.filter(data_emissao__gte=inicio, data_emissao__lt=fim_exclusivo)
    notas_qs = _filtrar_por_obra_contexto(request, notas_qs)
    for nota in notas_qs:
        m = date(nota.data_emissao.year, nota.data_emissao.month, 1)
        idx = idx_by_month.get(m)
        if idx is not None:
            saidas[idx] += nota.valor_total or Decimal("0.00")

    series = []
    for i, ms in enumerate(month_starts):
        entrada = arredondar_moeda(entradas[i])
        saida = arredondar_moeda(saidas[i])
        saldo = arredondar_moeda(entrada - saida)
        series.append({"label": ms.strftime("%m/%Y"), "entrada": entrada, "saida": saida, "saldo": saldo})

    return {
        "meses_opcoes": meses_opcoes,
        "meses_qtd": meses_qtd,
        "series": series,
        "total_orcado": total_orcado,
        "total_entradas": arredondar_moeda(sum(s["entrada"] for s in series)),
        "total_saidas": arredondar_moeda(sum(s["saida"] for s in series)),
        "total_saldo": arredondar_moeda(sum(s["saldo"] for s in series)),
    }


class ProjecaoFinanceiraView(TemplateView):
    template_name = "app/projecao_financeira.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(dados_projecao_financeira_request(self.request))
        context["jobs_recentes"] = listar_jobs_recentes(
            empresa=_get_empresa_operacional(self.request),
            obra=_obter_obra_contexto(self.request),
            limite=8,
        )
        return context


@login_required
def projecao_financeira_export_view(request):
    dados = dados_projecao_financeira_request(request)
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="TERCEIRO",
        entidade="ProjecaoFinanceira",
        identificador="Horizonte financeiro",
        acao="EXPORT",
        finalidade="Exportacao de previsao financeira da obra",
        detalhes="Exportacao Excel da projecao financeira.",
    )
    linhas = [
        {"Mes": item["label"], "Entradas": item["entrada"], "Saidas": item["saida"], "Saldo": item["saldo"]}
        for item in dados["series"]
    ]
    return _exportar_excel_response("projecao_financeira.xlsx", "Projecao Financeira", linhas)


@login_required
def projecao_financeira_pdf_view(request):
    dados = dados_projecao_financeira_request(request)
    resumo = {
        "Total Orcado": money_br(dados["total_orcado"]),
        "Total Entradas": money_br(dados["total_entradas"]),
        "Total Saidas": money_br(dados["total_saidas"]),
        "Saldo no Horizonte": money_br(dados["total_saldo"]),
    }
    extras = [
        {"Mes": item["label"], "Entradas": money_br(item["entrada"]), "Saidas": money_br(item["saida"]), "Saldo": money_br(item["saldo"])}
        for item in dados["series"]
    ]
    return _pdf_relatorio_probatorio_response(
        "projecao_financeira.pdf",
        "Projecao Financeira",
        resumo,
        [],
        extras,
        extras_titulo="Visao Mensal",
        extras_colunas=[("Mes", 80), ("Entradas", 135), ("Saidas", 135), ("Saldo", 145)],
        incluir_historico=False,
    )


def nota_fiscal_delete_view(request):
    return _apagar_objeto(request, NotaFiscal.objects.all(), "nota_fiscal_list")


def nota_fiscal_export_view(request):
    queryset = (
        NotaFiscal.objects.select_related("medicao", "pedido_compra")
        .prefetch_related("centros_custo__centro_custo")
        .order_by("-data_emissao", "-id")
    )
    queryset = _filtrar_por_obra_contexto(request, queryset)
    queryset = _filtros_notas(request, queryset)
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="FORNECEDOR",
        entidade="NotaFiscal",
        identificador="Notas fiscais",
        acao="EXPORT",
        finalidade="Exportacao de dados fiscais e cadastrais de fornecedores",
        detalhes="Exportacao Excel da lista de notas fiscais.",
    )
    linhas = [
        {
            "ID": nota.id,
            "Numero da Nota": nota.numero,
            "Origem": str(nota.medicao or nota.pedido_compra or ""),
            "CNPJ": nota.cnpj,
            "Fornecedor": nota.fornecedor,
            "Descricao": nota.descricao,
            "Centro de Custo": " | ".join(
                f"{item.centro_custo.codigo} - {item.centro_custo.descricao}" for item in nota.centros_custo.all()
            ),
            "Valor": nota.valor_total,
            "Data de Emissao": nota.data_emissao.strftime("%d/%m/%Y"),
            "Data de Vencimento": nota.data_vencimento.strftime("%d/%m/%Y") if nota.data_vencimento else "-",
            "Data de Cadastro": _datahora_local(nota.criado_em).strftime("%d/%m/%Y %H:%M") if nota.criado_em else "-",
        }
        for nota in queryset
    ]
    return _exportar_excel_response("notas_fiscais.xlsx", "Notas Fiscais", linhas)


@login_required
def nota_fiscal_lista_pdf_view(request):
    queryset = (
        NotaFiscal.objects.select_related("medicao", "pedido_compra")
        .prefetch_related("centros_custo__centro_custo")
        .order_by("-data_emissao", "-id")
    )
    queryset = _filtrar_por_obra_contexto(request, queryset)
    queryset = _filtros_notas(request, queryset)
    resumo = {"Quantidade de Registros": queryset.count(), "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")}
    extras = [
        {
            "Numero": nota.numero,
            "Fornecedor": nota.fornecedor,
            "Valor": money_br(nota.valor_total),
            "Emissao": nota.data_emissao.strftime("%d/%m/%Y"),
            "Vencimento": nota.data_vencimento.strftime("%d/%m/%Y") if nota.data_vencimento else "-",
        }
        for nota in queryset
    ]
    return _pdf_relatorio_probatorio_response(
        "notas_fiscais_lista.pdf",
        "Lista de Notas Fiscais",
        resumo,
        [],
        extras,
        extras_titulo="Lista de Notas Fiscais",
        extras_colunas=[("Numero", 85), ("Fornecedor", 190), ("Valor", 70), ("Emissao", 75), ("Vencimento", 75)],
        incluir_historico=False,
    )


@login_required
def nota_fiscal_dados_view(request, pk):
    nota = get_object_or_404(
        _filtrar_por_obra_contexto(
            request,
            NotaFiscal.objects.prefetch_related("centros_custo__centro_custo").select_related("medicao", "pedido_compra"),
        ),
        pk=pk,
    )
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="FORNECEDOR",
        entidade="NotaFiscal",
        objeto=nota,
        identificador=nota.numero,
        acao="VIEW",
        finalidade="Consulta detalhada de dados fiscais e rateios para operacao autorizada",
        detalhes="Consulta detalhada via endpoint JSON de nota fiscal.",
    )
    return JsonResponse(
        {
            "id": nota.pk,
            "numero": nota.numero,
            "fornecedor": nota.fornecedor,
            "cnpj": nota.cnpj,
            "descricao": nota.descricao,
            "valor_total": money_br(nota.valor_total),
            "data_emissao": nota.data_emissao.strftime("%d/%m/%Y"),
            "data_vencimento": nota.data_vencimento.strftime("%d/%m/%Y") if nota.data_vencimento else "-",
            "data_cadastro": _datahora_local(nota.criado_em).strftime("%d/%m/%Y %H:%M") if nota.criado_em else "-",
            "rateios": [
                {
                    "centro_custo": f"{item.centro_custo.codigo} - {item.centro_custo.descricao}",
                    "valor": money_br(item.valor or Decimal('0.00')),
                }
                for item in nota.centros_custo.all()
            ],
        }
    )


def contrato_dados_view(request, pk):
    contrato = get_object_or_404(
        _filtrar_por_obra_contexto(
            request,
            Compromisso.objects.prefetch_related(Prefetch("itens", queryset=CompromissoItem.objects.select_related("centro_custo"))),
        ),
        pk=pk,
    )
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="FORNECEDOR",
        entidade="Compromisso",
        objeto=contrato,
        identificador=contrato.numero,
        acao="VIEW",
        finalidade="Consulta detalhada de contrato ou pedido para operacao autorizada",
        detalhes="Consulta detalhada via endpoint JSON de contrato.",
    )
    return JsonResponse(obter_dados_contrato(contrato))


def medicao_dados_view(request, pk):
    medicao = get_object_or_404(
        _filtrar_por_obra_contexto(
            request,
            Medicao.objects.prefetch_related(Prefetch("itens", queryset=MedicaoItem.objects.select_related("centro_custo"))),
        ),
        pk=pk,
    )
    registrar_acesso_dado_pessoal(
        request,
        categoria_titular="FORNECEDOR",
        entidade="Medicao",
        objeto=medicao,
        identificador=medicao.numero_da_medicao,
        acao="VIEW",
        finalidade="Consulta detalhada de medicao para conferencias e operacao autorizada",
        detalhes="Consulta detalhada via endpoint JSON de medicao.",
    )
    return JsonResponse(obter_dados_medicao(medicao))


@login_required
def plano_contas_importar_view(request):
    """View para importar plano de contas vinculado a obra do contexto."""
    obra_contexto = _obter_obra_contexto(request)
    
    if not obra_contexto:
        messages.error(request, "Selecione uma obra no menu antes de importar o plano de contas.")
        return redirect("plano_contas_list")
    
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        if not arquivo:
            messages.error(request, "Selecione um arquivo para importar.")
            return render(request, "app/plano_contas_importar.html", {"obra_contexto": obra_contexto})
        
        try:
            importar_plano_contas_excel(arquivo, obra=obra_contexto)
            messages.success(request, "Plano de contas importado com sucesso!")
            return redirect("plano_contas_list")
        except ValidationError as e:
            messages.error(request, str(e.message) if hasattr(e, "message") else str(e))
        except Exception as e:
            messages.error(request, f"Erro ao importar: {str(e)}")
    
    return render(request, "app/plano_contas_importar.html", {"obra_contexto": obra_contexto})
