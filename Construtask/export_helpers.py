import os
import re
import struct
import unicodedata
import zlib
from io import BytesIO

import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .text_normalization import corrigir_mojibake

_STATIC_APP_DIR = os.path.join(os.path.dirname(__file__), "static", "app")
_PDF_LOGO_PATH = os.path.join(_STATIC_APP_DIR, "logo-construtask.png")
_EXCEL_FILL_RED = PatternFill(fill_type="solid", fgColor="840B0B")
_EXCEL_FONT_WHITE = Font(color="FFFFFF", bold=True)
_EXCEL_FONT_BLACK_BOLD = Font(color="000000", bold=True)
_EXCEL_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def _datahora_local(datahora):
    if not datahora:
        return None
    return timezone.localtime(datahora)


def _normalizar_texto_exportacao(valor):
    if valor is None:
        return "-"
    texto = str(valor)
    return corrigir_mojibake(texto)


def _sanear_texto_exportacao(texto):
    texto = "-" if texto is None else str(texto)
    return corrigir_mojibake(texto)


def _sanear_texto_exportacao_seguro(texto):
    texto = "-" if texto is None else str(texto)
    texto = corrigir_mojibake(texto)
    texto = _sanear_texto_exportacao(texto)
    reparos = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2026": "...",
        "\xa0": " ",
        "\xa2": "\u00e2",
        "\xb3": "\u00f3",
        "\xb4": "\u00f4",
        "\xb5": "\u00f5",
        "\xa3": "\u00e3",
        "\xa9": "\u00e9",
        "\xaa": "\u00ea",
        "\xad": "\u00ed",
        "\xa1": "\u00e1",
        "\xa7": "\u00e7",
        "\xba": "\u00fa",
    }
    for antigo, novo in reparos.items():
        texto = texto.replace(antigo, novo)
    return unicodedata.normalize("NFC", texto)


def _normalizar_linhas_exportacao(linhas):
    normalizadas = []
    for linha in linhas:
        if isinstance(linha, dict):
            normalizadas.append({chave: _sanear_texto_exportacao_seguro(valor) for chave, valor in linha.items()})
        else:
            normalizadas.append(_sanear_texto_exportacao_seguro(linha))
    return normalizadas


def _ajustar_larguras_excel(worksheet):
    for column_cells in worksheet.columns:
        comprimento = 0
        coluna = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                comprimento = max(comprimento, len(str(cell.value or "")))
            except Exception:
                continue
        worksheet.column_dimensions[coluna].width = min(max(comprimento + 2, 12), 42)


def _estilizar_tabela_excel(worksheet, linha_cabecalho, linha_inicio_dados, linha_fim_dados):
    for cell in worksheet[linha_cabecalho]:
        cell.fill = _EXCEL_FILL_RED
        cell.font = _EXCEL_FONT_WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _EXCEL_BORDER

    for linha in worksheet.iter_rows(min_row=linha_inicio_dados, max_row=linha_fim_dados):
        for cell in linha:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = _EXCEL_BORDER


def _aplicar_layout_excel_relatorio(worksheet, titulo_relatorio, subtitulo=None):
    ultima_coluna = max(worksheet.max_column, 1)
    worksheet.insert_rows(1, amount=4)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_coluna)
    worksheet["A1"] = titulo_relatorio
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    worksheet["A1"].fill = _EXCEL_FILL_RED
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].border = _EXCEL_BORDER

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_coluna)
    worksheet["A2"] = "Construtask"
    worksheet["A2"].font = _EXCEL_FONT_BLACK_BOLD
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A2"].border = _EXCEL_BORDER

    if subtitulo:
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ultima_coluna)
        worksheet["A3"] = subtitulo
        worksheet["A3"].font = _EXCEL_FONT_BLACK_BOLD
        worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.freeze_panes = "A5"
    _estilizar_tabela_excel(worksheet, 4, 5, worksheet.max_row)
    _ajustar_larguras_excel(worksheet)


def _exportar_excel_response(nome_arquivo, sheet_name, linhas):
    output = BytesIO()
    linhas = _normalizar_linhas_exportacao(linhas)
    dataframe = pd.DataFrame(linhas)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        _aplicar_layout_excel_relatorio(worksheet, sheet_name)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _pdf_escape(texto):
    texto = "-" if texto is None else str(texto)
    texto = _sanear_texto_exportacao_seguro(texto)
    for termo_antigo, termo_novo in {
        "HISTRICO": "HIST\u00d3RICO",
        "APROVAO": "APROVA\u00c7\u00c3O",
        "DESCRIO": "DESCRI\u00c7\u00c3O",
        "CONTRATAES": "CONTRATA\u00c7\u00d5ES",
        "MEDIO": "MEDI\u00c7\u00c3O",
    }.items():
        texto = texto.replace(termo_antigo, termo_novo)
    return texto.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_wrap_text(texto, largura_maxima, tamanho_fonte):
    texto = _sanear_texto_exportacao_seguro(texto or "-").strip()
    if not texto:
        return ["-"]

    largura_maxima = max(float(largura_maxima or 0), 20.0)
    tamanho_fonte = max(float(tamanho_fonte or 0), 6.0)
    largura_util = max(largura_maxima - 6.0, 10.0)

    paragrafos = texto.splitlines()
    linhas_finais = []

    for paragrafo in paragrafos:
        paragrafo = re.sub(r"\s+", " ", paragrafo).strip()
        if not paragrafo:
            linhas_finais.append("")
            continue

        palavras = paragrafo.split(" ")
        linha_atual = ""

        for palavra in palavras:
            while palavra and _pdf_estimar_largura_texto(palavra, tamanho_fonte) > largura_util:
                if linha_atual:
                    linhas_finais.append(linha_atual)
                    linha_atual = ""
                parte, resto = _pdf_quebrar_palavra_longa(palavra, largura_util, tamanho_fonte)
                linhas_finais.append(parte)
                palavra = resto

            if not linha_atual:
                linha_atual = palavra
            else:
                candidato = f"{linha_atual} {palavra}"
                if _pdf_estimar_largura_texto(candidato, tamanho_fonte) <= largura_util:
                    linha_atual = candidato
                else:
                    linhas_finais.append(linha_atual)
                    linha_atual = palavra

        if linha_atual:
            linhas_finais.append(linha_atual)

    return linhas_finais or ["-"]


def _pdf_estimar_largura_texto(texto, tamanho_fonte):
    texto = _sanear_texto_exportacao_seguro(texto or "")
    largura = 0.0
    for caractere in texto:
        if caractere in "W@%MmQGOD":
            fator = 0.92
        elif caractere in "ABCDEFGHKNOPRSTUVXYZ":
            fator = 0.78
        elif caractere in "mw":
            fator = 0.82
        elif caractere in "ijlI1|.,:;!'` ":
            fator = 0.30
        elif caractere in "-_/\\()[]{}":
            fator = 0.42
        elif caractere.isdigit():
            fator = 0.60
        else:
            fator = 0.56
        largura += tamanho_fonte * fator
    return largura


def _pdf_ajustar_texto_para_largura(texto, largura_maxima, tamanho_fonte, sufixo="..."):
    texto = _sanear_texto_exportacao_seguro(texto or "")
    if _pdf_estimar_largura_texto(texto, tamanho_fonte) <= largura_maxima:
        return texto
    base = texto
    while base:
        candidato = f"{base}{sufixo}"
        if _pdf_estimar_largura_texto(candidato, tamanho_fonte) <= largura_maxima:
            return candidato
        base = base[:-1].rstrip()
    return sufixo


def _pdf_quebrar_palavra_longa(palavra, largura_maxima, tamanho_fonte):
    ponto_quebra = 1
    for indice in range(1, len(palavra) + 1):
        trecho = palavra[:indice]
        if _pdf_estimar_largura_texto(trecho, tamanho_fonte) > largura_maxima:
            break
        ponto_quebra = indice
    if ponto_quebra >= len(palavra):
        return palavra, ""
    if ponto_quebra > 3:
        return f"{palavra[:ponto_quebra - 1]}-", palavra[ponto_quebra - 1 :]
    return palavra[:ponto_quebra], palavra[ponto_quebra:]


def _pdf_text_commands(x, y, texto, *, fonte="F1", tamanho=10):
    return [
        "0 0 0 rg",
        "BT",
        f"/{fonte} {tamanho} Tf",
        f"{x:.2f} {y:.2f} Td",
        f"({_pdf_escape(texto)}) Tj",
        "ET",
    ]


def _pdf_text_commands_color(x, y, texto, *, fonte="F1", tamanho=10, rgb=(0, 0, 0)):
    r, g, b = rgb
    return [
        f"{r} {g} {b} rg",
        "BT",
        f"/{fonte} {tamanho} Tf",
        f"{x:.2f} {y:.2f} Td",
        f"({_pdf_escape(texto)}) Tj",
        "ET",
        "0 0 0 rg",
    ]


def _pdf_normalizar_colunas(colunas):
    colunas_normalizadas = []
    for coluna in colunas:
        if isinstance(coluna, dict):
            colunas_normalizadas.append(
                {
                    "chave": coluna.get("chave") or coluna.get("titulo") or "valor",
                    "titulo": coluna.get("titulo") or coluna.get("chave") or "Valor",
                    "largura": float(coluna.get("largura", 80)),
                    "align": coluna.get("align") or _pdf_inferir_alinhamento_coluna(coluna.get("titulo") or coluna.get("chave") or ""),
                }
            )
        else:
            titulo, largura = coluna
            colunas_normalizadas.append(
                {
                    "chave": titulo,
                    "titulo": titulo,
                    "largura": float(largura),
                    "align": _pdf_inferir_alinhamento_coluna(titulo),
                }
            )
    return colunas_normalizadas


def _pdf_inferir_alinhamento_coluna(titulo):
    titulo_normalizado = _sanear_texto_exportacao_seguro(titulo or "").lower()
    if any(chave in titulo_normalizado for chave in ["valor", "saldo", "total", "quantidade", "%", "percent", "nivel"]):
        return "right"
    if any(chave in titulo_normalizado for chave in ["data", "emissao", "vencimento", "validade"]):
        return "center"
    return "left"


def _pdf_x_texto_alinhado(x_cursor, largura, texto, tamanho_fonte, alinhamento, padding_x):
    largura_texto = _pdf_estimar_largura_texto(texto, tamanho_fonte)
    if alinhamento == "right":
        return max(x_cursor + padding_x, x_cursor + largura - padding_x - largura_texto)
    if alinhamento == "center":
        return max(x_cursor + padding_x, x_cursor + ((largura - largura_texto) / 2))
    return x_cursor + padding_x


def _pdf_valor_documento(valor, *, vazio="N\u00e3o informado"):
    if valor is None:
        return vazio
    if isinstance(valor, str):
        texto = _sanear_texto_exportacao_seguro(valor).strip()
        return texto if texto and texto not in {"-", "--", "- - - -"} else vazio
    return str(valor)


def _pdf_normalizar_linhas_documento(linhas, colunas):
    colunas_normalizadas = _pdf_normalizar_colunas(colunas)
    linhas_normalizadas = []
    for linha in list(linhas or []):
        linha_normalizada = {}
        for coluna in colunas_normalizadas:
            linha_normalizada[coluna["chave"]] = _pdf_valor_documento(linha.get(coluna["chave"]))
        linhas_normalizadas.append(linha_normalizada)
    if linhas_normalizadas:
        return linhas_normalizadas
    return [{colunas_normalizadas[0]["chave"]: "Nenhum registro encontrado"}]


def _pdf_obter_metadados_relatorio(titulo, resumo):
    resumo = resumo or {}
    return {
        "sistema": "CONSTRUTASK",
        "relatorio": _pdf_valor_documento(titulo, vazio="Relat\u00f3rio"),
        "codigo_documento": _pdf_valor_documento(
            resumo.get("Numero")
            or resumo.get("C\u00f3digo")
            or resumo.get("C\u00f3digo Documento")
            or resumo.get("Identificador")
            or resumo.get("Identificador da Evid\u00eancia")
        ),
        "obra": _pdf_valor_documento(resumo.get("Obra")),
        "data_emissao": _pdf_valor_documento(resumo.get("Emitido em") or _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")),
        "id_interno": _pdf_valor_documento(
            resumo.get("ID Interno") or resumo.get("Identificador da Evid\u00eancia"),
            vazio="N\u00e3o informado",
        ),
    }


def _pdf_titulo_limpo(titulo):
    texto = _sanear_texto_exportacao_seguro(titulo or "").strip()
    if " - EVD-" in texto:
        return texto.split(" - EVD-", 1)[0].strip()
    return texto


def _png_scanline_unfilter(filtro, linha_filtrada, anterior, bpp):
    resultado = bytearray(len(linha_filtrada))
    if filtro == 0:
        return bytearray(linha_filtrada)
    if filtro == 1:
        for indice, valor in enumerate(linha_filtrada):
            esquerda = resultado[indice - bpp] if indice >= bpp else 0
            resultado[indice] = (valor + esquerda) & 0xFF
        return resultado
    if filtro == 2:
        for indice, valor in enumerate(linha_filtrada):
            acima = anterior[indice] if anterior else 0
            resultado[indice] = (valor + acima) & 0xFF
        return resultado
    if filtro == 3:
        for indice, valor in enumerate(linha_filtrada):
            esquerda = resultado[indice - bpp] if indice >= bpp else 0
            acima = anterior[indice] if anterior else 0
            resultado[indice] = (valor + ((esquerda + acima) // 2)) & 0xFF
        return resultado
    if filtro == 4:
        def _paeth(a, b, c):
            p = a + b - c
            pa = abs(p - a)
            pb = abs(p - b)
            pc = abs(p - c)
            if pa <= pb and pa <= pc:
                return a
            if pb <= pc:
                return b
            return c

        for indice, valor in enumerate(linha_filtrada):
            esquerda = resultado[indice - bpp] if indice >= bpp else 0
            acima = anterior[indice] if anterior else 0
            diagonal = anterior[indice - bpp] if anterior and indice >= bpp else 0
            resultado[indice] = (valor + _paeth(esquerda, acima, diagonal)) & 0xFF
        return resultado
    raise ValueError("Filtro PNG nao suportado.")


def _carregar_png_para_pdf(caminho_logo):
    if not os.path.exists(caminho_logo):
        return None

    with open(caminho_logo, "rb") as arquivo_logo:
        dados = arquivo_logo.read()

    assinatura = b"\x89PNG\r\n\x1a\n"
    if not dados.startswith(assinatura):
        return None

    cursor = len(assinatura)
    largura = altura = None
    bit_depth = color_type = interlace = None
    idat_chunks = []

    while cursor < len(dados):
        tamanho = struct.unpack(">I", dados[cursor : cursor + 4])[0]
        cursor += 4
        tipo = dados[cursor : cursor + 4]
        cursor += 4
        payload = dados[cursor : cursor + tamanho]
        cursor += tamanho + 4

        if tipo == b"IHDR":
            largura, altura, bit_depth, color_type, _, _, interlace = struct.unpack(">IIBBBBB", payload)
        elif tipo == b"IDAT":
            idat_chunks.append(payload)
        elif tipo == b"IEND":
            break

    if not largura or not altura or bit_depth != 8 or interlace != 0 or color_type not in (2, 6):
        return None

    canais = 4 if color_type == 6 else 3
    bytes_por_pixel = canais
    bytes_por_linha = largura * canais
    dados_descomprimidos = zlib.decompress(b"".join(idat_chunks))

    cursor = 0
    linha_anterior = None
    rgb = bytearray()
    alpha_bytes = bytearray() if canais == 4 else None
    for _ in range(altura):
        filtro = dados_descomprimidos[cursor]
        cursor += 1
        linha_filtrada = dados_descomprimidos[cursor : cursor + bytes_por_linha]
        cursor += bytes_por_linha
        linha = _png_scanline_unfilter(filtro, linha_filtrada, linha_anterior, bytes_por_pixel)
        linha_anterior = linha

        if canais == 3:
            rgb.extend(linha)
            continue

        for indice in range(0, len(linha), 4):
            vermelho, verde, azul, alpha = linha[indice : indice + 4]
            rgb.extend((vermelho, verde, azul))
            alpha_bytes.extend((alpha,))

    return {
        "width": largura,
        "height": altura,
        "stream": zlib.compress(bytes(rgb)),
        "alpha_stream": zlib.compress(bytes(alpha_bytes)) if alpha_bytes is not None else None,
    }


def desenhar_cabecalho_pdf(y_topo, titulo, resumo, logo_pdf):
    metadados = _pdf_obter_metadados_relatorio(_pdf_titulo_limpo(titulo), resumo)
    x_caixa = 50
    largura_caixa = 495
    y_caixa = y_topo - 76
    altura_caixa = 72
    padding_x = 10
    padding_y = 8
    comandos = [
        "0.96 0.96 0.96 rg",
        "0 0 0 RG",
        f"{x_caixa} {y_caixa:.2f} {largura_caixa} {altura_caixa} re",
        "B",
    ]
    if logo_pdf:
        largura_logo = 125
        altura_logo = round((logo_pdf["height"] / logo_pdf["width"]) * largura_logo, 2)
        x_logo = x_caixa + largura_caixa - padding_x - largura_logo
        y_logo = y_caixa + altura_caixa - padding_y - altura_logo
        comandos.extend(
            [
                "q",
                f"{largura_logo} 0 0 {altura_logo} {x_logo:.2f} {y_logo:.2f} cm",
                "/Im1 Do",
                "Q",
            ]
        )
        largura_titulo = max(140, x_logo - (x_caixa + padding_x) - 14)
    else:
        comandos.extend(
            [
                "0.85 0.85 0.85 rg",
                "0 0 0 RG",
                f"{x_caixa + largura_caixa - padding_x - 125:.2f} {y_caixa + altura_caixa - padding_y - 28:.2f} 125 28 re",
                "B",
                *_pdf_text_commands(x_caixa + largura_caixa - padding_x - 105, y_caixa + altura_caixa - padding_y - 18, "LOGO", fonte="F2", tamanho=12),
            ]
        )
        largura_titulo = largura_caixa - (padding_x * 2) - 14
    titulo_cabecalho = _pdf_ajustar_texto_para_largura(metadados["relatorio"], largura_titulo, 11)
    comandos.extend(
        [
            *_pdf_text_commands(x_caixa + padding_x, y_caixa + altura_caixa - 18, titulo_cabecalho, fonte="F2", tamanho=11),
            *_pdf_text_commands(x_caixa + padding_x, y_caixa + altura_caixa - 34, f"C\u00f3digo: {metadados['codigo_documento']}", tamanho=8),
            *_pdf_text_commands(x_caixa + padding_x, y_caixa + altura_caixa - 50, f"Obra: {metadados['obra']}", tamanho=8),
            "0.75 0.75 0.75 RG",
            f"{x_caixa} {y_caixa - 4:.2f} m {x_caixa + largura_caixa} {y_caixa - 4:.2f} l S",
        ]
    )
    return comandos


def desenhar_rodape_pdf(numero_pagina, total_paginas, resumo, y_base=26):
    metadados = _pdf_obter_metadados_relatorio("", resumo)
    texto_geracao = f"Gerado por Construtask em {metadados['data_emissao']}"
    x_geracao = _pdf_x_texto_alinhado(225, 320, texto_geracao, 8, "right", 0)
    return [
        "0.75 0.75 0.75 RG",
        f"40 {y_base + 10:.2f} m 555 {y_base + 10:.2f} l S",
        *_pdf_text_commands(42, y_base, f"P\u00e1gina {numero_pagina} de {total_paginas}", tamanho=8),
        *_pdf_text_commands(x_geracao, y_base, texto_geracao, tamanho=8),
    ]


def _pdf_section_title_commands(y, titulo):
    titulo_bruto = _sanear_texto_exportacao_seguro(titulo or "")
    titulo_base = titulo_bruto.lower()
    if "histor" in titulo_base and "aprov" in titulo_base:
        titulo_normalizado = "HIST\u00d3RICO DE APROVA\u00c7\u00c3O"
    elif "histor" in titulo_base and "aditivo" in titulo_base:
        titulo_normalizado = "HIST\u00d3RICO DOS ADITIVOS"
    else:
        titulo_normalizado = titulo_bruto.upper()
    for termo_antigo, termo_novo in {
        "HISTRICO": "HIST\u00d3RICO",
        "APROVAO": "APROVA\u00c7\u00c3O",
        "DESCRIO": "DESCRI\u00c7\u00c3O",
        "MEDIO": "MEDI\u00c7\u00c3O",
        "CONTRATAES": "CONTRATA\u00c7\u00d5ES",
    }.items():
        titulo_normalizado = titulo_normalizado.replace(termo_antigo, termo_novo)
    return [
        "0.87 0.87 0.87 rg",
        "0 0 0 RG",
        f"50 {y - 20:.2f} 495 20 re",
        "B",
        *_pdf_text_commands_color(56, y - 14, titulo_normalizado, fonte="F2", tamanho=10, rgb=(0, 0, 0)),
        "0 0 0 rg",
    ]


def desenhar_titulo_secao(y, titulo):
    return _pdf_section_title_commands(y, titulo)


def desenhar_bloco_informacoes(y_topo, titulo, resumo):
    linhas = [{"Campo": campo, "Valor": _pdf_valor_documento(valor)} for campo, valor in (resumo or {}).items()]
    return _pdf_table_commands(
        y_topo,
        [
            {"chave": "Campo", "titulo": "Campo", "largura": 160, "align": "left"},
            {"chave": "Valor", "titulo": "Valor", "largura": 335, "align": "left"},
        ],
        linhas,
        titulo=titulo,
    )


def desenhar_tabela_padrao(y_topo, titulo, colunas, linhas, *, max_linhas=None):
    linhas_normalizadas = _pdf_normalizar_linhas_documento(linhas, colunas)
    return _pdf_table_commands(y_topo, colunas, linhas_normalizadas, titulo=titulo, max_linhas=max_linhas)


def _pdf_table_commands(y_topo, colunas, linhas, *, titulo=None, max_linhas=None):
    colunas = _pdf_normalizar_colunas(colunas)
    comandos = []
    y_atual = y_topo
    if titulo:
        comandos.extend(_pdf_section_title_commands(y_atual, titulo))
        y_atual -= 30

    if max_linhas is not None:
        linhas = list(linhas)[:max_linhas]

    altura_linha = 22
    padding_x = 5
    tamanho_fonte_cabecalho = 7
    tamanho_fonte_corpo = 7
    espacamento_linha = 9

    if not linhas:
        linhas = [{colunas[0]["chave"]: "-"}]

    header_wraps = []
    for coluna in colunas:
        header_wraps.append(
            _pdf_wrap_text(coluna["titulo"], coluna["largura"] - (padding_x * 2), tamanho_fonte_cabecalho)
        )
    header_height = max(20, 10 + (max(len(item) for item in header_wraps) * espacamento_linha))
    comandos.extend(
        [
            "0.90 0.90 0.90 rg",
            "0 0 0 RG",
            f"50 {y_atual - header_height:.2f} 495 {header_height} re",
            "B",
        ]
    )
    x_cursor = 50
    for indice_coluna, coluna in enumerate(colunas):
        largura = coluna["largura"]
        comandos.extend(
            [
                "0.90 0.90 0.90 rg",
                "0 0 0 RG",
                f"{x_cursor:.2f} {y_atual - header_height:.2f} {largura:.2f} {header_height} re",
                "B",
            ]
        )
        y_header = y_atual - 12
        for sublinha in header_wraps[indice_coluna]:
            comandos.extend(
                _pdf_text_commands_color(
                    _pdf_x_texto_alinhado(
                        x_cursor,
                        largura,
                        sublinha,
                        tamanho_fonte_cabecalho,
                        coluna.get("align", "left"),
                        padding_x,
                    ),
                    y_header,
                    sublinha,
                    fonte="F2",
                    tamanho=tamanho_fonte_cabecalho,
                    rgb=(0, 0, 0),
                )
            )
            y_header -= espacamento_linha
        x_cursor += largura
    y_atual -= header_height

    for linha in linhas:
        alturas = []
        conteudos = []
        for coluna in colunas:
            valor = _pdf_valor_documento(linha.get(coluna["chave"]))
            quebrado = _pdf_wrap_text(valor, coluna["largura"] - (padding_x * 2), tamanho_fonte_corpo)
            conteudos.append(quebrado)
            alturas.append(max(altura_linha, 10 + (len(quebrado) * espacamento_linha)))
        altura = max(alturas)
        x_cursor = 50
        for indice, coluna in enumerate(colunas):
            largura = coluna["largura"]
            comandos.append(f"{x_cursor:.2f} {y_atual - altura:.2f} {largura:.2f} {altura:.2f} re")
            comandos.append("S")
            y_texto = y_atual - 13
            for sublinha in conteudos[indice]:
                comandos.extend(
                    _pdf_text_commands(
                        _pdf_x_texto_alinhado(
                            x_cursor,
                            largura,
                            sublinha,
                            tamanho_fonte_corpo,
                            coluna.get("align", "left"),
                            padding_x,
                        ),
                        y_texto,
                        sublinha,
                        fonte="F1",
                        tamanho=tamanho_fonte_corpo,
                    )
                )
                y_texto -= espacamento_linha
            x_cursor += largura
        y_atual -= altura

    return comandos, y_atual


def _pdf_estimar_altura_tabela(colunas, linhas, *, titulo=None):
    colunas = _pdf_normalizar_colunas(colunas)
    padding_x = 5
    tamanho_fonte_cabecalho = 7
    tamanho_fonte_corpo = 7
    espacamento_linha = 9
    altura_total = 0
    if titulo:
        altura_total += 30
    header_wraps = [
        _pdf_wrap_text(coluna["titulo"], coluna["largura"] - (padding_x * 2), tamanho_fonte_cabecalho)
        for coluna in colunas
    ]
    altura_total += max(20, 10 + (max(len(item) for item in header_wraps) * espacamento_linha))
    for linha in linhas:
        alturas = []
        for coluna in colunas:
            quebrado = _pdf_wrap_text(
                _pdf_valor_documento(linha.get(coluna["chave"])),
                coluna["largura"] - (padding_x * 2),
                tamanho_fonte_corpo,
            )
            alturas.append(max(22, 10 + (len(quebrado) * espacamento_linha)))
        altura_total += max(alturas)
    return altura_total


def _pdf_relatorio_tabelas_response(nome_arquivo, titulo, resumo, secoes):
    logo_pdf = _carregar_png_para_pdf(_PDF_LOGO_PATH)
    altura_pagina = 842
    margem_superior = 40
    margem_inferior = 60
    y_inicial_pagina = altura_pagina - margem_superior - 102
    paginas = []

    def _novo_conteudo_pagina():
        comandos_pagina = desenhar_cabecalho_pdf(790, titulo, resumo, logo_pdf)
        return comandos_pagina, y_inicial_pagina

    def _fechar_pagina(comandos_pagina):
        paginas.append(comandos_pagina)

    conteudo, y_atual = _novo_conteudo_pagina()
    comandos_resumo, y_atual = desenhar_bloco_informacoes(y_atual, "RESUMO", resumo)
    conteudo.extend(comandos_resumo)

    for secao in secoes:
        colunas_secao = _pdf_normalizar_colunas(secao["colunas"])
        linhas = list(secao["linhas"] or [{colunas_secao[0]["chave"]: "-"}])
        max_linhas = secao.get("max_linhas")
        if max_linhas is not None:
            linhas = linhas[:max_linhas]
        primeiro_bloco = True
        while linhas:
            y_bloco = y_atual - 18
            titulo_secao = secao["titulo"] if primeiro_bloco else f'{secao["titulo"]} (continua)'
            linhas_bloco = []
            for linha in linhas:
                candidato = linhas_bloco + [linha]
                altura_prevista = _pdf_estimar_altura_tabela(colunas_secao, candidato, titulo=titulo_secao)
                if y_bloco - altura_prevista < margem_inferior:
                    break
                linhas_bloco = candidato
            if not linhas_bloco:
                _fechar_pagina(conteudo)
                conteudo, y_atual = _novo_conteudo_pagina()
                continue
            comandos_secao, y_atual = desenhar_tabela_padrao(y_bloco, titulo_secao, colunas_secao, linhas_bloco)
            conteudo.extend(comandos_secao)
            linhas = linhas[len(linhas_bloco) :]
            primeiro_bloco = False
            if linhas:
                _fechar_pagina(conteudo)
                conteudo, y_atual = _novo_conteudo_pagina()

    _fechar_pagina(conteudo)

    page_count = len(paginas)
    paginas_stream = []
    for indice_pagina, comandos_pagina in enumerate(paginas, start=1):
        comandos_pagina.extend(desenhar_rodape_pdf(indice_pagina, page_count, resumo))
        paginas_stream.append("\n".join(comandos_pagina).encode("cp1252", "replace"))
    font1_id = 3 + (page_count * 2)
    font2_id = font1_id + 1
    image_id = font2_id + 1 if logo_pdf else None
    alpha_image_id = image_id + 1 if logo_pdf and logo_pdf.get("alpha_stream") else None

    recursos_pagina = f"/Font << /F1 {font1_id} 0 R /F2 {font2_id} 0 R >>".encode("ascii")
    if image_id:
        recursos_pagina += f" /XObject << /Im1 {image_id} 0 R >>".encode("ascii")

    objetos = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        (
            b"2 0 obj << /Type /Pages /Kids ["
            + b" ".join(f"{3 + (indice * 2)} 0 R".encode("ascii") for indice in range(page_count))
            + b"] /Count "
            + str(page_count).encode("ascii")
            + b" >> endobj\n"
        ),
    ]
    for indice_pagina, stream in enumerate(paginas_stream):
        page_obj_id = 3 + (indice_pagina * 2)
        content_obj_id = page_obj_id + 1
        objetos.append(
            f"{page_obj_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << ".encode("ascii")
            + recursos_pagina
            + f" >> /Contents {content_obj_id} 0 R >> endobj\n".encode("ascii")
        )
        objetos.append(
            f"{content_obj_id} 0 obj << /Length {len(stream)} >> stream\n".encode("ascii")
            + stream
            + b"\nendstream endobj\n"
        )
    objetos.extend(
        [
            f"{font1_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >> endobj\n".encode("ascii"),
            f"{font2_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >> endobj\n".encode("ascii"),
        ]
    )
    if logo_pdf and image_id:
        objetos.append(
            f"{image_id} 0 obj << /Type /XObject /Subtype /Image /Width ".encode("ascii")
            + str(logo_pdf["width"]).encode("ascii")
            + b" /Height "
            + str(logo_pdf["height"]).encode("ascii")
            + b" /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length "
            + str(len(logo_pdf["stream"])).encode("ascii")
            + (f" /SMask {alpha_image_id} 0 R".encode("ascii") if alpha_image_id else b"")
            + b" >> stream\n"
            + logo_pdf["stream"]
            + b"\nendstream endobj\n"
        )
    if logo_pdf and alpha_image_id:
        objetos.append(
            f"{alpha_image_id} 0 obj << /Type /XObject /Subtype /Image /Width ".encode("ascii")
            + str(logo_pdf["width"]).encode("ascii")
            + b" /Height "
            + str(logo_pdf["height"]).encode("ascii")
            + b" /ColorSpace /DeviceGray /BitsPerComponent 8 /Filter /FlateDecode /Length "
            + str(len(logo_pdf["alpha_stream"])).encode("ascii")
            + b" >> stream\n"
            + logo_pdf["alpha_stream"]
            + b"\nendstream endobj\n"
        )

    pdf = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objetos:
        offsets.append(len(pdf))
        pdf += obj
    xref = len(pdf)
    pdf += f"xref\n0 {len(offsets)}\n".encode("ascii")
    pdf += b"0000000000 65535 f \n"
    for offset in offsets[1:]:
        pdf += f"{offset:010d} 00000 n \n".encode("ascii")
    pdf += f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode("ascii")

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _pdf_relatorio_probatorio_response(
    nome_arquivo,
    titulo,
    resumo,
    historico,
    extras,
    *,
    extras_titulo,
    extras_colunas,
    incluir_historico=True,
    extras_max_linhas=None,
    secoes_extras=None,
):
    secoes = []
    if incluir_historico:
        secoes.append(
            {
                "titulo": "Hist\u00f3rico de Aprova\u00e7\u00e3o",
                "colunas": [("Data", 75), ("A\u00e7\u00e3o", 75), ("Usu\u00e1rio", 105), ("Descri\u00e7\u00e3o", 240)],
                "linhas": [
                    {
                        "Data": linha.get("Data", "-"),
                        "A\u00e7\u00e3o": linha.get("Acao", linha.get("A\u00e7\u00e3o", "-")),
                        "Usu\u00e1rio": linha.get("Usuario", linha.get("Usu\u00e1rio", "-")),
                        "Descri\u00e7\u00e3o": linha.get("Descricao", linha.get("Descri\u00e7\u00e3o", "-")),
                    }
                    for linha in (historico or [{"Data": "-", "Acao": "-", "Usuario": "-", "Descricao": "-"}])
                ],
            }
        )
    secoes.append(
        {
            "titulo": extras_titulo,
            "colunas": extras_colunas,
            "linhas": extras or [{extras_colunas[0][0]: "-"}],
            **({"max_linhas": extras_max_linhas} if extras_max_linhas is not None else {}),
        }
    )
    for secao_extra in secoes_extras or []:
        secoes.append(secao_extra)
    return _pdf_relatorio_tabelas_response(nome_arquivo, titulo, resumo, secoes)


def _pdf_simples_response(nome_arquivo, titulo, linhas):
    resumo = []
    for linha in list(linhas):
        texto = str(linha)
        if ":" in texto:
            campo, valor = texto.split(":", 1)
            resumo.append({"Campo": campo.strip(), "Valor": valor.strip() or "-"})
        else:
            resumo.append({"Campo": "Informacao", "Valor": texto})

    return _pdf_relatorio_probatorio_response(
        nome_arquivo,
        titulo,
        {"Documento": titulo, "Emitido em": _datahora_local(timezone.now()).strftime("%d/%m/%Y %H:%M")},
        [],
        resumo,
        extras_titulo="Dados Exportados",
        extras_colunas=[("Campo", 165), ("Valor", 330)],
    )


def _exportar_relatorio_probatorio_excel_response(
    nome_arquivo,
    sheet_resumo,
    resumo,
    historico_linhas,
    *,
    extras_sheet_name=None,
    extras_linhas=None,
):
    output = BytesIO()
    resumo_linhas = _normalizar_linhas_exportacao([{"Campo": chave, "Valor": valor} for chave, valor in resumo.items()])
    historico_normalizado = _normalizar_linhas_exportacao(
        historico_linhas or [{"Data": "-", "Acao": "-", "Usuario": "-", "Descricao": "-"}]
    )
    extras_normalizados = _normalizar_linhas_exportacao(extras_linhas or [{"Informacao": "-"}])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(resumo_linhas).to_excel(writer, index=False, sheet_name=sheet_resumo)
        _aplicar_layout_excel_relatorio(writer.book[sheet_resumo], "RELATORIO PROBATORIO DE APROVACAO", sheet_resumo)
        pd.DataFrame(historico_normalizado).to_excel(writer, index=False, sheet_name="Historico")
        _aplicar_layout_excel_relatorio(writer.book["Historico"], "RELATORIO PROBATORIO DE APROVACAO", "Historico")
        if extras_sheet_name:
            pd.DataFrame(extras_normalizados).to_excel(writer, index=False, sheet_name=extras_sheet_name)
            _aplicar_layout_excel_relatorio(writer.book[extras_sheet_name], "RELATORIO PROBATORIO DE APROVACAO", extras_sheet_name)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response
